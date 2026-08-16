from flask import Blueprint, render_template, request, jsonify, session, send_file, flash
from flask_login import login_required, current_user
from database.models.company import Company
from database.models.employee import Employee
from database.models.customer import Customer
from database.models.supplier import Supplier
from database.models.inventory import Inventory
from database.models.branch import Branch
from database.models.department import Department
from database.models.financial import FinancialRecord, FinancialAccount
from database.models.custom_twin import CustomTwin, CustomTwinRecord

from services.analytics_service import AnalyticsService
from database.db import db
from sqlalchemy import func
from collections import Counter
import csv
import io
import os
from datetime import datetime, date

business_twin_bp = Blueprint('business_twin', __name__)


def _get_currency_symbol(company_id):
    """Get the currency symbol for a company."""
    from config.constants import CURRENCY_SYMBOLS
    company = Company.query.get(company_id)
    currency = company.currency if company and company.currency else 'USD'
    return CURRENCY_SYMBOLS.get(currency, '$')


def _sync_dept_headcount(company_id, dept_id):
    """Sync Department.employee_count to actual active employees."""
    if not dept_id:
        return
    count = Employee.query.filter_by(
        company_id=company_id,
        department_id=dept_id,
        status='active'
    ).count()
    dept = Department.query.filter_by(id=dept_id, company_id=company_id).first()
    if dept:
        dept.employee_count = count
        db.session.commit()


@business_twin_bp.route('/')
@login_required
def overview():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    summary = AnalyticsService.get_executive_summary(company_id)

    employees = Employee.query.filter_by(company_id=company_id, status='active').count()
    customers = Customer.query.filter_by(company_id=company_id, status='active').count()
    suppliers = Supplier.query.filter_by(company_id=company_id, status='active').count()
    inventory_items = Inventory.query.filter_by(company_id=company_id, is_active=True).count()
    branches = Branch.query.filter_by(company_id=company_id, status='active').count()
    departments = Department.query.filter_by(company_id=company_id, is_active=True).all()

    if company:
        if employees == 0 and company.employee_count:
            employees = company.employee_count
        if customers == 0:
            if hasattr(company, 'customer_count') and company.customer_count:
                customers = company.customer_count
            else:
                customers = AnalyticsService._estimate_customers(company)

    # Get custom twins for the company
    custom_twins = CustomTwin.query.filter_by(company_id=company_id, is_active=True).order_by(CustomTwin.name).all()

    return render_template('business_twin/twin_overview.html',
                         company=company, summary=summary,
                         employees=employees, customers=customers,
                         suppliers=suppliers, inventory_items=inventory_items,
                         branches=branches, departments=departments,
                         custom_twins=custom_twins)


@business_twin_bp.route('/departments')
@login_required
def departments():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    
    depts = Department.query.filter_by(company_id=company_id, is_active=True).all()
    
    dept_data = []
    total_headcount = 0
    total_payroll = 0
    total_budget = 0
    total_spent = 0
    
    for dept in depts:
        # Real-time employee aggregation
        emp_stats = db.session.query(
            func.count(Employee.id).label('headcount'),
            func.sum(Employee.salary).label('total_salary'),
            func.sum(Employee.benefits_cost).label('total_benefits'),
            func.avg(Employee.salary).label('avg_salary')
        ).filter(
            Employee.company_id == company_id,
            Employee.department_id == dept.id,
            Employee.status == 'active'
        ).first()
        
        headcount = emp_stats.headcount or 0
        dept_salary = float(emp_stats.total_salary or 0)
        dept_benefits = float(emp_stats.total_benefits or 0)
        dept_payroll = dept_salary + dept_benefits
        avg_salary = float(emp_stats.avg_salary or 0)
        
        # Keep stored count in sync
        if dept.employee_count != headcount:
            dept.employee_count = headcount
            db.session.commit()
        
        utilization = round((dept.spent / dept.budget * 100), 1) if dept.budget > 0 else 0
        
        # Employment type breakdown
        type_rows = db.session.query(
            Employee.employment_type,
            func.count(Employee.id).label('count')
        ).filter(
            Employee.company_id == company_id,
            Employee.department_id == dept.id,
            Employee.status == 'active'
        ).group_by(Employee.employment_type).all()
        type_breakdown = {t: c for t, c in type_rows}
        
        # Top skills
        skills_list = []
        for emp in Employee.query.filter_by(company_id=company_id, department_id=dept.id, status='active').all():
            if emp.skills:
                skills_list.extend([s.strip() for s in emp.skills.split(',') if s.strip()])
        top_skills = [s[0] for s in Counter(skills_list).most_common(3)]
        
        dept_data.append({
            'id': dept.id,
            'name': dept.name,
            'code': dept.code,
            'color': dept.color or '#00D4FF',
            'head_name': dept.head_name,
            'headcount': headcount,
            'total_salary': dept_salary,
            'total_benefits': dept_benefits,
            'total_payroll': dept_payroll,
            'avg_salary': avg_salary,
            'budget': dept.budget,
            'spent': dept.spent,
            'remaining': max(0, dept.budget - dept.spent),
            'utilization': utilization,
            'type_breakdown': type_breakdown,
            'top_skills': top_skills,
            'is_active': dept.is_active
        })
        
        total_headcount += headcount
        total_payroll += dept_payroll
        total_budget += dept.budget
        total_spent += dept.spent
    
    company_avg_salary = db.session.query(func.avg(Employee.salary)).filter(
        Employee.company_id == company_id,
        Employee.status == 'active'
    ).scalar() or 0
    
    dept_data.sort(key=lambda x: x['headcount'], reverse=True)
    
    return render_template('business_twin/twin_departments.html',
                         company=company,
                         departments=dept_data,
                         total_headcount=total_headcount,
                         total_payroll=total_payroll,
                         total_budget=total_budget,
                         total_spent=total_spent,
                         avg_salary=float(company_avg_salary),
                         dept_count=len(depts))


@business_twin_bp.route('/departments/add', methods=['POST'])
@login_required
def add_department():
    company_id = session.get('company_id')
    try:
        name = request.form.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'Department name is required'}), 400
        
        code = request.form.get('code', name[:3].upper()).strip()
        dept = Department(
            company_id=company_id,
            name=name,
            code=code,
            description=request.form.get('description', '').strip(),
            head_name=request.form.get('head_name', '').strip(),
            budget=float(request.form.get('budget', 0) or 0),
            color=request.form.get('color', '#00D4FF'),
            is_active=True
        )
        db.session.add(dept)
        db.session.commit()
        return jsonify({'success': True, 'department': dept.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# REVENUE TWIN — ENHANCED ROUTES
# =============================================================================

@business_twin_bp.route('/revenue')
@login_required
def revenue():
    company_id = session.get('company_id')
    summary = AnalyticsService.get_executive_summary(company_id)
    trend = AnalyticsService.get_trend_data(company_id, 180)
    
    records = FinancialRecord.query.filter_by(
        company_id=company_id, transaction_type='revenue'
    ).order_by(FinancialRecord.transaction_date.desc()).all()
    
    departments = Department.query.filter_by(company_id=company_id, is_active=True).all()
    accounts = FinancialAccount.query.filter_by(company_id=company_id, is_active=True).all()
    
    return render_template('business_twin/twin_revenue.html',
                         summary=summary, trend=trend,
                         records=[r.to_dict() for r in records],
                         departments=departments,
                         accounts=accounts,
                         currency_symbol=_get_currency_symbol(company_id))


@business_twin_bp.route('/revenue/add', methods=['POST'])
@login_required
def add_revenue():
    company_id = session.get('company_id')
    try:
        amount = float(request.form.get('amount', 0))
        if amount <= 0:
            return jsonify({'success': False, 'error': 'Amount must be greater than 0'}), 400
        
        tx_date_str = request.form.get('transaction_date')
        if tx_date_str:
            tx_date = datetime.strptime(tx_date_str, '%Y-%m-%d').date()
        else:
            tx_date = date.today()
        
        dept_id = request.form.get('department_id')
        account_id = request.form.get('account_id')
        
        record = FinancialRecord(
            company_id=company_id,
            account_id=int(account_id) if account_id and account_id.strip() else None,
            transaction_date=tx_date,
            transaction_type='revenue',
            category=request.form.get('category', 'Revenue'),
            subcategory=request.form.get('subcategory', ''),
            amount=amount,
            description=request.form.get('description', ''),
            reference_number=request.form.get('reference_number', ''),
            department_id=int(dept_id) if dept_id and dept_id.strip() else None,
            is_recurring=request.form.get('is_recurring') == 'on',
            recurring_frequency=request.form.get('recurring_frequency', ''),
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        db.session.add(record)
        db.session.commit()
        
        symbol = _get_currency_symbol(company_id)
        return jsonify({
            'success': True,
            'message': f'Revenue of {symbol}{amount:,.2f} added successfully',
            'record': record.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/revenue/edit/<int:record_id>', methods=['POST'])
@login_required
def edit_revenue(record_id):
    company_id = session.get('company_id')
    record = FinancialRecord.query.filter_by(
        id=record_id, company_id=company_id, transaction_type='revenue'
    ).first()
    if not record:
        return jsonify({'success': False, 'error': 'Revenue record not found'}), 404
    
    try:
        amount = request.form.get('amount')
        if amount is not None and amount.strip():
            record.amount = float(amount)
        
        tx_date_str = request.form.get('transaction_date')
        if tx_date_str and tx_date_str.strip():
            record.transaction_date = datetime.strptime(tx_date_str, '%Y-%m-%d').date()
        
        record.category = request.form.get('category', record.category or 'Revenue')
        record.subcategory = request.form.get('subcategory', record.subcategory or '')
        record.description = request.form.get('description', record.description or '')
        record.reference_number = request.form.get('reference_number', record.reference_number or '')
        
        dept_id = request.form.get('department_id')
        if dept_id is not None:
            record.department_id = int(dept_id) if dept_id.strip() else None
        
        account_id = request.form.get('account_id')
        if account_id is not None:
            record.account_id = int(account_id) if account_id.strip() else None
        
        record.is_recurring = request.form.get('is_recurring') == 'on'
        record.recurring_frequency = request.form.get('recurring_frequency', record.recurring_frequency or '')
        
        db.session.commit()
        return jsonify({'success': True, 'record': record.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/revenue/delete/<int:record_id>', methods=['DELETE'])
@login_required
def delete_revenue(record_id):
    company_id = session.get('company_id')
    record = FinancialRecord.query.filter_by(
        id=record_id, company_id=company_id, transaction_type='revenue'
    ).first()
    if not record:
        return jsonify({'success': False, 'error': 'Revenue record not found'}), 404
    
    try:
        db.session.delete(record)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Revenue record deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/revenue/delete-all', methods=['DELETE'])
@login_required
def delete_all_revenue():
    company_id = session.get('company_id')
    try:
        count = FinancialRecord.query.filter_by(
            company_id=company_id, transaction_type='revenue'
        ).count()
        FinancialRecord.query.filter_by(
            company_id=company_id, transaction_type='revenue'
        ).delete()
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'All {count} revenue records deleted successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/revenue/import-csv', methods=['POST'])
@login_required
def import_revenue_csv():
    company_id = session.get('company_id')

    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)

        added_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                amount = float(row.get('amount', 0) or 0)
                if amount <= 0:
                    errors.append(f"Row {row_num}: Invalid amount")
                    continue

                tx_date = date.today()
                date_str = row.get('transaction_date', '').strip()
                if date_str:
                    try:
                        tx_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    except:
                        try:
                            tx_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass

                ref = row.get('reference_number', '').strip()
                
                # Upsert logic: match by reference_number + transaction_date
                existing = None
                if ref:
                    existing = FinancialRecord.query.filter_by(
                        company_id=company_id,
                        transaction_type='revenue',
                        reference_number=ref,
                        transaction_date=tx_date
                    ).first()

                dept_id = None
                dept_name = row.get('department', '').strip()
                if dept_name:
                    dept = Department.query.filter_by(company_id=company_id, name=dept_name).first()
                    if dept:
                        dept_id = dept.id

                account_id = None
                account_name = row.get('account', '').strip()
                if account_name:
                    acc = FinancialAccount.query.filter_by(company_id=company_id, account_name=account_name).first()
                    if acc:
                        account_id = acc.id

                if existing:
                    existing.category = row.get('category', existing.category or 'Revenue').strip() or existing.category
                    existing.subcategory = row.get('subcategory', existing.subcategory or '').strip()
                    existing.amount = amount
                    existing.description = row.get('description', existing.description or '').strip()
                    existing.department_id = dept_id or existing.department_id
                    existing.account_id = account_id or existing.account_id
                    existing.is_recurring = row.get('is_recurring', '').strip().lower() in ['yes', 'true', '1', 'on']
                    existing.recurring_frequency = row.get('recurring_frequency', existing.recurring_frequency or '').strip()
                    updated_count += 1
                else:
                    record = FinancialRecord(
                        company_id=company_id,
                        account_id=account_id,
                        transaction_date=tx_date,
                        transaction_type='revenue',
                        category=row.get('category', 'Revenue').strip() or 'Revenue',
                        subcategory=row.get('subcategory', '').strip(),
                        amount=amount,
                        description=row.get('description', '').strip(),
                        reference_number=ref,
                        department_id=dept_id,
                        is_recurring=row.get('is_recurring', '').strip().lower() in ['yes', 'true', '1', 'on'],
                        recurring_frequency=row.get('recurring_frequency', '').strip()
                    )
                    db.session.add(record)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        return jsonify({
            'success': True,
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/revenue/import-excel', methods=['POST'])
@login_required
def import_revenue_excel():
    company_id = session.get('company_id')

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active

        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_').replace('-', '_')
                headers.append(h)
            else:
                headers.append('')

        header_map = {}
        for idx, h in enumerate(headers):
            if h in ['transaction_date', 'date', 'revenue_date', 'tx_date']:
                header_map['transaction_date'] = idx
            elif h in ['category', 'revenue_category', 'type']:
                header_map['category'] = idx
            elif h in ['subcategory', 'sub_category', 'item']:
                header_map['subcategory'] = idx
            elif h in ['amount', 'value', 'total', 'revenue']:
                header_map['amount'] = idx
            elif h in ['description', 'desc', 'note', 'details']:
                header_map['description'] = idx
            elif h in ['reference_number', 'ref_no', 'ref_number', 'invoice_no']:
                header_map['reference_number'] = idx
            elif h in ['department', 'dept', 'department_name']:
                header_map['department'] = idx
            elif h in ['account', 'account_name', 'financial_account']:
                header_map['account'] = idx
            elif h in ['is_recurring', 'recurring', 'repeat']:
                header_map['is_recurring'] = idx
            elif h in ['recurring_frequency', 'frequency', 'interval']:
                header_map['recurring_frequency'] = idx

        if 'amount' not in header_map:
            return jsonify({'success': False, 'error': 'Excel file must have an amount column'}), 400

        added_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                amount_idx = header_map.get('amount', 0)
                amount_val = row[amount_idx] if amount_idx < len(row) else None

                if amount_val is None:
                    continue

                try:
                    amount = float(amount_val)
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Invalid amount value")
                    continue

                if amount <= 0:
                    errors.append(f"Row {row_num}: Amount must be greater than 0")
                    continue

                tx_date = date.today()
                date_idx = header_map.get('transaction_date')
                if date_idx is not None and date_idx < len(row) and row[date_idx]:
                    date_val = row[date_idx]
                    if isinstance(date_val, datetime):
                        tx_date = date_val.date()
                    elif isinstance(date_val, date):
                        tx_date = date_val
                    else:
                        date_str = str(date_val).strip()
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                            try:
                                tx_date = datetime.strptime(date_str, fmt).date()
                                break
                            except:
                                continue

                ref = ''
                ref_idx = header_map.get('reference_number')
                if ref_idx is not None and ref_idx < len(row) and row[ref_idx] is not None:
                    ref = str(row[ref_idx]).strip()

                # Upsert logic
                existing = None
                if ref:
                    existing = FinancialRecord.query.filter_by(
                        company_id=company_id,
                        transaction_type='revenue',
                        reference_number=ref,
                        transaction_date=tx_date
                    ).first()

                dept_id = None
                dept_idx = header_map.get('department')
                if dept_idx is not None and dept_idx < len(row) and row[dept_idx]:
                    dept_name = str(row[dept_idx]).strip()
                    dept = Department.query.filter_by(company_id=company_id, name=dept_name).first()
                    if dept:
                        dept_id = dept.id

                account_id = None
                account_idx = header_map.get('account')
                if account_idx is not None and account_idx < len(row) and row[account_idx]:
                    account_name = str(row[account_idx]).strip()
                    acc = FinancialAccount.query.filter_by(company_id=company_id, account_name=account_name).first()
                    if acc:
                        account_id = acc.id

                def get_str_val(key, default=''):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return default

                def get_bool_val(key):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        val = str(row[idx]).lower().strip()
                        return val in ['yes', 'true', '1', 'on', 'y']
                    return False

                if existing:
                    existing.category = get_str_val('category', existing.category or 'Revenue') or existing.category
                    existing.subcategory = get_str_val('subcategory', existing.subcategory or '')
                    existing.amount = amount
                    existing.description = get_str_val('description', existing.description or '')
                    existing.department_id = dept_id or existing.department_id
                    existing.account_id = account_id or existing.account_id
                    existing.is_recurring = get_bool_val('is_recurring')
                    existing.recurring_frequency = get_str_val('recurring_frequency', existing.recurring_frequency or '')
                    updated_count += 1
                else:
                    record = FinancialRecord(
                        company_id=company_id,
                        account_id=account_id,
                        transaction_date=tx_date,
                        transaction_type='revenue',
                        category=get_str_val('category', 'Revenue') or 'Revenue',
                        subcategory=get_str_val('subcategory'),
                        amount=amount,
                        description=get_str_val('description'),
                        reference_number=ref,
                        department_id=dept_id,
                        is_recurring=get_bool_val('is_recurring'),
                        recurring_frequency=get_str_val('recurring_frequency')
                    )
                    db.session.add(record)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        return jsonify({
            'success': True,
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/revenue/download-template')
@login_required
def download_revenue_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'transaction_date', 'category', 'subcategory', 'amount', 'description',
        'reference_number', 'department', 'account', 'is_recurring', 'recurring_frequency'
    ])

    sample_data = [
        ['2026-06-15', 'Revenue', 'Product Sales', 12500, 'Q2 software licenses', 'INV-001', 'Sales', 'Main Account', 'No', ''],
        ['2026-06-14', 'Revenue', 'Services', 45000, 'Consulting retainer', 'INV-002', 'Operations', 'Main Account', 'Yes', 'monthly'],
        ['2026-06-13', 'Revenue', 'Subscription', 28000, 'Annual SaaS renewal', 'INV-003', 'IT', 'Savings', 'No', ''],
        ['2026-06-12', 'Revenue', 'Licensing', 8500, 'Patent license fee', 'INV-004', 'Finance', 'Main Account', 'Yes', 'yearly'],
        ['2026-06-11', 'Revenue', 'Product Sales', 15000, 'Hardware shipment', 'INV-005', 'Sales', 'Main Account', 'No', ''],
        ['2026-06-10', 'Revenue', 'Services', 32000, 'Managed support contract', 'INV-006', 'IT', 'Main Account', 'Yes', 'monthly']
    ]

    writer.writerows(sample_data)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='revenue_import_template.csv')


@business_twin_bp.route('/revenue/download-excel-template')
@login_required
def download_revenue_excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Revenue"

    headers = [
        'transaction_date', 'category', 'subcategory', 'amount', 'description',
        'reference_number', 'department', 'account', 'is_recurring', 'recurring_frequency'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    sample_data = [
        ['2026-06-15', 'Revenue', 'Product Sales', 12500, 'Q2 software licenses', 'INV-001', 'Sales', 'Main Account', 'No', ''],
        ['2026-06-14', 'Revenue', 'Services', 45000, 'Consulting retainer', 'INV-002', 'Operations', 'Main Account', 'Yes', 'monthly'],
        ['2026-06-13', 'Revenue', 'Subscription', 28000, 'Annual SaaS renewal', 'INV-003', 'IT', 'Savings', 'No', ''],
        ['2026-06-12', 'Revenue', 'Licensing', 8500, 'Patent license fee', 'INV-004', 'Finance', 'Main Account', 'Yes', 'yearly'],
        ['2026-06-11', 'Revenue', 'Product Sales', 15000, 'Hardware shipment', 'INV-005', 'Sales', 'Main Account', 'No', ''],
        ['2026-06-10', 'Revenue', 'Services', 32000, 'Managed support contract', 'INV-006', 'IT', 'Main Account', 'Yes', 'monthly']
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='revenue_import_template.xlsx')


# =============================================================================
# EXPENSES
# =============================================================================

@business_twin_bp.route('/expenses')
@login_required
def expenses():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)

    departments = Department.query.filter_by(company_id=company_id, is_active=True).all()

    if not departments:
        default_depts = [
            {'name': 'Sales', 'code': 'SAL', 'color': '#00D4FF', 'budget': 150000},
            {'name': 'Marketing', 'code': 'MKT', 'color': '#8B5CF6', 'budget': 120000},
            {'name': 'Operations', 'code': 'OPS', 'color': '#F59E0B', 'budget': 200000},
            {'name': 'Finance', 'code': 'FIN', 'color': '#10B981', 'budget': 80000},
            {'name': 'HR', 'code': 'HRM', 'color': '#F472B6', 'budget': 60000},
            {'name': 'IT', 'code': 'ITD', 'color': '#EF4444', 'budget': 180000}
        ]
        for d in default_depts:
            dept = Department(company_id=company_id, **d)
            db.session.add(dept)
        db.session.commit()
        departments = Department.query.filter_by(company_id=company_id, is_active=True).all()

    dept_perf = []
    total_budget = 0
    total_spent = 0

    for dept in departments:
        actual_spent = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.department_id == dept.id,
            FinancialRecord.transaction_type == 'expense'
        ).scalar() or 0.0

        dept.spent = float(actual_spent)
        utilization = round((dept.spent / dept.budget * 100), 2) if dept.budget > 0 else 0
        remaining = dept.budget - dept.spent

        dept_perf.append({
            'id': dept.id,
            'name': dept.name,
            'code': dept.code,
            'budget': dept.budget,
            'spent': dept.spent,
            'remaining': round(remaining, 2),
            'utilization': utilization,
            'employee_count': dept.employee_count,
            'color': dept.color or '#00D4FF',
            'variance': round(remaining, 2)
        })

        total_budget += dept.budget
        total_spent += dept.spent

    db.session.commit()

    recent_expenses = FinancialRecord.query.filter_by(
        company_id=company_id, transaction_type='expense'
    ).order_by(FinancialRecord.transaction_date.desc()).limit(50).all()

    category_breakdown = db.session.query(
        FinancialRecord.category,
        func.sum(FinancialRecord.amount).label('total')
    ).filter(
        FinancialRecord.company_id == company_id,
        FinancialRecord.transaction_type == 'expense'
    ).group_by(FinancialRecord.category).all()

    monthly_trend = db.session.query(
        func.strftime('%Y-%m', FinancialRecord.transaction_date).label('month'),
        func.sum(FinancialRecord.amount).label('total')
    ).filter(
        FinancialRecord.company_id == company_id,
        FinancialRecord.transaction_type == 'expense'
    ).group_by('month').order_by('month').all()

    summary = {
        'total_budget': total_budget,
        'total_spent': total_spent,
        'total_remaining': total_budget - total_spent,
        'overall_utilization': round((total_spent / total_budget * 100), 2) if total_budget > 0 else 0
    }

    return render_template('business_twin/twin_expenses.html', 
                         summary=summary, 
                         dept_perf=dept_perf,
                         recent_expenses=recent_expenses,
                         recent_expenses_json=[e.to_dict() for e in recent_expenses],
                         category_breakdown=category_breakdown,
                         monthly_trend=monthly_trend,
                         departments=departments)


@business_twin_bp.route('/expenses/add', methods=['POST'])
@login_required
def add_expense():
    company_id = session.get('company_id')

    try:
        dept_id = request.form.get('department_id')
        amount = float(request.form.get('amount', 0))

        if amount <= 0:
            return jsonify({'success': False, 'error': 'Amount must be greater than 0'}), 400

        tx_date_str = request.form.get('transaction_date')
        if tx_date_str:
            tx_date = datetime.strptime(tx_date_str, '%Y-%m-%d').date()
        else:
            tx_date = date.today()

        expense = FinancialRecord(
            company_id=company_id,
            department_id=int(dept_id) if dept_id else None,
            transaction_date=tx_date,
            transaction_type='expense',
            category=request.form.get('category', 'Operating Expenses'),
            subcategory=request.form.get('subcategory', 'General'),
            amount=amount,
            description=request.form.get('description', ''),
            reference_number=request.form.get('reference_number', ''),
            is_recurring=request.form.get('is_recurring') == 'on',
            recurring_frequency=request.form.get('recurring_frequency', ''),
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )

        db.session.add(expense)

        if dept_id:
            dept = Department.query.get(int(dept_id))
            if dept:
                dept.spent = (dept.spent or 0) + amount

        db.session.commit()

        symbol = _get_currency_symbol(company_id)

        return jsonify({
            'success': True, 
            'message': f'Expense of {symbol}{amount:,.2f} added successfully',
            'expense': expense.to_dict()
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/expenses/edit/<int:expense_id>', methods=['POST'])
@login_required
def edit_expense(expense_id):
    company_id = session.get('company_id')
    expense = FinancialRecord.query.filter_by(
        id=expense_id, company_id=company_id, transaction_type='expense'
    ).first()
    if not expense:
        return jsonify({'success': False, 'error': 'Expense not found'}), 404

    try:
        old_dept_id = expense.department_id
        old_amount = expense.amount

        amount = request.form.get('amount')
        if amount is not None and amount.strip():
            expense.amount = float(amount)

        tx_date_str = request.form.get('transaction_date')
        if tx_date_str and tx_date_str.strip():
            expense.transaction_date = datetime.strptime(tx_date_str, '%Y-%m-%d').date()

        expense.category = request.form.get('category', expense.category or 'Operating Expenses')
        expense.subcategory = request.form.get('subcategory', expense.subcategory or '')
        expense.description = request.form.get('description', expense.description or '')
        expense.reference_number = request.form.get('reference_number', expense.reference_number or '')

        dept_id = request.form.get('department_id')
        if dept_id is not None:
            expense.department_id = int(dept_id) if dept_id.strip() else None

        expense.is_recurring = request.form.get('is_recurring') == 'on'
        expense.recurring_frequency = request.form.get('recurring_frequency', expense.recurring_frequency or '')

        new_amount = expense.amount
        new_dept_id = expense.department_id

        # Adjust department spending
        if old_dept_id and old_dept_id == new_dept_id:
            dept = Department.query.get(old_dept_id)
            if dept:
                dept.spent = (dept.spent or 0) + (new_amount - old_amount)
        elif old_dept_id and new_dept_id and old_dept_id != new_dept_id:
            old_dept = Department.query.get(old_dept_id)
            if old_dept:
                old_dept.spent = max(0, (old_dept.spent or 0) - old_amount)
            new_dept = Department.query.get(new_dept_id)
            if new_dept:
                new_dept.spent = (new_dept.spent or 0) + new_amount
        elif not old_dept_id and new_dept_id:
            new_dept = Department.query.get(new_dept_id)
            if new_dept:
                new_dept.spent = (new_dept.spent or 0) + new_amount
        elif old_dept_id and not new_dept_id:
            old_dept = Department.query.get(old_dept_id)
            if old_dept:
                old_dept.spent = max(0, (old_dept.spent or 0) - old_amount)

        db.session.commit()
        return jsonify({'success': True, 'expense': expense.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/expenses/delete/<int:expense_id>', methods=['DELETE'])
@login_required
def delete_expense(expense_id):
    company_id = session.get('company_id')

    expense = FinancialRecord.query.filter_by(id=expense_id, company_id=company_id).first()
    if not expense:
        return jsonify({'success': False, 'error': 'Expense not found'}), 404

    try:
        if expense.department_id:
            dept = Department.query.get(expense.department_id)
            if dept and dept.spent:
                dept.spent = max(0, dept.spent - expense.amount)

        db.session.delete(expense)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Expense deleted successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/expenses/delete-all', methods=['DELETE'])
@login_required
def delete_all_expenses():
    company_id = session.get('company_id')
    try:
        count = FinancialRecord.query.filter_by(
            company_id=company_id, transaction_type='expense'
        ).count()
        FinancialRecord.query.filter_by(
            company_id=company_id, transaction_type='expense'
        ).delete()

        # Reset all department spent to 0
        departments = Department.query.filter_by(company_id=company_id).all()
        for dept in departments:
            dept.spent = 0

        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'All {count} expense records deleted successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/expenses/import-csv', methods=['POST'])
@login_required
def import_expenses_csv():
    company_id = session.get('company_id')

    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)

        added_count = 0
        errors = []
        affected_dept_ids = set()

        depts = {d.name.lower(): d for d in Department.query.filter_by(company_id=company_id).all()}

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                dept_name = row.get('department', '').strip()
                dept = depts.get(dept_name.lower())

                if not dept and dept_name:
                    for d_name, d_obj in depts.items():
                        if dept_name.lower() in d_name or d_name in dept_name.lower():
                            dept = d_obj
                            break

                amount = float(row.get('amount', 0) or 0)
                if amount <= 0:
                    errors.append(f"Row {row_num}: Invalid amount")
                    continue

                tx_date = date.today()
                date_str = row.get('transaction_date', '').strip()
                if date_str:
                    try:
                        tx_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    except:
                        try:
                            tx_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass

                expense = FinancialRecord(
                    company_id=company_id,
                    department_id=dept.id if dept else None,
                    transaction_date=tx_date,
                    transaction_type='expense',
                    category=row.get('category', 'Operating Expenses').strip() or 'Operating Expenses',
                    subcategory=row.get('subcategory', 'General').strip() or 'General',
                    amount=amount,
                    description=row.get('description', '').strip(),
                    reference_number=row.get('reference_number', '').strip(),
                    is_recurring=row.get('is_recurring', '').lower() in ['yes', 'true', '1', 'on'],
                    recurring_frequency=row.get('recurring_frequency', '').strip()
                )
                db.session.add(expense)

                if dept:
                    dept.spent = (dept.spent or 0) + amount
                    affected_dept_ids.add(dept.id)

                added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        
        for dept_id in affected_dept_ids:
            _sync_dept_headcount(company_id, dept_id)

        return jsonify({
            'success': True, 
            'added': added_count,
            'errors': errors,
            'message': f'Successfully imported {added_count} expenses'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/expenses/import-excel', methods=['POST'])
@login_required
def import_expenses_excel():
    company_id = session.get('company_id')

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active

        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_').replace('-', '_')
                headers.append(h)
            else:
                headers.append('')

        header_map = {}
        for idx, h in enumerate(headers):
            if h in ['department', 'dept', 'department_name', 'dept_name']:
                header_map['department'] = idx
            elif h in ['category', 'expense_category', 'type']:
                header_map['category'] = idx
            elif h in ['subcategory', 'sub_category', 'item']:
                header_map['subcategory'] = idx
            elif h in ['amount', 'cost', 'price', 'value', 'total']:
                header_map['amount'] = idx
            elif h in ['description', 'desc', 'note', 'details']:
                header_map['description'] = idx
            elif h in ['transaction_date', 'date', 'expense_date', 'tx_date']:
                header_map['transaction_date'] = idx
            elif h in ['reference_number', 'ref_no', 'ref_number', 'invoice_no', 'receipt']:
                header_map['reference_number'] = idx
            elif h in ['is_recurring', 'recurring', 'repeat']:
                header_map['is_recurring'] = idx
            elif h in ['recurring_frequency', 'frequency', 'interval']:
                header_map['recurring_frequency'] = idx

        if 'amount' not in header_map:
            return jsonify({'success': False, 'error': 'Excel file must have an amount column'}), 400

        depts = {d.name.lower(): d for d in Department.query.filter_by(company_id=company_id).all()}

        added_count = 0
        errors = []
        affected_dept_ids = set()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                amount_idx = header_map.get('amount', 0)
                amount_val = row[amount_idx] if amount_idx < len(row) else None

                if amount_val is None:
                    continue

                try:
                    amount = float(amount_val)
                except (ValueError, TypeError):
                    errors.append(f"Row {row_num}: Invalid amount value")
                    continue

                if amount <= 0:
                    errors.append(f"Row {row_num}: Amount must be greater than 0")
                    continue

                dept = None
                dept_idx = header_map.get('department')
                if dept_idx is not None and dept_idx < len(row) and row[dept_idx]:
                    dept_name = str(row[dept_idx]).strip()
                    dept = depts.get(dept_name.lower())
                    if not dept:
                        for d_name, d_obj in depts.items():
                            if dept_name.lower() in d_name or d_name in dept_name.lower():
                                dept = d_obj
                                break

                tx_date = date.today()
                date_idx = header_map.get('transaction_date')
                if date_idx is not None and date_idx < len(row) and row[date_idx]:
                    date_val = row[date_idx]
                    if isinstance(date_val, datetime):
                        tx_date = date_val.date()
                    elif isinstance(date_val, date):
                        tx_date = date_val
                    else:
                        date_str = str(date_val).strip()
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                            try:
                                tx_date = datetime.strptime(date_str, fmt).date()
                                break
                            except:
                                continue

                def get_str_val(key, default=''):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        return str(row[idx]).strip()
                    return default

                def get_bool_val(key):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row) and row[idx] is not None:
                        val = str(row[idx]).lower().strip()
                        return val in ['yes', 'true', '1', 'on', 'y']
                    return False

                expense = FinancialRecord(
                    company_id=company_id,
                    department_id=dept.id if dept else None,
                    transaction_date=tx_date,
                    transaction_type='expense',
                    category=get_str_val('category', 'Operating Expenses'),
                    subcategory=get_str_val('subcategory', 'General'),
                    amount=amount,
                    description=get_str_val('description'),
                    reference_number=get_str_val('reference_number'),
                    is_recurring=get_bool_val('is_recurring'),
                    recurring_frequency=get_str_val('recurring_frequency')
                )
                db.session.add(expense)

                if dept:
                    dept.spent = (dept.spent or 0) + amount
                    affected_dept_ids.add(dept.id)

                added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        
        for dept_id in affected_dept_ids:
            _sync_dept_headcount(company_id, dept_id)

        return jsonify({
            'success': True, 
            'added': added_count,
            'errors': errors,
            'message': f'Successfully imported {added_count} expenses from Excel'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/expenses/download-template')
@login_required
def download_expense_template():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'department', 'category', 'subcategory', 'amount', 'description',
        'transaction_date', 'reference_number', 'is_recurring', 'recurring_frequency'
    ])

    sample_data = [
        ['Sales', 'Operating Expenses', 'Travel & Entertainment', 12500, 'Q2 client meetings', '2026-06-15', 'EXP-001', 'No', ''],
        ['Marketing', 'Operating Expenses', 'Digital Advertising', 45000, 'Google Ads campaign', '2026-06-14', 'EXP-002', 'Yes', 'monthly'],
        ['Operations', 'Operating Expenses', 'Logistics', 28000, 'Warehouse freight', '2026-06-13', 'EXP-003', 'No', ''],
        ['Finance', 'Operating Expenses', 'Software', 8500, 'Accounting license', '2026-06-12', 'EXP-004', 'Yes', 'yearly'],
        ['HR', 'Operating Expenses', 'Recruitment', 15000, 'Agency fees', '2026-06-11', 'EXP-005', 'No', ''],
        ['IT', 'Operating Expenses', 'Cloud', 32000, 'AWS services', '2026-06-10', 'EXP-006', 'Yes', 'monthly']
    ]

    writer.writerows(sample_data)

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='expense_import_template.csv')


@business_twin_bp.route('/expenses/download-excel-template')
@login_required
def download_expense_excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"

    headers = [
        'department', 'category', 'subcategory', 'amount', 'description',
        'transaction_date', 'reference_number', 'is_recurring', 'recurring_frequency'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    sample_data = [
        ['Sales', 'Operating Expenses', 'Travel & Entertainment', 12500, 'Q2 client meetings', '2026-06-15', 'EXP-001', 'No', ''],
        ['Marketing', 'Operating Expenses', 'Digital Advertising', 45000, 'Google Ads campaign', '2026-06-14', 'EXP-002', 'Yes', 'monthly'],
        ['Operations', 'Operating Expenses', 'Logistics', 28000, 'Warehouse freight', '2026-06-13', 'EXP-003', 'No', ''],
        ['Finance', 'Operating Expenses', 'Software', 8500, 'Accounting license', '2026-06-12', 'EXP-004', 'Yes', 'yearly'],
        ['HR', 'Operating Expenses', 'Recruitment', 15000, 'Agency fees', '2026-06-11', 'EXP-005', 'No', ''],
        ['IT', 'Operating Expenses', 'Cloud', 32000, 'AWS services', '2026-06-10', 'EXP-006', 'Yes', 'monthly']
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='expense_import_template.xlsx')


@business_twin_bp.route('/customers')
@login_required
def customers():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    customers = Customer.query.filter_by(company_id=company_id).order_by(Customer.created_at.desc()).all()

    total_db_customers = len(customers)
    estimated_total = 0
    if total_db_customers == 0 and company:
        estimated_total = AnalyticsService._estimate_customers(company)

    return render_template('business_twin/twin_customers.html', 
                         customers=customers, 
                         company=company,
                         total_db_customers=total_db_customers,
                         estimated_total=estimated_total)


@business_twin_bp.route('/customers/add', methods=['POST'])
@login_required
def add_customer():
    company_id = session.get('company_id')

    customer = Customer(
        company_id=company_id,
        customer_code=request.form.get('customer_code', 'CUST-' + str(datetime.utcnow().timestamp())[-6:]),
        first_name=request.form.get('first_name', ''),
        last_name=request.form.get('last_name', ''),
        company_name=request.form.get('company_name', ''),
        email=request.form.get('email', ''),
        phone=request.form.get('phone', ''),
        address=request.form.get('address', ''),
        city=request.form.get('city', ''),
        country=request.form.get('country', ''),
        segment=request.form.get('segment', 'General'),
        lifetime_value=float(request.form.get('lifetime_value', 0)),
        acquisition_cost=float(request.form.get('acquisition_cost', 0)),
        total_orders=int(request.form.get('total_orders', 0)),
        total_spent=float(request.form.get('total_spent', 0)),
        satisfaction_score=float(request.form.get('satisfaction_score', 0)) if request.form.get('satisfaction_score') else None,
        nps_score=float(request.form.get('nps_score', 0)) if request.form.get('nps_score') else None,
        status=request.form.get('status', 'active')
    )

    db.session.add(customer)
    db.session.commit()

    company = Company.query.get(company_id)
    if company and hasattr(company, 'customer_count'):
        company.customer_count = Customer.query.filter_by(company_id=company_id).count()
        db.session.commit()

    flash('Customer added successfully!', 'success')
    return jsonify({'success': True, 'customer': customer.to_dict()})


@business_twin_bp.route('/customers/edit/<int:customer_id>', methods=['POST'])
@login_required
def edit_customer(customer_id):
    company_id = session.get('company_id')
    customer = Customer.query.filter_by(id=customer_id, company_id=company_id).first()
    
    if not customer:
        return jsonify({'success': False, 'error': 'Customer not found'}), 404
    
    try:
        customer.customer_code = request.form.get('customer_code') or customer.customer_code
        customer.first_name = request.form.get('first_name', customer.first_name or '')
        customer.last_name = request.form.get('last_name', customer.last_name or '')
        customer.company_name = request.form.get('company_name', customer.company_name or '')
        customer.email = request.form.get('email', customer.email or '')
        customer.phone = request.form.get('phone', customer.phone or '')
        customer.address = request.form.get('address', customer.address or '')
        customer.city = request.form.get('city', customer.city or '')
        customer.country = request.form.get('country', customer.country or '')
        customer.segment = request.form.get('segment', customer.segment or 'General')
        customer.lifetime_value = float(request.form.get('lifetime_value', customer.lifetime_value or 0) or 0)
        customer.acquisition_cost = float(request.form.get('acquisition_cost', customer.acquisition_cost or 0) or 0)
        customer.total_orders = int(request.form.get('total_orders', customer.total_orders or 0) or 0)
        customer.total_spent = float(request.form.get('total_spent', customer.total_spent or 0) or 0)
        
        satisfaction = request.form.get('satisfaction_score', '')
        customer.satisfaction_score = float(satisfaction) if satisfaction and satisfaction.strip() else None
        
        nps = request.form.get('nps_score', '')
        customer.nps_score = float(nps) if nps and nps.strip() else None
        
        customer.status = request.form.get('status', customer.status or 'active')
        
        db.session.commit()
        
        return jsonify({'success': True, 'customer': customer.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/customers/delete/<int:customer_id>', methods=['DELETE'])
@login_required
def delete_customer(customer_id):
    company_id = session.get('company_id')
    customer = Customer.query.filter_by(id=customer_id, company_id=company_id).first()
    
    if not customer:
        return jsonify({'success': False, 'error': 'Customer not found'}), 404
    
    try:
        db.session.delete(customer)
        db.session.commit()
        
        company = Company.query.get(company_id)
        if company and hasattr(company, 'customer_count'):
            company.customer_count = Customer.query.filter_by(company_id=company_id).count()
            db.session.commit()
        
        return jsonify({'success': True, 'message': 'Customer deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/customers/delete-all', methods=['DELETE'])
@login_required
def delete_all_customers():
    company_id = session.get('company_id')
    try:
        count = Customer.query.filter_by(company_id=company_id).count()
        Customer.query.filter_by(company_id=company_id).delete()
        db.session.commit()
        
        company = Company.query.get(company_id)
        if company and hasattr(company, 'customer_count'):
            company.customer_count = 0
            db.session.commit()
        
        return jsonify({'success': True, 'message': f'All {count} customers deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/customers/import-csv', methods=['POST'])
@login_required
def import_customers_csv():
    company_id = session.get('company_id')

    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)

        added_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                if not row.get('first_name') or not row.get('last_name') or not row.get('email'):
                    errors.append(f"Row {row_num}: Missing required fields (first_name, last_name, email)")
                    continue

                email = row.get('email', '').strip()
                existing = Customer.query.filter_by(company_id=company_id, email=email).first()
                
                if existing:
                    existing.customer_code = row.get('customer_code', existing.customer_code or '').strip() or existing.customer_code
                    existing.first_name = row.get('first_name', '').strip()
                    existing.last_name = row.get('last_name', '').strip()
                    existing.company_name = row.get('company_name', '').strip()
                    existing.phone = row.get('phone', '').strip()
                    existing.address = row.get('address', '').strip()
                    existing.city = row.get('city', '').strip()
                    existing.country = row.get('country', '').strip()
                    existing.segment = row.get('segment', 'General').strip()
                    existing.lifetime_value = float(row.get('lifetime_value', 0) or 0)
                    existing.acquisition_cost = float(row.get('acquisition_cost', 0) or 0)
                    existing.total_orders = int(row.get('total_orders', 0) or 0)
                    existing.total_spent = float(row.get('total_spent', 0) or 0)
                    existing.satisfaction_score = float(row.get('satisfaction_score', 0)) if row.get('satisfaction_score') else None
                    existing.nps_score = float(row.get('nps_score', 0)) if row.get('nps_score') else None
                    existing.status = row.get('status', 'active').strip()
                    updated_count += 1
                else:
                    customer = Customer(
                        company_id=company_id,
                        customer_code=row.get('customer_code', 'CUST-' + str(datetime.utcnow().timestamp())[-6:] + str(row_num)),
                        first_name=row.get('first_name', '').strip(),
                        last_name=row.get('last_name', '').strip(),
                        company_name=row.get('company_name', '').strip(),
                        email=email,
                        phone=row.get('phone', '').strip(),
                        address=row.get('address', '').strip(),
                        city=row.get('city', '').strip(),
                        country=row.get('country', '').strip(),
                        segment=row.get('segment', 'General').strip(),
                        lifetime_value=float(row.get('lifetime_value', 0) or 0),
                        acquisition_cost=float(row.get('acquisition_cost', 0) or 0),
                        total_orders=int(row.get('total_orders', 0) or 0),
                        total_spent=float(row.get('total_spent', 0) or 0),
                        satisfaction_score=float(row.get('satisfaction_score', 0)) if row.get('satisfaction_score') else None,
                        nps_score=float(row.get('nps_score', 0)) if row.get('nps_score') else None,
                        status=row.get('status', 'active').strip()
                    )
                    db.session.add(customer)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        company = Company.query.get(company_id)
        if company and hasattr(company, 'customer_count'):
            company.customer_count = Customer.query.filter_by(company_id=company_id).count()
            db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/customers/import-excel', methods=['POST'])
@login_required
def import_customers_excel():
    company_id = session.get('company_id')

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active

        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_').replace('-', '_')
                headers.append(h)
            else:
                headers.append('')

        header_map = {}
        for idx, h in enumerate(headers):
            if h in ['customer_code', 'code', 'cust_code', 'id']:
                header_map['customer_code'] = idx
            elif h in ['first_name', 'firstname', 'first_name']:
                header_map['first_name'] = idx
            elif h in ['last_name', 'lastname', 'last_name']:
                header_map['last_name'] = idx
            elif h in ['company_name', 'company', 'organization', 'org_name']:
                header_map['company_name'] = idx
            elif h in ['email', 'email_address', 'e_mail', 'e-mail', 'mail']:
                header_map['email'] = idx
            elif h in ['phone', 'phone_number', 'telephone', 'mobile', 'contact']:
                header_map['phone'] = idx
            elif h in ['address', 'street', 'street_address']:
                header_map['address'] = idx
            elif h in ['city', 'town']:
                header_map['city'] = idx
            elif h in ['country', 'nation', 'region']:
                header_map['country'] = idx
            elif h in ['segment', 'category', 'type', 'customer_segment']:
                header_map['segment'] = idx
            elif h in ['lifetime_value', 'ltv', 'lifetime_value', 'customer_lifetime_value']:
                header_map['lifetime_value'] = idx
            elif h in ['acquisition_cost', 'acq_cost', 'cost', 'cac', 'customer_acquisition_cost']:
                header_map['acquisition_cost'] = idx
            elif h in ['total_orders', 'orders', 'order_count', 'num_orders']:
                header_map['total_orders'] = idx
            elif h in ['total_spent', 'spent', 'revenue', 'amount_spent']:
                header_map['total_spent'] = idx
            elif h in ['satisfaction_score', 'satisfaction', 'csat', 'customer_satisfaction']:
                header_map['satisfaction_score'] = idx
            elif h in ['nps_score', 'nps', 'net_promoter_score', 'promoter_score']:
                header_map['nps_score'] = idx
            elif h in ['status', 'state', 'customer_status']:
                header_map['status'] = idx

        if 'first_name' not in header_map or 'last_name' not in header_map or 'email' not in header_map:
            return jsonify({'success': False, 'error': 'Excel file must have columns: first_name, last_name, email'}), 400

        added_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                first_name_idx = header_map.get('first_name', 0)
                last_name_idx = header_map.get('last_name', 1)
                email_idx = header_map.get('email', 2)

                first_name = str(row[first_name_idx] if first_name_idx < len(row) else '').strip()
                last_name = str(row[last_name_idx] if last_name_idx < len(row) else '').strip()
                email = str(row[email_idx] if email_idx < len(row) else '').strip()

                if not first_name or not last_name or not email:
                    errors.append(f"Row {row_num}: Missing required fields (first_name, last_name, email)")
                    continue

                def get_val(key, default=''):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val is not None else default
                    return default

                def get_float_val(key, default=0.0):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        if val is not None:
                            try:
                                return float(val)
                            except (ValueError, TypeError):
                                return default
                    return default

                def get_int_val(key, default=0):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        if val is not None:
                            try:
                                return int(float(val))
                            except (ValueError, TypeError):
                                return default
                    return default

                existing = Customer.query.filter_by(company_id=company_id, email=email).first()
                
                if existing:
                    existing.customer_code = get_val('customer_code', existing.customer_code or '') or existing.customer_code
                    existing.first_name = first_name
                    existing.last_name = last_name
                    existing.company_name = get_val('company_name')
                    existing.phone = get_val('phone')
                    existing.address = get_val('address')
                    existing.city = get_val('city')
                    existing.country = get_val('country')
                    existing.segment = get_val('segment', 'General')
                    existing.lifetime_value = get_float_val('lifetime_value')
                    existing.acquisition_cost = get_float_val('acquisition_cost')
                    existing.total_orders = get_int_val('total_orders')
                    existing.total_spent = get_float_val('total_spent')
                    existing.satisfaction_score = get_float_val('satisfaction_score') if get_val('satisfaction_score') else None
                    existing.nps_score = get_float_val('nps_score') if get_val('nps_score') else None
                    existing.status = get_val('status', 'active')
                    updated_count += 1
                else:
                    customer = Customer(
                        company_id=company_id,
                        customer_code=get_val('customer_code', 'CUST-' + str(datetime.utcnow().timestamp())[-6:] + str(row_num)),
                        first_name=first_name,
                        last_name=last_name,
                        company_name=get_val('company_name'),
                        email=email,
                        phone=get_val('phone'),
                        address=get_val('address'),
                        city=get_val('city'),
                        country=get_val('country'),
                        segment=get_val('segment', 'General'),
                        lifetime_value=get_float_val('lifetime_value'),
                        acquisition_cost=get_float_val('acquisition_cost'),
                        total_orders=get_int_val('total_orders'),
                        total_spent=get_float_val('total_spent'),
                        satisfaction_score=get_float_val('satisfaction_score') if get_val('satisfaction_score') else None,
                        nps_score=get_float_val('nps_score') if get_val('nps_score') else None,
                        status=get_val('status', 'active')
                    )
                    db.session.add(customer)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        company = Company.query.get(company_id)
        if company and hasattr(company, 'customer_count'):
            company.customer_count = Customer.query.filter_by(company_id=company_id).count()
            db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/customers/export-excel')
@login_required
def export_customers_excel():
    company_id = session.get('company_id')
    customers = Customer.query.filter_by(company_id=company_id).order_by(Customer.created_at.desc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ['ID', 'Customer Code', 'First Name', 'Last Name', 'Company Name', 'Email', 
               'Phone', 'Address', 'City', 'Country', 'Segment', 'Lifetime Value', 
               'Acquisition Cost', 'Total Orders', 'Total Spent', 'Satisfaction Score', 
               'NPS Score', 'Status', 'Created At']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_num, customer in enumerate(customers, 2):
        ws.cell(row=row_num, column=1, value=customer.id)
        ws.cell(row=row_num, column=2, value=customer.customer_code)
        ws.cell(row=row_num, column=3, value=customer.first_name)
        ws.cell(row=row_num, column=4, value=customer.last_name)
        ws.cell(row=row_num, column=5, value=customer.company_name)
        ws.cell(row=row_num, column=6, value=customer.email)
        ws.cell(row=row_num, column=7, value=customer.phone)
        ws.cell(row=row_num, column=8, value=customer.address)
        ws.cell(row=row_num, column=9, value=customer.city)
        ws.cell(row=row_num, column=10, value=customer.country)
        ws.cell(row=row_num, column=11, value=customer.segment)
        ws.cell(row=row_num, column=12, value=customer.lifetime_value)
        ws.cell(row=row_num, column=13, value=customer.acquisition_cost)
        ws.cell(row=row_num, column=14, value=customer.total_orders)
        ws.cell(row=row_num, column=15, value=customer.total_spent)
        ws.cell(row=row_num, column=16, value=customer.satisfaction_score)
        ws.cell(row=row_num, column=17, value=customer.nps_score)
        ws.cell(row=row_num, column=18, value=customer.status)
        ws.cell(row=row_num, column=19, value=customer.created_at.strftime('%Y-%m-%d %H:%M:%S') if customer.created_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='customers_export.xlsx')


@business_twin_bp.route('/customers/export-pdf')
@login_required
def export_customers_pdf():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    customers = Customer.query.filter_by(company_id=company_id).order_by(Customer.created_at.desc()).all()

    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    symbol = _get_currency_symbol(company_id)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a1f3a'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph(f"{company.company_name if company else 'Company'} - Customer Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=15)
    elements.append(Paragraph(f"Total Customers: {len(customers)} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    elements.append(Spacer(1, 0.1*inch))

    table_data = [['Name', 'Email', 'Segment', 'LTV', 'Orders', 'Spent', 'Status']]
    for c in customers:
        table_data.append([
            c.full_name() or '-',
            c.email or '-',
            c.segment or 'General',
            f"{symbol}{c.lifetime_value:,.0f}" if c.lifetime_value else f'{symbol}0',
            str(c.total_orders),
            f"{symbol}{c.total_spent:,.0f}" if c.total_spent else f'{symbol}0',
            c.status.title()
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f3a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8f9fa')])
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name='customers_report.pdf')


@business_twin_bp.route('/customers/download-template')
@login_required
def download_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'customer_code', 'first_name', 'last_name', 'company_name', 'email', 'phone',
        'address', 'city', 'country', 'segment', 'lifetime_value', 'acquisition_cost',
        'total_orders', 'total_spent', 'satisfaction_score', 'nps_score', 'status'
    ])

    writer.writerow([
        'CUST-001', 'John', 'Doe', 'Acme Corp', 'john.doe@example.com', '+1-555-0101',
        '123 Main St', 'New York', 'USA', 'Enterprise', '50000', '500',
        '10', '45000', '4.5', '8', 'active'
    ])
    writer.writerow([
        'CUST-002', 'Jane', 'Smith', 'TechStart Inc', 'jane.smith@example.com', '+1-555-0102',
        '456 Oak Ave', 'San Francisco', 'USA', 'SMB', '15000', '200',
        '5', '12000', '4.2', '7', 'active'
    ])
    writer.writerow([
        'CUST-003', 'Ahmed', 'Khan', 'Global Solutions', 'ahmed.khan@example.com', '+91-88888-88888',
        '789 Park Road', 'Bangalore', 'India', 'Enterprise', '80000', '800',
        '15', '75000', '4.8', '9', 'active'
    ])

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='customer_import_template.csv')


@business_twin_bp.route('/customers/download-excel-template')
@login_required
def download_excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Customers"

    headers = [
        'customer_code', 'first_name', 'last_name', 'company_name', 'email', 'phone',
        'address', 'city', 'country', 'segment', 'lifetime_value', 'acquisition_cost',
        'total_orders', 'total_spent', 'satisfaction_score', 'nps_score', 'status'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    sample_data = [
        ['CUST-001', 'John', 'Doe', 'Acme Corp', 'john.doe@example.com', '+1-555-0101',
         '123 Main St', 'New York', 'USA', 'Enterprise', 50000, 500, 10, 45000, 4.5, 8, 'active'],
        ['CUST-002', 'Jane', 'Smith', 'TechStart Inc', 'jane.smith@example.com', '+1-555-0102',
         '456 Oak Ave', 'San Francisco', 'USA', 'SMB', 15000, 200, 5, 12000, 4.2, 7, 'active'],
        ['CUST-003', 'Ahmed', 'Khan', 'Global Solutions', 'ahmed.khan@example.com', '+91-88888-88888',
         '789 Park Road', 'Bangalore', 'India', 'Enterprise', 80000, 800, 15, 75000, 4.8, 9, 'active']
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='customer_import_template.xlsx')


@business_twin_bp.route('/employees')
@login_required
def employees():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    employees = Employee.query.filter_by(company_id=company_id).order_by(Employee.created_at.desc()).limit(100).all()
    departments = Department.query.filter_by(company_id=company_id, is_active=True).all()

    total_db_employees = len(employees)
    active_count = len([e for e in employees if e.status == 'active'])
    unique_dept_ids = set(e.department_id for e in employees if e.department_id)
    dept_count = len(unique_dept_ids)
    avg_salary = sum(e.salary for e in employees) / len(employees) if employees else 0

    dept_map = {d.id: d.name for d in departments}

    return render_template('business_twin/twin_employees.html', 
                         employees=employees, 
                         departments=departments,
                         dept_map=dept_map,
                         total_db_employees=total_db_employees,
                         active_count=active_count,
                         dept_count=dept_count,
                         avg_salary=avg_salary,
                         company=company)


@business_twin_bp.route('/employees/add', methods=['POST'])
@login_required
def add_employee():
    company_id = session.get('company_id')

    try:
        dept_id = request.form.get('department_id')

        employee = Employee(
            company_id=company_id,
            employee_id=request.form.get('employee_id', ''),
            first_name=request.form.get('first_name', ''),
            last_name=request.form.get('last_name', ''),
            email=request.form.get('email', ''),
            phone=request.form.get('phone', ''),
            department_id=int(dept_id) if dept_id and dept_id.strip() else None,
            job_title=request.form.get('job_title', ''),
            employment_type=request.form.get('employment_type', 'full_time'),
            salary=float(request.form.get('salary', 0) or 0),
            benefits_cost=float(request.form.get('benefits_cost', 0) or 0),
            hire_date=datetime.strptime(request.form.get('hire_date'), '%Y-%m-%d').date() if request.form.get('hire_date') else None,
            status=request.form.get('status', 'active'),
            performance_rating=float(request.form.get('performance_rating', 0)) if request.form.get('performance_rating') else None
        )

        db.session.add(employee)
        db.session.commit()
        
        # Sync department headcount
        if dept_id:
            _sync_dept_headcount(company_id, int(dept_id))

        company = Company.query.get(company_id)
        if company and hasattr(company, 'employee_count'):
            company.employee_count = Employee.query.filter_by(company_id=company_id).count()
            db.session.commit()

        return jsonify({'success': True, 'employee': employee.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/employees/edit/<int:employee_id>', methods=['POST'])
@login_required
def edit_employee(employee_id):
    company_id = session.get('company_id')
    employee = Employee.query.filter_by(id=employee_id, company_id=company_id).first()
    
    if not employee:
        return jsonify({'success': False, 'error': 'Employee not found'}), 404
    
    try:
        old_dept_id = employee.department_id
        
        employee.employee_id = request.form.get('employee_id') or employee.employee_id
        employee.first_name = request.form.get('first_name', employee.first_name or '')
        employee.last_name = request.form.get('last_name', employee.last_name or '')
        employee.email = request.form.get('email', employee.email or '')
        employee.phone = request.form.get('phone', employee.phone or '')
        employee.job_title = request.form.get('job_title', employee.job_title or '')
        
        dept_id = request.form.get('department_id')
        if dept_id and dept_id.strip():
            employee.department_id = int(dept_id)
        else:
            employee.department_id = None
            
        employee.employment_type = request.form.get('employment_type', employee.employment_type or 'full_time')
        employee.salary = float(request.form.get('salary', employee.salary or 0) or 0)
        employee.benefits_cost = float(request.form.get('benefits_cost', employee.benefits_cost or 0) or 0)
        
        hire_date_str = request.form.get('hire_date', '')
        if hire_date_str and hire_date_str.strip():
            employee.hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
        else:
            employee.hire_date = None
        
        employee.status = request.form.get('status', employee.status or 'active')
        
        rating = request.form.get('performance_rating', '')
        employee.performance_rating = float(rating) if rating and rating.strip() else None
        
        db.session.commit()
        
        # Sync department headcounts
        if old_dept_id:
            _sync_dept_headcount(company_id, old_dept_id)
        if employee.department_id and employee.department_id != old_dept_id:
            _sync_dept_headcount(company_id, employee.department_id)
        
        return jsonify({'success': True, 'employee': employee.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/employees/delete/<int:employee_id>', methods=['DELETE'])
@login_required
def delete_employee(employee_id):
    company_id = session.get('company_id')
    employee = Employee.query.filter_by(id=employee_id, company_id=company_id).first()
    
    if not employee:
        return jsonify({'success': False, 'error': 'Employee not found'}), 404
    
    try:
        old_dept_id = employee.department_id
        
        db.session.delete(employee)
        db.session.commit()
        
        # Sync department headcount
        if old_dept_id:
            _sync_dept_headcount(company_id, old_dept_id)
        
        # Sync company employee count
        company = Company.query.get(company_id)
        if company and hasattr(company, 'employee_count'):
            company.employee_count = Employee.query.filter_by(company_id=company_id).count()
            db.session.commit()
        
        return jsonify({'success': True, 'message': 'Employee deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/employees/delete-all', methods=['DELETE'])
@login_required
def delete_all_employees():
    company_id = session.get('company_id')
    try:
        # Collect department IDs before deletion for sync
        dept_ids = set()
        employees = Employee.query.filter_by(company_id=company_id).all()
        for emp in employees:
            if emp.department_id:
                dept_ids.add(emp.department_id)
        
        count = Employee.query.filter_by(company_id=company_id).count()
        Employee.query.filter_by(company_id=company_id).delete()
        db.session.commit()
        
        # Sync all affected departments
        for dept_id in dept_ids:
            _sync_dept_headcount(company_id, dept_id)
        
        company = Company.query.get(company_id)
        if company and hasattr(company, 'employee_count'):
            company.employee_count = 0
            db.session.commit()
        
        return jsonify({'success': True, 'message': f'All {count} employees deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/employees/import-csv', methods=['POST'])
@login_required
def import_employees_csv():
    company_id = session.get('company_id')

    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)

        added_count = 0
        updated_count = 0
        errors = []
        affected_dept_ids = set()

        depts = {d.name.lower(): d.id for d in Department.query.filter_by(company_id=company_id).all()}

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                first_name = row.get('first_name', '').strip()
                last_name = row.get('last_name', '').strip()

                if not first_name or not last_name:
                    errors.append(f"Row {row_num}: Missing required fields (first_name, last_name)")
                    continue

                email = row.get('email', '').strip()
                existing = Employee.query.filter_by(company_id=company_id, email=email).first() if email else None

                dept_id = None
                dept_name = row.get('department', '').strip()
                if dept_name:
                    dept_id = depts.get(dept_name.lower())

                hire_date = None
                date_str = row.get('hire_date', '').strip()
                if date_str:
                    try:
                        hire_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    except:
                        try:
                            hire_date = datetime.strptime(date_str, '%m/%d/%Y').date()
                        except:
                            pass

                if existing:
                    old_dept_id = existing.department_id
                    existing.employee_id = row.get('employee_id', '').strip() or existing.employee_id
                    existing.first_name = first_name
                    existing.last_name = last_name
                    existing.email = email
                    existing.phone = row.get('phone', '').strip()
                    existing.job_title = row.get('job_title', '').strip()
                    existing.department_id = dept_id
                    existing.employment_type = row.get('employment_type', 'full_time').strip()
                    existing.salary = float(row.get('salary', 0) or 0)
                    existing.benefits_cost = float(row.get('benefits_cost', 0) or 0)
                    existing.hire_date = hire_date
                    existing.status = row.get('status', 'active').strip()
                    existing.performance_rating = float(row.get('performance_rating', 0)) if row.get('performance_rating') else None
                    updated_count += 1
                    if old_dept_id:
                        affected_dept_ids.add(old_dept_id)
                    if dept_id and dept_id != old_dept_id:
                        affected_dept_ids.add(dept_id)
                else:
                    employee = Employee(
                        company_id=company_id,
                        employee_id=row.get('employee_id', '').strip(),
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        phone=row.get('phone', '').strip(),
                        department_id=dept_id,
                        job_title=row.get('job_title', '').strip(),
                        employment_type=row.get('employment_type', 'full_time').strip(),
                        salary=float(row.get('salary', 0) or 0),
                        benefits_cost=float(row.get('benefits_cost', 0) or 0),
                        hire_date=hire_date,
                        status=row.get('status', 'active').strip(),
                        performance_rating=float(row.get('performance_rating', 0)) if row.get('performance_rating') else None
                    )
                    db.session.add(employee)
                    added_count += 1
                    if dept_id:
                        affected_dept_ids.add(dept_id)

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        
        for dept_id in affected_dept_ids:
            _sync_dept_headcount(company_id, dept_id)

        company = Company.query.get(company_id)
        if company and hasattr(company, 'employee_count'):
            company.employee_count = Employee.query.filter_by(company_id=company_id).count()
            db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/employees/import-excel', methods=['POST'])
@login_required
def import_employees_excel():
    company_id = session.get('company_id')

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active

        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_').replace('-', '_')
                headers.append(h)
            else:
                headers.append('')

        header_map = {}
        for idx, h in enumerate(headers):
            if h in ['first_name', 'firstname', 'first']:
                header_map['first_name'] = idx
            elif h in ['last_name', 'lastname', 'last']:
                header_map['last_name'] = idx
            elif h in ['email', 'email_address', 'e_mail']:
                header_map['email'] = idx
            elif h in ['phone', 'phone_number', 'telephone', 'mobile']:
                header_map['phone'] = idx
            elif h in ['employee_id', 'emp_id', 'id', 'employee_code']:
                header_map['employee_id'] = idx
            elif h in ['job_title', 'title', 'position', 'role']:
                header_map['job_title'] = idx
            elif h in ['department', 'dept', 'department_name', 'dept_name']:
                header_map['department'] = idx
            elif h in ['employment_type', 'type', 'emp_type']:
                header_map['employment_type'] = idx
            elif h in ['salary', 'annual_salary', 'base_salary']:
                header_map['salary'] = idx
            elif h in ['benefits_cost', 'benefits', 'benefit_cost']:
                header_map['benefits_cost'] = idx
            elif h in ['hire_date', 'date_hired', 'start_date']:
                header_map['hire_date'] = idx
            elif h in ['status', 'state', 'employee_status']:
                header_map['status'] = idx
            elif h in ['performance_rating', 'rating', 'performance', 'score']:
                header_map['performance_rating'] = idx

        if 'first_name' not in header_map or 'last_name' not in header_map:
            return jsonify({'success': False, 'error': 'Excel file must have columns: first_name, last_name'}), 400

        depts = {d.name.lower(): d.id for d in Department.query.filter_by(company_id=company_id).all()}

        added_count = 0
        updated_count = 0
        errors = []
        affected_dept_ids = set()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                first_name_idx = header_map.get('first_name', 0)
                last_name_idx = header_map.get('last_name', 1)

                first_name = str(row[first_name_idx] if first_name_idx < len(row) else '').strip()
                last_name = str(row[last_name_idx] if last_name_idx < len(row) else '').strip()

                if not first_name or not last_name:
                    errors.append(f"Row {row_num}: Missing required fields (first_name, last_name)")
                    continue

                def get_val(key, default=''):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        return str(val).strip() if val is not None else default
                    return default

                def get_float_val(key, default=0.0):
                    idx = header_map.get(key)
                    if idx is not None and idx < len(row):
                        val = row[idx]
                        if val is not None:
                            try:
                                return float(val)
                            except (ValueError, TypeError):
                                return default
                    return default

                email = get_val('email')
                existing = Employee.query.filter_by(company_id=company_id, email=email).first() if email else None

                dept_id = None
                dept_name = get_val('department')
                if dept_name:
                    dept_id = depts.get(dept_name.lower())
                    if not dept_id:
                        for d_name, d_id in depts.items():
                            if dept_name.lower() in d_name or d_name in dept_name.lower():
                                dept_id = d_id
                                break

                hire_date = None
                hire_idx = header_map.get('hire_date')
                if hire_idx is not None and hire_idx < len(row) and row[hire_idx]:
                    date_val = row[hire_idx]
                    if isinstance(date_val, datetime):
                        hire_date = date_val.date()
                    elif isinstance(date_val, date):
                        hire_date = date_val
                    else:
                        date_str = str(date_val).strip()
                        for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                            try:
                                hire_date = datetime.strptime(date_str, fmt).date()
                                break
                            except:
                                continue

                rating_val = get_val('performance_rating')
                performance_rating = None
                if rating_val:
                    try:
                        performance_rating = float(rating_val)
                    except:
                        pass

                if existing:
                    old_dept_id = existing.department_id
                    existing.employee_id = get_val('employee_id') or existing.employee_id
                    existing.first_name = first_name
                    existing.last_name = last_name
                    existing.email = email
                    existing.phone = get_val('phone')
                    existing.department_id = dept_id
                    existing.job_title = get_val('job_title')
                    existing.employment_type = get_val('employment_type', 'full_time')
                    existing.salary = get_float_val('salary')
                    existing.benefits_cost = get_float_val('benefits_cost')
                    existing.hire_date = hire_date
                    existing.status = get_val('status', 'active')
                    existing.performance_rating = performance_rating
                    updated_count += 1
                    if old_dept_id:
                        affected_dept_ids.add(old_dept_id)
                    if dept_id and dept_id != old_dept_id:
                        affected_dept_ids.add(dept_id)
                else:
                    employee = Employee(
                        company_id=company_id,
                        employee_id=get_val('employee_id'),
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        phone=get_val('phone'),
                        department_id=dept_id,
                        job_title=get_val('job_title'),
                        employment_type=get_val('employment_type', 'full_time'),
                        salary=get_float_val('salary'),
                        benefits_cost=get_float_val('benefits_cost'),
                        hire_date=hire_date,
                        status=get_val('status', 'active'),
                        performance_rating=performance_rating
                    )
                    db.session.add(employee)
                    added_count += 1
                    if dept_id:
                        affected_dept_ids.add(dept_id)

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()
        
        for dept_id in affected_dept_ids:
            _sync_dept_headcount(company_id, dept_id)

        company = Company.query.get(company_id)
        if company and hasattr(company, 'employee_count'):
            company.employee_count = Employee.query.filter_by(company_id=company_id).count()
            db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/employees/export-excel')
@login_required
def export_employees_excel():
    company_id = session.get('company_id')
    employees = Employee.query.filter_by(company_id=company_id).order_by(Employee.created_at.desc()).all()
    departments = Department.query.filter_by(company_id=company_id).all()
    dept_map = {d.id: d.name for d in departments}

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ['ID', 'Employee ID', 'First Name', 'Last Name', 'Email', 'Phone', 
               'Department', 'Job Title', 'Employment Type', 'Salary', 'Benefits Cost', 
               'Total Cost', 'Hire Date', 'Status', 'Performance Rating', 'Created At']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_num, emp in enumerate(employees, 2):
        ws.cell(row=row_num, column=1, value=emp.id)
        ws.cell(row=row_num, column=2, value=emp.employee_id)
        ws.cell(row=row_num, column=3, value=emp.first_name)
        ws.cell(row=row_num, column=4, value=emp.last_name)
        ws.cell(row=row_num, column=5, value=emp.email)
        ws.cell(row=row_num, column=6, value=emp.phone)
        ws.cell(row=row_num, column=7, value=dept_map.get(emp.department_id, 'Unassigned'))
        ws.cell(row=row_num, column=8, value=emp.job_title)
        ws.cell(row=row_num, column=9, value=emp.employment_type)
        ws.cell(row=row_num, column=10, value=emp.salary)
        ws.cell(row=row_num, column=11, value=emp.benefits_cost)
        ws.cell(row=row_num, column=12, value=emp.total_cost())
        ws.cell(row=row_num, column=13, value=emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '')
        ws.cell(row=row_num, column=14, value=emp.status)
        ws.cell(row=row_num, column=15, value=emp.performance_rating)
        ws.cell(row=row_num, column=16, value=emp.created_at.strftime('%Y-%m-%d %H:%M:%S') if emp.created_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='employees_export.xlsx')


@business_twin_bp.route('/employees/export-pdf')
@login_required
def export_employees_pdf():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    employees = Employee.query.filter_by(company_id=company_id).order_by(Employee.created_at.desc()).all()
    departments = Department.query.filter_by(company_id=company_id).all()
    dept_map = {d.id: d.name for d in departments}

    symbol = _get_currency_symbol(company_id)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a1f3a'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph(f"{company.company_name if company else 'Company'} - Employee Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=15)
    elements.append(Paragraph(f"Total Employees: {len(employees)} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    elements.append(Spacer(1, 0.1*inch))

    table_data = [['Name', 'Email', 'Department', 'Job Title', 'Salary', 'Status', 'Rating']]
    for e in employees:
        table_data.append([
            e.full_name() or '-',
            e.email or '-',
            dept_map.get(e.department_id, 'Unassigned'),
            e.job_title or '-',
            f"{symbol}{e.salary:,.0f}" if e.salary else f'{symbol}0',
            e.status.title(),
            str(e.performance_rating) if e.performance_rating else '-'
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f3a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8f9fa')])
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name='employees_report.pdf')


@business_twin_bp.route('/employees/download-template')
@login_required
def download_employee_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'first_name', 'last_name', 'email', 'phone', 'employee_id', 'job_title',
        'department', 'employment_type', 'salary', 'benefits_cost', 'hire_date', 'status', 'performance_rating'
    ])

    writer.writerow([
        'John', 'Doe', 'john.doe@company.com', '+1-555-0101', 'EMP-001',
        'Software Engineer', 'IT', 'full_time', '85000', '12000', '2024-01-15', 'active', '4.5'
    ])
    writer.writerow([
        'Jane', 'Smith', 'jane.smith@company.com', '+1-555-0102', 'EMP-002',
        'Marketing Manager', 'Marketing', 'full_time', '75000', '10000', '2023-08-20', 'active', '4.2'
    ])
    writer.writerow([
        'Ahmed', 'Khan', 'ahmed.khan@company.com', '+91-88888-88888', 'EMP-003',
        'Sales Lead', 'Sales', 'full_time', '95000', '15000', '2022-05-10', 'active', '4.8'
    ])

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='employee_import_template.csv')


@business_twin_bp.route('/employees/download-excel-template')
@login_required
def download_employee_excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    headers = [
        'first_name', 'last_name', 'email', 'phone', 'employee_id', 'job_title',
        'department', 'employment_type', 'salary', 'benefits_cost', 'hire_date', 'status', 'performance_rating'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    sample_data = [
        ['John', 'Doe', 'john.doe@company.com', '+1-555-0101', 'EMP-001',
         'Software Engineer', 'IT', 'full_time', 85000, 12000, '2024-01-15', 'active', 4.5],
        ['Jane', 'Smith', 'jane.smith@company.com', '+1-555-0102', 'EMP-002',
         'Marketing Manager', 'Marketing', 'full_time', 75000, 10000, '2023-08-20', 'active', 4.2],
        ['Ahmed', 'Khan', 'ahmed.khan@company.com', '+91-88888-88888', 'EMP-003',
         'Sales Lead', 'Sales', 'full_time', 95000, 15000, '2022-05-10', 'active', 4.8]
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='employee_import_template.xlsx')


@business_twin_bp.route('/suppliers')
@login_required
def suppliers():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    suppliers = Supplier.query.filter_by(company_id=company_id).order_by(Supplier.created_at.desc()).all()

    total_db_suppliers = len(suppliers)
    active_count = len([s for s in suppliers if s.status == 'active'])
    primary_count = len([s for s in suppliers if s.is_primary])
    avg_rating = sum(s.rating for s in suppliers if s.rating) / len([s for s in suppliers if s.rating]) if any(s.rating for s in suppliers) else 0
    total_spend = sum(s.total_spend for s in suppliers)

    return render_template('business_twin/twin_suppliers.html', 
                         suppliers=suppliers,
                         total_db_suppliers=total_db_suppliers,
                         active_count=active_count,
                         primary_count=primary_count,
                         avg_rating=avg_rating,
                         total_spend=total_spend,
                         company=company,
                         currency_symbol=_get_currency_symbol(company_id))


@business_twin_bp.route('/suppliers/add', methods=['POST'])
@login_required
def add_supplier():
    company_id = session.get('company_id')

    try:
        supplier = Supplier(
            company_id=company_id,
            name=request.form.get('name', ''),
            contact_person=request.form.get('contact_person', ''),
            email=request.form.get('email', ''),
            phone=request.form.get('phone', ''),
            address=request.form.get('address', ''),
            city=request.form.get('city', ''),
            country=request.form.get('country', ''),
            category=request.form.get('category', 'General'),
            payment_terms=request.form.get('payment_terms', ''),
            lead_time_days=int(request.form.get('lead_time_days', 0) or 0),
            rating=float(request.form.get('rating', 0)) if request.form.get('rating') else None,
            quality_score=float(request.form.get('quality_score', 0)) if request.form.get('quality_score') else None,
            reliability_score=float(request.form.get('reliability_score', 0)) if request.form.get('reliability_score') else None,
            cost_rating=float(request.form.get('cost_rating', 0)) if request.form.get('cost_rating') else None,
            total_spend=float(request.form.get('total_spend', 0) or 0),
            total_orders=int(request.form.get('total_orders', 0) or 0),
            is_primary=request.form.get('is_primary') == 'true',
            status=request.form.get('status', 'active'),
            contract_start=datetime.strptime(request.form.get('contract_start'), '%Y-%m-%d').date() if request.form.get('contract_start') else None,
            contract_end=datetime.strptime(request.form.get('contract_end'), '%Y-%m-%d').date() if request.form.get('contract_end') else None
        )

        db.session.add(supplier)
        db.session.commit()

        return jsonify({'success': True, 'supplier': supplier.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/suppliers/edit/<int:supplier_id>', methods=['POST'])
@login_required
def edit_supplier(supplier_id):
    company_id = session.get('company_id')
    supplier = Supplier.query.filter_by(id=supplier_id, company_id=company_id).first()
    if not supplier:
        return jsonify({'success': False, 'error': 'Supplier not found'}), 404

    try:
        supplier.name = request.form.get('name', supplier.name)
        supplier.contact_person = request.form.get('contact_person', supplier.contact_person or '')
        supplier.email = request.form.get('email', supplier.email or '')
        supplier.phone = request.form.get('phone', supplier.phone or '')
        supplier.address = request.form.get('address', supplier.address or '')
        supplier.city = request.form.get('city', supplier.city or '')
        supplier.country = request.form.get('country', supplier.country or '')
        supplier.category = request.form.get('category', supplier.category or 'General')
        supplier.payment_terms = request.form.get('payment_terms', supplier.payment_terms or '')
        supplier.lead_time_days = int(request.form.get('lead_time_days', supplier.lead_time_days or 0) or 0)
        supplier.rating = float(request.form.get('rating', supplier.rating or 0)) if request.form.get('rating') else supplier.rating
        supplier.quality_score = float(request.form.get('quality_score', supplier.quality_score or 0)) if request.form.get('quality_score') else supplier.quality_score
        supplier.reliability_score = float(request.form.get('reliability_score', supplier.reliability_score or 0)) if request.form.get('reliability_score') else supplier.reliability_score
        supplier.cost_rating = float(request.form.get('cost_rating', supplier.cost_rating or 0)) if request.form.get('cost_rating') else supplier.cost_rating
        supplier.total_spend = float(request.form.get('total_spend', supplier.total_spend or 0) or 0)
        supplier.total_orders = int(request.form.get('total_orders', supplier.total_orders or 0) or 0)
        supplier.is_primary = request.form.get('is_primary') == 'true'
        supplier.status = request.form.get('status', supplier.status or 'active')

        contract_start = request.form.get('contract_start')
        if contract_start and contract_start.strip():
            supplier.contract_start = datetime.strptime(contract_start, '%Y-%m-%d').date()
        else:
            supplier.contract_start = None

        contract_end = request.form.get('contract_end')
        if contract_end and contract_end.strip():
            supplier.contract_end = datetime.strptime(contract_end, '%Y-%m-%d').date()
        else:
            supplier.contract_end = None

        db.session.commit()
        return jsonify({'success': True, 'supplier': supplier.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/suppliers/delete/<int:supplier_id>', methods=['DELETE'])
@login_required
def delete_supplier(supplier_id):
    company_id = session.get('company_id')
    supplier = Supplier.query.filter_by(id=supplier_id, company_id=company_id).first()
    if not supplier:
        return jsonify({'success': False, 'error': 'Supplier not found'}), 404

    try:
        db.session.delete(supplier)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Supplier deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/suppliers/delete-all', methods=['DELETE'])
@login_required
def delete_all_suppliers():
    company_id = session.get('company_id')
    try:
        count = Supplier.query.filter_by(company_id=company_id).count()
        Supplier.query.filter_by(company_id=company_id).delete()
        db.session.commit()
        return jsonify({
            'success': True,
            'message': f'All {count} supplier records deleted successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/suppliers/import-csv', methods=['POST'])
@login_required
def import_suppliers_csv():
    company_id = session.get('company_id')

    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)

        added_count = 0
        updated_count = 0
        errors = []

        def parse_date(date_str):
            if not date_str or not date_str.strip():
                return None
            for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                try:
                    return datetime.strptime(date_str.strip(), fmt).date()
                except:
                    continue
            return None

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                name = row.get('name', '').strip()
                if not name:
                    errors.append(f"Row {row_num}: Missing required field (name)")
                    continue

                email = row.get('email', '').strip()

                # Upsert logic: match by email if present, otherwise by name
                existing = None
                if email:
                    existing = Supplier.query.filter_by(company_id=company_id, email=email).first()
                if not existing and name:
                    existing = Supplier.query.filter_by(company_id=company_id, name=name).first()

                if existing:
                    existing.name = name
                    if row.get('contact_person') and row['contact_person'].strip():
                        existing.contact_person = row['contact_person'].strip()
                    if email:
                        existing.email = email
                    if row.get('phone') and row['phone'].strip():
                        existing.phone = row['phone'].strip()
                    if row.get('address') and row['address'].strip():
                        existing.address = row['address'].strip()
                    if row.get('city') and row['city'].strip():
                        existing.city = row['city'].strip()
                    if row.get('country') and row['country'].strip():
                        existing.country = row['country'].strip()
                    if row.get('category') and row['category'].strip():
                        existing.category = row['category'].strip()
                    if row.get('payment_terms') and row['payment_terms'].strip():
                        existing.payment_terms = row['payment_terms'].strip()
                    if row.get('lead_time_days') and row['lead_time_days'].strip():
                        existing.lead_time_days = int(row['lead_time_days'].strip() or 0)
                    if row.get('rating') and row['rating'].strip():
                        existing.rating = float(row['rating'].strip())
                    if row.get('quality_score') and row['quality_score'].strip():
                        existing.quality_score = float(row['quality_score'].strip())
                    if row.get('reliability_score') and row['reliability_score'].strip():
                        existing.reliability_score = float(row['reliability_score'].strip())
                    if row.get('cost_rating') and row['cost_rating'].strip():
                        existing.cost_rating = float(row['cost_rating'].strip())
                    if row.get('total_spend') and row['total_spend'].strip():
                        existing.total_spend = float(row['total_spend'].strip() or 0)
                    if row.get('total_orders') and row['total_orders'].strip():
                        existing.total_orders = int(row['total_orders'].strip() or 0)
                    if row.get('is_primary') and row['is_primary'].strip():
                        existing.is_primary = row['is_primary'].strip().lower() in ['yes', 'true', '1', 'on']
                    if row.get('status') and row['status'].strip():
                        existing.status = row['status'].strip()
                    if row.get('contract_start') and row['contract_start'].strip():
                        existing.contract_start = parse_date(row['contract_start'].strip())
                    if row.get('contract_end') and row['contract_end'].strip():
                        existing.contract_end = parse_date(row['contract_end'].strip())
                    updated_count += 1
                else:
                    supplier = Supplier(
                        company_id=company_id,
                        name=name,
                        contact_person=row.get('contact_person', '').strip(),
                        email=email,
                        phone=row.get('phone', '').strip(),
                        address=row.get('address', '').strip(),
                        city=row.get('city', '').strip(),
                        country=row.get('country', '').strip(),
                        category=row.get('category', 'General').strip(),
                        payment_terms=row.get('payment_terms', '').strip(),
                        lead_time_days=int(row.get('lead_time_days', 0) or 0),
                        rating=float(row.get('rating', 0)) if row.get('rating') else None,
                        quality_score=float(row.get('quality_score', 0)) if row.get('quality_score') else None,
                        reliability_score=float(row.get('reliability_score', 0)) if row.get('reliability_score') else None,
                        cost_rating=float(row.get('cost_rating', 0)) if row.get('cost_rating') else None,
                        total_spend=float(row.get('total_spend', 0) or 0),
                        total_orders=int(row.get('total_orders', 0) or 0),
                        is_primary=row.get('is_primary', '').strip().lower() in ['yes', 'true', '1', 'on'],
                        status=row.get('status', 'active').strip(),
                        contract_start=parse_date(row.get('contract_start', '')),
                        contract_end=parse_date(row.get('contract_end', ''))
                    )
                    db.session.add(supplier)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/suppliers/import-excel', methods=['POST'])
@login_required
def import_suppliers_excel():
    company_id = session.get('company_id')

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active

        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_').replace('-', '_')
                headers.append(h)
            else:
                headers.append('')

        header_map = {}
        for idx, h in enumerate(headers):
            if h in ['name', 'supplier_name', 'company_name', 'vendor_name']:
                header_map['name'] = idx
            elif h in ['contact_person', 'contact', 'representative']:
                header_map['contact_person'] = idx
            elif h in ['email', 'email_address', 'e_mail']:
                header_map['email'] = idx
            elif h in ['phone', 'phone_number', 'telephone', 'mobile']:
                header_map['phone'] = idx
            elif h in ['address', 'street', 'street_address']:
                header_map['address'] = idx
            elif h in ['city', 'town']:
                header_map['city'] = idx
            elif h in ['country', 'nation', 'region']:
                header_map['country'] = idx
            elif h in ['category', 'type', 'supplier_category']:
                header_map['category'] = idx
            elif h in ['payment_terms', 'terms', 'payment']:
                header_map['payment_terms'] = idx
            elif h in ['lead_time_days', 'lead_time', 'delivery_days']:
                header_map['lead_time_days'] = idx
            elif h in ['rating', 'score', 'supplier_rating']:
                header_map['rating'] = idx
            elif h in ['quality_score', 'quality']:
                header_map['quality_score'] = idx
            elif h in ['reliability_score', 'reliability']:
                header_map['reliability_score'] = idx
            elif h in ['cost_rating', 'cost_score']:
                header_map['cost_rating'] = idx
            elif h in ['total_spend', 'spend', 'total_spent']:
                header_map['total_spend'] = idx
            elif h in ['total_orders', 'orders', 'order_count']:
                header_map['total_orders'] = idx
            elif h in ['is_primary', 'primary', 'is_primary_supplier']:
                header_map['is_primary'] = idx
            elif h in ['status', 'state', 'supplier_status']:
                header_map['status'] = idx
            elif h in ['contract_start', 'start_date', 'contract_start_date']:
                header_map['contract_start'] = idx
            elif h in ['contract_end', 'end_date', 'contract_end_date']:
                header_map['contract_end'] = idx

        if 'name' not in header_map:
            return jsonify({'success': False, 'error': 'Excel file must have a name column'}), 400

        added_count = 0
        updated_count = 0
        errors = []

        def parse_excel_date(date_val):
            if date_val is None:
                return None
            if isinstance(date_val, datetime):
                return date_val.date()
            elif isinstance(date_val, date):
                return date_val
            else:
                date_str = str(date_val).strip()
                for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y']:
                    try:
                        return datetime.strptime(date_str, fmt).date()
                    except:
                        continue
                return None

        def get_str_val(row, key, default=''):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else default
            return default

        def get_float_val(row, key, default=None):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                if val is not None:
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return default
            return default

        def get_int_val(row, key, default=0):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                if val is not None:
                    try:
                        return int(float(val))
                    except (ValueError, TypeError):
                        return default
            return default

        def get_bool_val(row, key):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                if val is not None:
                    return str(val).strip().lower() in ['yes', 'true', '1', 'on']
            return False

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name_idx = header_map.get('name', 0)
                name = str(row[name_idx] if name_idx < len(row) else '').strip()

                if not name:
                    errors.append(f"Row {row_num}: Missing required field (name)")
                    continue

                email = get_str_val(row, 'email')
                
                # Upsert logic: match by email if present, otherwise by name
                existing = None
                if email:
                    existing = Supplier.query.filter_by(company_id=company_id, email=email).first()
                if not existing and name:
                    existing = Supplier.query.filter_by(company_id=company_id, name=name).first()

                if existing:
                    existing.name = name
                    existing.contact_person = get_str_val(row, 'contact_person') or existing.contact_person
                    existing.email = email or existing.email
                    existing.phone = get_str_val(row, 'phone') or existing.phone
                    existing.address = get_str_val(row, 'address') or existing.address
                    existing.city = get_str_val(row, 'city') or existing.city
                    existing.country = get_str_val(row, 'country') or existing.country
                    existing.category = get_str_val(row, 'category', 'General') or existing.category
                    existing.payment_terms = get_str_val(row, 'payment_terms') or existing.payment_terms
                    existing.lead_time_days = get_int_val(row, 'lead_time_days') or existing.lead_time_days
                    existing.rating = get_float_val(row, 'rating') if header_map.get('rating') is not None and header_map.get('rating') < len(row) and row[header_map.get('rating')] is not None else existing.rating
                    existing.quality_score = get_float_val(row, 'quality_score') if header_map.get('quality_score') is not None and header_map.get('quality_score') < len(row) and row[header_map.get('quality_score')] is not None else existing.quality_score
                    existing.reliability_score = get_float_val(row, 'reliability_score') if header_map.get('reliability_score') is not None and header_map.get('reliability_score') < len(row) and row[header_map.get('reliability_score')] is not None else existing.reliability_score
                    existing.cost_rating = get_float_val(row, 'cost_rating') if header_map.get('cost_rating') is not None and header_map.get('cost_rating') < len(row) and row[header_map.get('cost_rating')] is not None else existing.cost_rating
                    existing.total_spend = get_float_val(row, 'total_spend', 0.0) or existing.total_spend
                    existing.total_orders = get_int_val(row, 'total_orders') or existing.total_orders
                    existing.is_primary = get_bool_val(row, 'is_primary')
                    existing.status = get_str_val(row, 'status', 'active') or existing.status
                    existing.contract_start = parse_excel_date(row[header_map.get('contract_start')] if header_map.get('contract_start') is not None and header_map.get('contract_start') < len(row) else None) or existing.contract_start
                    existing.contract_end = parse_excel_date(row[header_map.get('contract_end')] if header_map.get('contract_end') is not None and header_map.get('contract_end') < len(row) else None) or existing.contract_end
                    updated_count += 1
                else:
                    supplier = Supplier(
                        company_id=company_id,
                        name=name,
                        contact_person=get_str_val(row, 'contact_person'),
                        email=email,
                        phone=get_str_val(row, 'phone'),
                        address=get_str_val(row, 'address'),
                        city=get_str_val(row, 'city'),
                        country=get_str_val(row, 'country'),
                        category=get_str_val(row, 'category', 'General'),
                        payment_terms=get_str_val(row, 'payment_terms'),
                        lead_time_days=get_int_val(row, 'lead_time_days'),
                        rating=get_float_val(row, 'rating'),
                        quality_score=get_float_val(row, 'quality_score'),
                        reliability_score=get_float_val(row, 'reliability_score'),
                        cost_rating=get_float_val(row, 'cost_rating'),
                        total_spend=get_float_val(row, 'total_spend', 0.0),
                        total_orders=get_int_val(row, 'total_orders'),
                        is_primary=get_bool_val(row, 'is_primary'),
                        status=get_str_val(row, 'status', 'active'),
                        contract_start=parse_excel_date(row[header_map.get('contract_start')] if header_map.get('contract_start') is not None and header_map.get('contract_start') < len(row) else None),
                        contract_end=parse_excel_date(row[header_map.get('contract_end')] if header_map.get('contract_end') is not None and header_map.get('contract_end') < len(row) else None)
                    )
                    db.session.add(supplier)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@business_twin_bp.route('/suppliers/export-excel')
@login_required
def export_suppliers_excel():
    company_id = session.get('company_id')
    suppliers = Supplier.query.filter_by(company_id=company_id).order_by(Supplier.created_at.desc()).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ['ID', 'Name', 'Contact Person', 'Email', 'Phone', 'Address', 'City', 'Country',
               'Category', 'Payment Terms', 'Lead Time (days)', 'Rating', 'Quality Score',
               'Reliability Score', 'Cost Rating', 'Total Spend', 'Total Orders', 'Primary',
               'Status', 'Contract Start', 'Contract End', 'Created At']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_num, s in enumerate(suppliers, 2):
        ws.cell(row=row_num, column=1, value=s.id)
        ws.cell(row=row_num, column=2, value=s.name)
        ws.cell(row=row_num, column=3, value=s.contact_person)
        ws.cell(row=row_num, column=4, value=s.email)
        ws.cell(row=row_num, column=5, value=s.phone)
        ws.cell(row=row_num, column=6, value=s.address)
        ws.cell(row=row_num, column=7, value=s.city)
        ws.cell(row=row_num, column=8, value=s.country)
        ws.cell(row=row_num, column=9, value=s.category)
        ws.cell(row=row_num, column=10, value=s.payment_terms)
        ws.cell(row=row_num, column=11, value=s.lead_time_days)
        ws.cell(row=row_num, column=12, value=s.rating)
        ws.cell(row=row_num, column=13, value=s.quality_score)
        ws.cell(row=row_num, column=14, value=s.reliability_score)
        ws.cell(row=row_num, column=15, value=s.cost_rating)
        ws.cell(row=row_num, column=16, value=s.total_spend)
        ws.cell(row=row_num, column=17, value=s.total_orders)
        ws.cell(row=row_num, column=18, value='Yes' if s.is_primary else 'No')
        ws.cell(row=row_num, column=19, value=s.status)
        ws.cell(row=row_num, column=20, value=s.contract_start.strftime('%Y-%m-%d') if s.contract_start else '')
        ws.cell(row=row_num, column=21, value=s.contract_end.strftime('%Y-%m-%d') if s.contract_end else '')
        ws.cell(row=row_num, column=22, value=s.created_at.strftime('%Y-%m-%d %H:%M:%S') if s.created_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='suppliers_export.xlsx')


@business_twin_bp.route('/suppliers/export-pdf')
@login_required
def export_suppliers_pdf():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    suppliers = Supplier.query.filter_by(company_id=company_id).order_by(Supplier.created_at.desc()).all()

    symbol = _get_currency_symbol(company_id)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a1f3a'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph(f"{company.company_name if company else 'Company'} - Supplier Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=15)
    elements.append(Paragraph(f"Total Suppliers: {len(suppliers)} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    elements.append(Spacer(1, 0.1*inch))

    table_data = [['Name', 'Category', 'Lead Time', 'Rating', 'Total Spend', 'Type', 'Status']]
    for s in suppliers:
        table_data.append([
            s.name or '-',
            s.category or 'General',
            f"{s.lead_time_days} days" if s.lead_time_days else '-',
            str(s.rating) if s.rating else '-',
            f"{symbol}{s.total_spend:,.0f}" if s.total_spend else f'{symbol}0',
            'Primary' if s.is_primary else 'Secondary',
            s.status.title()
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f3a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8f9fa')])
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name='suppliers_report.pdf')


@business_twin_bp.route('/suppliers/download-template')
@login_required
def download_supplier_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'name', 'contact_person', 'email', 'phone', 'address', 'city', 'country',
        'category', 'payment_terms', 'lead_time_days', 'rating', 'quality_score',
        'reliability_score', 'cost_rating', 'total_spend', 'total_orders',
        'is_primary', 'status', 'contract_start', 'contract_end'
    ])

    writer.writerow([
        'TechComponents Ltd', 'Ravi Menon', 'ravi@techcomponents.com', '+91-80-1234-5678',
        '42 Electronics Park, Whitefield', 'Bangalore', 'India', 'Hardware', 'Net 30',
        7, 4.5, 4.2, 4.8, 3.5, 450000, 120, 'Yes', 'active', '2024-01-01', '2026-12-31'
    ])
    writer.writerow([
        'CloudNine Services', 'Sarah Johnson', 'sarah@cloudnine.io', '+1-415-555-0199',
        '500 Cloud Way, Suite 300', 'San Francisco', 'USA', 'IT Services', 'Net 15',
        3, 4.8, 4.5, 4.9, 4.0, 320000, 48, 'Yes', 'active', '2023-06-01', '2025-05-31'
    ])
    writer.writerow([
        'Global Logistics Partners', 'Hassan Al-Rashid', 'hassan@glp.com', '+971-4-555-0123',
        'Jebel Ali Free Zone', 'Dubai', 'UAE', 'Logistics', 'Net 45',
        14, 3.9, 3.5, 4.2, 3.0, 180000, 85, 'No', 'active', '2024-03-15', '2025-03-14'
    ])

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='supplier_import_template.csv')


@business_twin_bp.route('/suppliers/download-excel-template')
@login_required
def download_supplier_excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    headers = [
        'name', 'contact_person', 'email', 'phone', 'address', 'city', 'country',
        'category', 'payment_terms', 'lead_time_days', 'rating', 'quality_score',
        'reliability_score', 'cost_rating', 'total_spend', 'total_orders',
        'is_primary', 'status', 'contract_start', 'contract_end'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    sample_data = [
        ['TechComponents Ltd', 'Ravi Menon', 'ravi@techcomponents.com', '+91-80-1234-5678',
         '42 Electronics Park, Whitefield', 'Bangalore', 'India', 'Hardware', 'Net 30',
         7, 4.5, 4.2, 4.8, 3.5, 450000, 120, 'Yes', 'active', '2024-01-01', '2026-12-31'],
        ['CloudNine Services', 'Sarah Johnson', 'sarah@cloudnine.io', '+1-415-555-0199',
         '500 Cloud Way, Suite 300', 'San Francisco', 'USA', 'IT Services', 'Net 15',
         3, 4.8, 4.5, 4.9, 4.0, 320000, 48, 'Yes', 'active', '2023-06-01', '2025-05-31'],
        ['Global Logistics Partners', 'Hassan Al-Rashid', 'hassan@glp.com', '+971-4-555-0123',
         'Jebel Ali Free Zone', 'Dubai', 'UAE', 'Logistics', 'Net 45',
         14, 3.9, 3.5, 4.2, 3.0, 180000, 85, 'No', 'active', '2024-03-15', '2025-03-14']
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='supplier_import_template.xlsx')


# =============================================================================
# INVENTORY TWIN
# =============================================================================

@business_twin_bp.route('/inventory')
@login_required
def inventory():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    items = Inventory.query.filter_by(company_id=company_id, is_active=True).order_by(Inventory.name).limit(200).all()

    total_db_items = len(items)
    total_quantity = sum(item.quantity_on_hand for item in items)
    low_stock_count = len([item for item in items if item.is_low_stock()])
    overstock_count = len([item for item in items if item.is_overstock()])
    total_stock_value = sum(item.stock_value() for item in items)

    return render_template('business_twin/twin_inventory.html', 
                         items=items,
                         total_db_items=total_db_items,
                         total_quantity=total_quantity,
                         low_stock_count=low_stock_count,
                         overstock_count=overstock_count,
                         total_stock_value=total_stock_value,
                         company=company,
                         currency_symbol=_get_currency_symbol(company_id))


@business_twin_bp.route('/inventory/add', methods=['POST'])
@login_required
def add_inventory():
    company_id = session.get('company_id')

    try:
        qoh = int(request.form.get('quantity_on_hand', 0) or 0)
        qres = int(request.form.get('quantity_reserved', 0) or 0)

        item = Inventory(
            company_id=company_id,
            sku=request.form.get('sku', ''),
            name=request.form.get('name', ''),
            description=request.form.get('description', ''),
            category=request.form.get('category', 'General'),
            location=request.form.get('location', ''),
            unit_cost=float(request.form.get('unit_cost', 0) or 0),
            selling_price=float(request.form.get('selling_price', 0) or 0),
            quantity_on_hand=qoh,
            quantity_reserved=qres,
            quantity_available=qoh - qres,
            reorder_point=int(request.form.get('reorder_point', 0) or 0),
            reorder_quantity=int(request.form.get('reorder_quantity', 0) or 0),
            max_stock=int(request.form.get('max_stock', 0) or 0),
            turnover_rate=float(request.form.get('turnover_rate', 0)) if request.form.get('turnover_rate') else None,
            days_on_hand=float(request.form.get('days_on_hand', 0)) if request.form.get('days_on_hand') else None,
            is_active=True
        )

        db.session.add(item)
        db.session.commit()

        return jsonify({'success': True, 'item': item.to_dict()})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/inventory/edit/<int:item_id>', methods=['POST'])
@login_required
def edit_inventory(item_id):
    company_id = session.get('company_id')
    item = Inventory.query.filter_by(id=item_id, company_id=company_id).first()
    if not item:
        return jsonify({'success': False, 'error': 'Item not found'}), 404
    
    try:
        item.sku = request.form.get('sku', item.sku)
        item.name = request.form.get('name', item.name)
        item.description = request.form.get('description', item.description or '')
        item.category = request.form.get('category', item.category or 'General')
        item.location = request.form.get('location', item.location or '')
        item.unit_cost = float(request.form.get('unit_cost', item.unit_cost or 0) or 0)
        item.selling_price = float(request.form.get('selling_price', item.selling_price or 0) or 0)
        
        qoh = int(request.form.get('quantity_on_hand', item.quantity_on_hand or 0) or 0)
        qres = int(request.form.get('quantity_reserved', item.quantity_reserved or 0) or 0)
        item.quantity_on_hand = qoh
        item.quantity_reserved = qres
        item.quantity_available = qoh - qres
        
        item.reorder_point = int(request.form.get('reorder_point', item.reorder_point or 0) or 0)
        item.reorder_quantity = int(request.form.get('reorder_quantity', item.reorder_quantity or 0) or 0)
        item.max_stock = int(request.form.get('max_stock', item.max_stock or 0) or 0)
        
        turnover = request.form.get('turnover_rate', '')
        item.turnover_rate = float(turnover) if turnover and turnover.strip() else item.turnover_rate
        
        days = request.form.get('days_on_hand', '')
        item.days_on_hand = float(days) if days and days.strip() else item.days_on_hand
        
        db.session.commit()
        return jsonify({'success': True, 'item': item.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/inventory/delete/<int:item_id>', methods=['DELETE'])
@login_required
def delete_inventory(item_id):
    company_id = session.get('company_id')
    item = Inventory.query.filter_by(id=item_id, company_id=company_id).first()
    if not item:
        return jsonify({'success': False, 'error': 'Item not found'}), 404
    
    try:
        db.session.delete(item)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Item deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/inventory/delete-all', methods=['DELETE'])
@login_required
def delete_all_inventory():
    company_id = session.get('company_id')
    try:
        count = Inventory.query.filter_by(company_id=company_id).count()
        Inventory.query.filter_by(company_id=company_id).delete()
        db.session.commit()
        return jsonify({'success': True, 'message': f'All {count} inventory items deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/inventory/import-csv', methods=['POST'])
@login_required
def import_inventory_csv():
    company_id = session.get('company_id')

    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400

    try:
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)

        added_count = 0
        updated_count = 0
        errors = []

        for row_num, row in enumerate(csv_reader, start=2):
            try:
                sku = row.get('sku', '').strip()
                name = row.get('name', '').strip()

                if not sku or not name:
                    errors.append(f"Row {row_num}: Missing required fields (sku, name)")
                    continue

                qoh = int(row.get('quantity_on_hand', 0) or 0)
                qres = int(row.get('quantity_reserved', 0) or 0)

                existing = Inventory.query.filter_by(company_id=company_id, sku=sku).first()

                if existing:
                    existing.name = name
                    existing.description = row.get('description', '').strip()
                    existing.category = row.get('category', 'General').strip()
                    existing.location = row.get('location', '').strip()
                    existing.unit_cost = float(row.get('unit_cost', 0) or 0)
                    existing.selling_price = float(row.get('selling_price', 0) or 0)
                    existing.quantity_on_hand = qoh
                    existing.quantity_reserved = qres
                    existing.quantity_available = qoh - qres
                    existing.reorder_point = int(row.get('reorder_point', 0) or 0)
                    existing.reorder_quantity = int(row.get('reorder_quantity', 0) or 0)
                    existing.max_stock = int(row.get('max_stock', 0) or 0)
                    existing.turnover_rate = float(row.get('turnover_rate', 0)) if row.get('turnover_rate') else existing.turnover_rate
                    existing.days_on_hand = float(row.get('days_on_hand', 0)) if row.get('days_on_hand') else existing.days_on_hand
                    existing.is_active = True
                    updated_count += 1
                else:
                    item = Inventory(
                        company_id=company_id,
                        sku=sku,
                        name=name,
                        description=row.get('description', '').strip(),
                        category=row.get('category', 'General').strip(),
                        location=row.get('location', '').strip(),
                        unit_cost=float(row.get('unit_cost', 0) or 0),
                        selling_price=float(row.get('selling_price', 0) or 0),
                        quantity_on_hand=qoh,
                        quantity_reserved=qres,
                        quantity_available=qoh - qres,
                        reorder_point=int(row.get('reorder_point', 0) or 0),
                        reorder_quantity=int(row.get('reorder_quantity', 0) or 0),
                        max_stock=int(row.get('max_stock', 0) or 0),
                        turnover_rate=float(row.get('turnover_rate', 0)) if row.get('turnover_rate') else None,
                        days_on_hand=float(row.get('days_on_hand', 0)) if row.get('days_on_hand') else None,
                        is_active=True
                    )
                    db.session.add(item)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/inventory/import-excel', methods=['POST'])
@login_required
def import_inventory_excel():
    company_id = session.get('company_id')

    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400

    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400

    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active

        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_').replace('-', '_')
                headers.append(h)
            else:
                headers.append('')

        header_map = {}
        for idx, h in enumerate(headers):
            if h in ['sku', 'sku_code', 'item_code', 'product_code']:
                header_map['sku'] = idx
            elif h in ['name', 'item_name', 'product_name', 'item']:
                header_map['name'] = idx
            elif h in ['description', 'desc', 'details']:
                header_map['description'] = idx
            elif h in ['category', 'type', 'item_category']:
                header_map['category'] = idx
            elif h in ['location', 'warehouse', 'storage']:
                header_map['location'] = idx
            elif h in ['unit_cost', 'cost', 'purchase_price', 'buy_price']:
                header_map['unit_cost'] = idx
            elif h in ['selling_price', 'price', 'sale_price', 'retail_price']:
                header_map['selling_price'] = idx
            elif h in ['quantity_on_hand', 'on_hand', 'stock', 'qty_on_hand']:
                header_map['quantity_on_hand'] = idx
            elif h in ['quantity_reserved', 'reserved', 'qty_reserved']:
                header_map['quantity_reserved'] = idx
            elif h in ['reorder_point', 'reorder_pt', 'min_stock']:
                header_map['reorder_point'] = idx
            elif h in ['reorder_quantity', 'reorder_qty', 'order_qty']:
                header_map['reorder_quantity'] = idx
            elif h in ['max_stock', 'maximum_stock', 'max']:
                header_map['max_stock'] = idx
            elif h in ['turnover_rate', 'turnover']:
                header_map['turnover_rate'] = idx
            elif h in ['days_on_hand', 'days_in_stock']:
                header_map['days_on_hand'] = idx

        if 'sku' not in header_map or 'name' not in header_map:
            return jsonify({'success': False, 'error': 'Excel file must have columns: sku, name'}), 400

        added_count = 0
        updated_count = 0
        errors = []

        def get_str_val(row, key, default=''):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else default
            return default

        def get_float_val(row, key, default=0.0):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                if val is not None:
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return default
            return default

        def get_int_val(row, key, default=0):
            idx = header_map.get(key)
            if idx is not None and idx < len(row):
                val = row[idx]
                if val is not None:
                    try:
                        return int(float(val))
                    except (ValueError, TypeError):
                        return default
            return default

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                sku_idx = header_map.get('sku', 0)
                name_idx = header_map.get('name', 1)

                sku = str(row[sku_idx] if sku_idx < len(row) else '').strip()
                name = str(row[name_idx] if name_idx < len(row) else '').strip()

                if not sku or not name:
                    errors.append(f"Row {row_num}: Missing required fields (sku, name)")
                    continue

                qoh = get_int_val(row, 'quantity_on_hand')
                qres = get_int_val(row, 'quantity_reserved')

                existing = Inventory.query.filter_by(company_id=company_id, sku=sku).first()

                if existing:
                    existing.name = name
                    existing.description = get_str_val(row, 'description')
                    existing.category = get_str_val(row, 'category', 'General')
                    existing.location = get_str_val(row, 'location')
                    existing.unit_cost = get_float_val(row, 'unit_cost')
                    existing.selling_price = get_float_val(row, 'selling_price')
                    existing.quantity_on_hand = qoh
                    existing.quantity_reserved = qres
                    existing.quantity_available = qoh - qres
                    existing.reorder_point = get_int_val(row, 'reorder_point')
                    existing.reorder_quantity = get_int_val(row, 'reorder_quantity')
                    existing.max_stock = get_int_val(row, 'max_stock')
                    turnover = get_str_val(row, 'turnover_rate')
                    existing.turnover_rate = float(turnover) if turnover else existing.turnover_rate
                    days = get_str_val(row, 'days_on_hand')
                    existing.days_on_hand = float(days) if days else existing.days_on_hand
                    existing.is_active = True
                    updated_count += 1
                else:
                    item = Inventory(
                        company_id=company_id,
                        sku=sku,
                        name=name,
                        description=get_str_val(row, 'description'),
                        category=get_str_val(row, 'category', 'General'),
                        location=get_str_val(row, 'location'),
                        unit_cost=get_float_val(row, 'unit_cost'),
                        selling_price=get_float_val(row, 'selling_price'),
                        quantity_on_hand=qoh,
                        quantity_reserved=qres,
                        quantity_available=qoh - qres,
                        reorder_point=get_int_val(row, 'reorder_point'),
                        reorder_quantity=get_int_val(row, 'reorder_quantity'),
                        max_stock=get_int_val(row, 'max_stock'),
                        turnover_rate=get_float_val(row, 'turnover_rate') if get_str_val(row, 'turnover_rate') else None,
                        days_on_hand=get_float_val(row, 'days_on_hand') if get_str_val(row, 'days_on_hand') else None,
                        is_active=True
                    )
                    db.session.add(item)
                    added_count += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        db.session.commit()

        return jsonify({
            'success': True, 
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@business_twin_bp.route('/inventory/export-excel')
@login_required
def export_inventory_excel():
    company_id = session.get('company_id')
    items = Inventory.query.filter_by(company_id=company_id, is_active=True).order_by(Inventory.name).all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    headers = ['ID', 'SKU', 'Name', 'Description', 'Category', 'Location', 'Unit Cost',
               'Selling Price', 'On Hand', 'Reserved', 'Available', 'Reorder Point',
               'Reorder Qty', 'Max Stock', 'Stock Value', 'Turnover Rate', 'Days On Hand', 'Created At']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item.id)
        ws.cell(row=row_num, column=2, value=item.sku)
        ws.cell(row=row_num, column=3, value=item.name)
        ws.cell(row=row_num, column=4, value=item.description)
        ws.cell(row=row_num, column=5, value=item.category)
        ws.cell(row=row_num, column=6, value=item.location)
        ws.cell(row=row_num, column=7, value=item.unit_cost)
        ws.cell(row=row_num, column=8, value=item.selling_price)
        ws.cell(row=row_num, column=9, value=item.quantity_on_hand)
        ws.cell(row=row_num, column=10, value=item.quantity_reserved)
        ws.cell(row=row_num, column=11, value=item.quantity_available)
        ws.cell(row=row_num, column=12, value=item.reorder_point)
        ws.cell(row=row_num, column=13, value=item.reorder_quantity)
        ws.cell(row=row_num, column=14, value=item.max_stock)
        ws.cell(row=row_num, column=15, value=item.stock_value())
        ws.cell(row=row_num, column=16, value=item.turnover_rate)
        ws.cell(row=row_num, column=17, value=item.days_on_hand)
        ws.cell(row=row_num, column=18, value=item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '')

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='inventory_export.xlsx')


@business_twin_bp.route('/inventory/export-pdf')
@login_required
def export_inventory_pdf():
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    items = Inventory.query.filter_by(company_id=company_id, is_active=True).order_by(Inventory.name).all()

    symbol = _get_currency_symbol(company_id)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)

    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#1a1f3a'), spaceAfter=20, alignment=TA_CENTER)
    elements.append(Paragraph(f"{company.company_name if company else 'Company'} - Inventory Report", title_style))
    elements.append(Spacer(1, 0.2*inch))

    summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=15)
    elements.append(Paragraph(f"Total Items: {len(items)} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    elements.append(Spacer(1, 0.1*inch))

    table_data = [['SKU', 'Name', 'Category', 'On Hand', 'Available', 'Unit Cost', 'Status', 'Value']]
    for item in items:
        status = 'Low' if item.is_low_stock() else ('Over' if item.is_overstock() else 'OK')
        table_data.append([
            item.sku or '-',
            item.name or '-',
            item.category or 'General',
            str(item.quantity_on_hand),
            str(item.quantity_available),
            f"{symbol}{item.unit_cost:,.2f}" if item.unit_cost else f'{symbol}0',
            status,
            f"{symbol}{item.stock_value():,.0f}" if item.stock_value() else f'{symbol}0'
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f3a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8f9fa')])
    ]))

    elements.append(table)
    doc.build(elements)
    output.seek(0)

    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name='inventory_report.pdf')


@business_twin_bp.route('/inventory/download-template')
@login_required
def download_inventory_csv_template():
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'sku', 'name', 'description', 'category', 'location', 'unit_cost', 'selling_price',
        'quantity_on_hand', 'quantity_reserved', 'reorder_point', 'reorder_quantity', 'max_stock',
        'turnover_rate', 'days_on_hand'
    ])

    writer.writerow([
        'SKU-001', 'MacBook Pro 16"', 'Apple laptop with M3 chip', 'Electronics', 'Warehouse A, Shelf 1',
        1999.00, 2499.00, 45, 5, 10, 20, 100, 8.5, 15
    ])
    writer.writerow([
        'SKU-002', 'Dell UltraSharp Monitor', '27-inch 4K USB-C monitor', 'Electronics', 'Warehouse A, Shelf 2',
        450.00, 599.00, 30, 3, 8, 15, 60, 6.2, 22
    ])
    writer.writerow([
        'SKU-003', 'Steel Rebar Grade 60', 'Construction rebar #4', 'Raw Materials', 'Yard B, Section 3',
        18.50, 24.00, 500, 50, 100, 200, 1000, 12.0, 8
    ])
    writer.writerow([
        'SKU-004', 'A4 Copy Paper (500 sheets)', 'Premium quality office paper', 'Office Supplies', 'Warehouse C, Rack 1',
        3.50, 5.00, 2000, 200, 500, 1000, 5000, 24.0, 4
    ])
    writer.writerow([
        'SKU-005', 'AWS EC2 Reserved Instance', 't3.large 1-year reserved', 'Software License', 'Digital',
        850.00, 1200.00, 10, 0, 2, 5, 50, 3.5, 45
    ])

    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='inventory_import_template.csv')


@business_twin_bp.route('/inventory/download-excel-template')
@login_required
def download_inventory_excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    headers = [
        'sku', 'name', 'description', 'category', 'location', 'unit_cost', 'selling_price',
        'quantity_on_hand', 'quantity_reserved', 'reorder_point', 'reorder_quantity', 'max_stock',
        'turnover_rate', 'days_on_hand'
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    sample_data = [
        ['SKU-001', 'MacBook Pro 16"', 'Apple laptop with M3 chip', 'Electronics', 'Warehouse A, Shelf 1',
         1999.00, 2499.00, 45, 5, 10, 20, 100, 8.5, 15],
        ['SKU-002', 'Dell UltraSharp Monitor', '27-inch 4K USB-C monitor', 'Electronics', 'Warehouse A, Shelf 2',
         450.00, 599.00, 30, 3, 8, 15, 60, 6.2, 22],
        ['SKU-003', 'Steel Rebar Grade 60', 'Construction rebar #4', 'Raw Materials', 'Yard B, Section 3',
         18.50, 24.00, 500, 50, 100, 200, 1000, 12.0, 8],
        ['SKU-004', 'A4 Copy Paper (500 sheets)', 'Premium quality office paper', 'Office Supplies', 'Warehouse C, Rack 1',
         3.50, 5.00, 2000, 200, 500, 1000, 5000, 24.0, 4],
        ['SKU-005', 'AWS EC2 Reserved Instance', 't3.large 1-year reserved', 'Software License', 'Digital',
         850.00, 1200.00, 10, 0, 2, 5, 50, 3.5, 45]
    ]

    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='inventory_import_template.xlsx')