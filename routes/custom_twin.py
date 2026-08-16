from flask import Blueprint, render_template, request, jsonify, session, send_file
from flask_login import login_required, current_user
from database.models.custom_twin import CustomTwin, CustomTwinRecord
from database.models.company import Company
from database.db import db
from sqlalchemy import func
import csv
import io
import json
from datetime import datetime, date
import re

custom_twin_bp = Blueprint('custom_twin', __name__)


# =============================================================================
# CUSTOM TWIN MANAGEMENT
# =============================================================================

@custom_twin_bp.route('/custom-twins')
@login_required
def list_custom_twins():
    """List all custom twins for the company."""
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    
    twins = CustomTwin.query.filter_by(company_id=company_id, is_active=True).order_by(CustomTwin.name).all()
    
    return render_template('business_twin/custom_twins.html',
                         company=company,
                         twins=twins)


@custom_twin_bp.route('/custom-twins/create', methods=['POST'])
@login_required
def create_custom_twin():
    """Create a new custom twin."""
    company_id = session.get('company_id')
    
    try:
        name = request.form.get('name', '').strip()
        if not name:
            return jsonify({'success': False, 'error': 'Twin name is required'}), 400
        
        # Generate slug from name
        slug = re.sub(r'[^\w\s-]', '', name.lower())
        slug = re.sub(r'[-\s]+', '-', slug)
        
        # Check if slug already exists for this company
        existing = CustomTwin.query.filter_by(company_id=company_id, slug=slug).first()
        if existing:
            return jsonify({'success': False, 'error': 'A twin with this name already exists'}), 400
        
        # Parse field definitions
        field_count = int(request.form.get('field_count', 0))
        fields = []
        for i in range(field_count):
            field_name = request.form.get(f'field_name_{i}', '').strip()
            field_label = request.form.get(f'field_label_{i}', '').strip()
            field_type = request.form.get(f'field_type_{i}', 'text')
            field_required = request.form.get(f'field_required_{i}') == 'on'
            field_default = request.form.get(f'field_default_{i}', '').strip()
            
            if field_name and field_label:
                fields.append({
                    'name': field_name,
                    'label': field_label,
                    'type': field_type,
                    'required': field_required,
                    'default': field_default
                })
        
        # Parse stats configuration
        stat_count = int(request.form.get('stat_count', 0))
        stats = []
        for i in range(stat_count):
            stat_name = request.form.get(f'stat_name_{i}', '').strip()
            stat_label = request.form.get(f'stat_label_{i}', '').strip()
            stat_type = request.form.get(f'stat_type_{i}', 'count')
            stat_field = request.form.get(f'stat_field_{i}', '').strip()
            stat_color = request.form.get(f'stat_color_{i}', '#00D4FF')
            
            if stat_name and stat_label:
                stats.append({
                    'name': stat_name,
                    'label': stat_label,
                    'type': stat_type,
                    'field': stat_field,
                    'color': stat_color
                })
        
        twin = CustomTwin(
            company_id=company_id,
            name=name,
            slug=slug,
            description=request.form.get('description', '').strip(),
            icon=request.form.get('icon', 'fa-cube'),
            color=request.form.get('color', '#00D4FF'),
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        twin.set_field_definitions(fields)
        twin.set_stats_config(stats)
        twin.set_table_columns([f['name'] for f in fields[:5]])  # First 5 fields by default
        
        db.session.add(twin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Custom twin "{name}" created successfully',
            'twin': twin.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/edit', methods=['POST'])
@login_required
def edit_custom_twin(twin_id):
    """Edit a custom twin configuration."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        name = request.form.get('name', '').strip()
        if name:
            twin.name = name
            # Update slug if name changed
            slug = re.sub(r'[^\w\s-]', '', name.lower())
            slug = re.sub(r'[-\s]+', '-', slug)
            # Only update slug if it doesn't conflict
            existing = CustomTwin.query.filter_by(company_id=company_id, slug=slug).first()
            if not existing or existing.id == twin.id:
                twin.slug = slug
        
        if request.form.get('description') is not None:
            twin.description = request.form.get('description', '').strip()
        if request.form.get('icon'):
            twin.icon = request.form.get('icon')
        if request.form.get('color'):
            twin.color = request.form.get('color')
        
        # Update field definitions if provided
        field_count = request.form.get('field_count')
        if field_count is not None:
            field_count = int(field_count)
            fields = []
            for i in range(field_count):
                field_name = request.form.get(f'field_name_{i}', '').strip()
                field_label = request.form.get(f'field_label_{i}', '').strip()
                field_type = request.form.get(f'field_type_{i}', 'text')
                field_required = request.form.get(f'field_required_{i}') == 'on'
                field_default = request.form.get(f'field_default_{i}', '').strip()
                
                if field_name and field_label:
                    fields.append({
                        'name': field_name,
                        'label': field_label,
                        'type': field_type,
                        'required': field_required,
                        'default': field_default
                    })
            twin.set_field_definitions(fields)
            twin.set_table_columns([f['name'] for f in fields[:5]])
        
        # Update stats configuration if provided
        stat_count = request.form.get('stat_count')
        if stat_count is not None:
            stat_count = int(stat_count)
            stats = []
            for i in range(stat_count):
                stat_name = request.form.get(f'stat_name_{i}', '').strip()
                stat_label = request.form.get(f'stat_label_{i}', '').strip()
                stat_type = request.form.get(f'stat_type_{i}', 'count')
                stat_field = request.form.get(f'stat_field_{i}', '').strip()
                stat_color = request.form.get(f'stat_color_{i}', '#00D4FF')
                
                if stat_name and stat_label:
                    stats.append({
                        'name': stat_name,
                        'label': stat_label,
                        'type': stat_type,
                        'field': stat_field,
                        'color': stat_color
                    })
            twin.set_stats_config(stats)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Custom twin "{twin.name}" updated successfully',
            'twin': twin.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/delete', methods=['DELETE'])
@login_required
def delete_custom_twin(twin_id):
    """Delete a custom twin and all its records."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        name = twin.name
        db.session.delete(twin)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Custom twin "{name}" deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# CUSTOM TWIN RECORDS
# =============================================================================

@custom_twin_bp.route('/custom-twins/<slug>')
@login_required
def view_custom_twin(slug):
    """View a custom twin and its records."""
    company_id = session.get('company_id')
    company = Company.query.get(company_id)
    
    twin = CustomTwin.query.filter_by(company_id=company_id, slug=slug, is_active=True).first()
    if not twin:
        return render_template('base.html', error='Twin not found'), 404
    
    records = CustomTwinRecord.query.filter_by(twin_id=twin.id, company_id=company_id, is_active=True).order_by(CustomTwinRecord.created_at.desc()).all()
    
    # Calculate stats
    stats = _calculate_stats(twin, records)
    
    return render_template('business_twin/custom_twin_detail.html',
                         company=company,
                         twin=twin,
                         records=records,
                         stats=stats,
                         field_definitions=twin.get_field_definitions(),
                         table_columns=twin.get_table_columns())


def _calculate_stats(twin, records):
    """Calculate statistics for a custom twin."""
    stats_config = twin.get_stats_config()
    results = []
    
    for stat in stats_config:
        value = 0
        stat_type = stat.get('type', 'count')
        field = stat.get('field', '')
        
        if stat_type == 'count':
            value = len(records)
        elif stat_type == 'sum' and field:
            value = sum(float(r.get_data().get(field, 0) or 0) for r in records)
        elif stat_type == 'avg' and field:
            values = [float(r.get_data().get(field, 0) or 0) for r in records if r.get_data().get(field)]
            value = sum(values) / len(values) if values else 0
        elif stat_type == 'min' and field:
            values = [float(r.get_data().get(field, 0) or 0) for r in records if r.get_data().get(field)]
            value = min(values) if values else 0
        elif stat_type == 'max' and field:
            values = [float(r.get_data().get(field, 0) or 0) for r in records if r.get_data().get(field)]
            value = max(values) if values else 0
        
        results.append({
            'name': stat.get('name', ''),
            'label': stat.get('label', ''),
            'value': value,
            'color': stat.get('color', '#00D4FF')
        })
    
    return results


@custom_twin_bp.route('/custom-twins/<int:twin_id>/records', methods=['POST'])
@login_required
def add_custom_record(twin_id):
    """Add a new record to a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        fields = twin.get_field_definitions()
        data = {}
        
        for field in fields:
            field_name = field['name']
            field_type = field.get('type', 'text')
            value = request.form.get(field_name, '').strip()
            
            # Convert value based on field type
            if field_type == 'number':
                value = float(value) if value else 0
            elif field_type == 'integer':
                value = int(float(value)) if value else 0
            elif field_type == 'boolean':
                value = value.lower() in ['true', 'yes', '1', 'on']
            elif field_type == 'date':
                if value:
                    try:
                        value = datetime.strptime(value, '%Y-%m-%d').date().isoformat()
                    except:
                        pass
            
            data[field_name] = value
        
        record = CustomTwinRecord(
            twin_id=twin_id,
            company_id=company_id,
            created_by=current_user.id if hasattr(current_user, 'id') else None
        )
        record.set_data(data)
        
        db.session.add(record)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Record added successfully',
            'record': record.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/records/<int:record_id>', methods=['POST'])
@login_required
def edit_custom_record(record_id):
    """Edit a custom twin record."""
    company_id = session.get('company_id')
    record = CustomTwinRecord.query.filter_by(id=record_id, company_id=company_id).first()
    
    if not record:
        return jsonify({'success': False, 'error': 'Record not found'}), 404
    
    try:
        twin = record.twin
        fields = twin.get_field_definitions()
        data = {}
        
        for field in fields:
            field_name = field['name']
            field_type = field.get('type', 'text')
            value = request.form.get(field_name, '').strip()
            
            # Convert value based on field type
            if field_type == 'number':
                value = float(value) if value else 0
            elif field_type == 'integer':
                value = int(float(value)) if value else 0
            elif field_type == 'boolean':
                value = value.lower() in ['true', 'yes', '1', 'on']
            elif field_type == 'date':
                if value:
                    try:
                        value = datetime.strptime(value, '%Y-%m-%d').date().isoformat()
                    except:
                        pass
            
            data[field_name] = value
        
        record.set_data(data)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Record updated successfully',
            'record': record.to_dict()
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/records/<int:record_id>', methods=['DELETE'])
@login_required
def delete_custom_record(record_id):
    """Delete a custom twin record."""
    company_id = session.get('company_id')
    record = CustomTwinRecord.query.filter_by(id=record_id, company_id=company_id).first()
    
    if not record:
        return jsonify({'success': False, 'error': 'Record not found'}), 404
    
    try:
        db.session.delete(record)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Record deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/records/delete-all', methods=['DELETE'])
@login_required
def delete_all_custom_records(twin_id):
    """Delete all records from a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        count = CustomTwinRecord.query.filter_by(twin_id=twin_id, company_id=company_id).count()
        CustomTwinRecord.query.filter_by(twin_id=twin_id, company_id=company_id).delete()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'All {count} records deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# IMPORT/EXPORT
# =============================================================================

@custom_twin_bp.route('/custom-twins/<int:twin_id>/import-csv', methods=['POST'])
@login_required
def import_custom_csv(twin_id):
    """Import records from CSV to a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    if 'csv_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['csv_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not file.filename.endswith('.csv'):
        return jsonify({'success': False, 'error': 'File must be a CSV'}), 400
    
    try:
        fields = twin.get_field_definitions()
        field_names = [f['name'] for f in fields]
        
        stream = io.StringIO(file.stream.read().decode('UTF-8'), newline=None)
        csv_reader = csv.DictReader(stream)
        
        added_count = 0
        updated_count = 0
        errors = []
        
        for row_num, row in enumerate(csv_reader, start=2):
            try:
                data = {}
                for field in fields:
                    field_name = field['name']
                    field_type = field.get('type', 'text')
                    value = row.get(field_name, '').strip()
                    
                    # Convert value based on field type
                    if field_type == 'number':
                        value = float(value) if value else 0
                    elif field_type == 'integer':
                        value = int(float(value)) if value else 0
                    elif field_type == 'boolean':
                        value = value.lower() in ['true', 'yes', '1', 'on']
                    
                    data[field_name] = value
                
                # Check for existing record by first field (usually an ID or name)
                existing = None
                if field_names and data.get(field_names[0]):
                    existing = CustomTwinRecord.query.filter_by(
                        twin_id=twin_id,
                        company_id=company_id
                    ).filter(
                        CustomTwinRecord.data.contains(json.dumps({field_names[0]: data[field_names[0]]}))
                    ).first()
                
                if existing:
                    existing.set_data(data)
                    updated_count += 1
                else:
                    record = CustomTwinRecord(
                        twin_id=twin_id,
                        company_id=company_id,
                        created_by=current_user.id if hasattr(current_user, 'id') else None
                    )
                    record.set_data(data)
                    db.session.add(record)
                    added_count += 1
                    
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/import-excel', methods=['POST'])
@login_required
def import_custom_excel(twin_id):
    """Import records from Excel to a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    
    file = request.files['excel_file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'success': False, 'error': 'File must be an Excel file (.xlsx or .xls)'}), 400
    
    try:
        from openpyxl import load_workbook
        
        fields = twin.get_field_definitions()
        field_names = [f['name'] for f in fields]
        
        wb = load_workbook(filename=io.BytesIO(file.read()))
        ws = wb.active
        
        # Read headers
        headers = []
        for cell in ws[1]:
            val = cell.value
            if val:
                h = str(val).strip().lower().replace(' ', '_')
                headers.append(h)
            else:
                headers.append('')
        
        # Create header map
        header_map = {}
        for idx, h in enumerate(headers):
            for field in fields:
                if h == field['name'].lower():
                    header_map[field['name']] = idx
                    break
        
        added_count = 0
        updated_count = 0
        errors = []
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {}
                for field in fields:
                    field_name = field['name']
                    field_type = field.get('type', 'text')
                    idx = header_map.get(field_name)
                    
                    if idx is not None and idx < len(row):
                        value = row[idx]
                        if value is not None:
                            value = str(value).strip()
                    else:
                        value = ''
                    
                    # Convert value based on field type
                    if field_type == 'number':
                        if isinstance(value, (int, float)):
                            value = float(value)
                        else:
                            value = float(value) if value else 0
                    elif field_type == 'integer':
                        if isinstance(value, (int, float)):
                            value = int(value)
                        else:
                            value = int(float(value)) if value else 0
                    elif field_type == 'boolean':
                        if isinstance(value, str):
                            value = value.lower() in ['true', 'yes', '1', 'on']
                        else:
                            value = bool(value)
                    
                    data[field_name] = value
                
                # Check for existing
                existing = None
                if field_names and data.get(field_names[0]):
                    existing = CustomTwinRecord.query.filter_by(
                        twin_id=twin_id,
                        company_id=company_id
                    ).filter(
                        CustomTwinRecord.data.contains(json.dumps({field_names[0]: data[field_names[0]]}))
                    ).first()
                
                if existing:
                    existing.set_data(data)
                    updated_count += 1
                else:
                    record = CustomTwinRecord(
                        twin_id=twin_id,
                        company_id=company_id,
                        created_by=current_user.id if hasattr(current_user, 'id') else None
                    )
                    record.set_data(data)
                    db.session.add(record)
                    added_count += 1
                    
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'added': added_count,
            'updated': updated_count,
            'errors': errors,
            'message': f'Import complete: {added_count} new, {updated_count} updated'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/export-excel')
@login_required
def export_custom_excel(twin_id):
    """Export records to Excel from a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        fields = twin.get_field_definitions()
        records = CustomTwinRecord.query.filter_by(twin_id=twin_id, company_id=company_id, is_active=True).all()
        
        wb = Workbook()
        ws = wb.active
        ws.title = twin.name[:31]  # Excel sheet name limit
        
        # Headers
        headers = [f['label'] for f in fields]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Data
        for row_num, record in enumerate(records, 2):
            data = record.get_data()
            for col_num, field in enumerate(fields, 1):
                value = data.get(field['name'], '')
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{twin.slug}_export.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/download-template')
@login_required
def download_custom_template(twin_id):
    """Download CSV template for a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        fields = twin.get_field_definitions()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header row with field names
        headers = [f['name'] for f in fields]
        writer.writerow(headers)
        
        # Sample data row
        sample = []
        for field in fields:
            field_type = field.get('type', 'text')
            if field_type == 'number':
                sample.append('100.00')
            elif field_type == 'integer':
                sample.append('10')
            elif field_type == 'boolean':
                sample.append('true')
            elif field_type == 'date':
                sample.append('2026-01-15')
            else:
                sample.append(f'Sample {field["label"]}')
        writer.writerow(sample)
        
        output.seek(0)
        filename = f"{twin.slug}_template.csv"
        return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                        as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@custom_twin_bp.route('/custom-twins/<int:twin_id>/download-excel-template')
@login_required
def download_custom_excel_template(twin_id):
    """Download Excel template for a custom twin."""
    company_id = session.get('company_id')
    twin = CustomTwin.query.filter_by(id=twin_id, company_id=company_id).first()
    
    if not twin:
        return jsonify({'success': False, 'error': 'Twin not found'}), 404
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        fields = twin.get_field_definitions()
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Headers
        headers = [f['name'] for f in fields]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a1f3a", end_color="1a1f3a", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Sample data row
        for col_num, field in enumerate(fields, 1):
            field_type = field.get('type', 'text')
            if field_type == 'number':
                value = 100.00
            elif field_type == 'integer':
                value = 10
            elif field_type == 'boolean':
                value = 'true'
            elif field_type == 'date':
                value = '2026-01-15'
            else:
                value = f'Sample {field["label"]}'
            ws.cell(row=2, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{twin.slug}_template.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
