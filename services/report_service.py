import os
import csv
import io
from datetime import datetime, timedelta
from database.db import db
from database.models.report import Report
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, ListFlowable, ListItem
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy import func

# ==================== OPTIONAL IMPORTS ====================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# ==================== UNICODE FONT REGISTRATION ====================
_PDF_FONT = 'Helvetica'
_PDF_FONT_BOLD = 'Helvetica-Bold'

def _register_unicode_fonts():
    global _PDF_FONT, _PDF_FONT_BOLD
    candidates = [
        ('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
         'DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
        ('LiberationSans', '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
         'LiberationSans-Bold', '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'),
        ('NotoSans', '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
         'NotoSans-Bold', '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf'),
        ('ArialUnicode', 'C:/Windows/Fonts/arial.ttf',
         'ArialUnicode-Bold', 'C:/Windows/Fonts/arialbd.ttf'),
        ('ArialUnicode', '/Library/Fonts/Arial.ttf',
         'ArialUnicode-Bold', '/Library/Fonts/Arial Bold.ttf'),
        ('ArialUnicode', '/System/Library/Fonts/Supplemental/Arial.ttf',
         'ArialUnicode-Bold', '/System/Library/Fonts/Supplemental/Arial Bold.ttf'),
    ]
    for reg_name, reg_path, bold_name, bold_path in candidates:
        try:
            pdfmetrics.registerFont(TTFont(reg_name, reg_path))
            pdfmetrics.registerFont(TTFont(bold_name, bold_path))
            _PDF_FONT = reg_name
            _PDF_FONT_BOLD = bold_name
            return
        except Exception:
            continue

_register_unicode_fonts()

def _pdf_currency_symbol(data):
    """Return PDF-safe currency symbol. Falls back to ISO code if no Unicode font."""
    cs = data.get('currency_symbol', '$')
    if _PDF_FONT == 'Helvetica':
        iso_map = {
            '$': 'USD', '€': 'EUR', '£': 'GBP', '₹': 'INR', '¥': 'JPY',
            '₩': 'KRW', 'A$': 'AUD', 'C$': 'CAD', 'S$': 'SGD', 'د.إ': 'AED',
            'Other': ''
        }
        return iso_map.get(cs, cs)
    return cs

# Absolute base directories for generated files
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
REPORTS_DIR = os.path.join(BASE_DIR, 'documents', 'reports')
EXPORTS_DIR = os.path.join(BASE_DIR, 'exports')

def _fmt_money(amount, currency_symbol='$'):
    return f"{currency_symbol}{abs(amount):,.0f}"

class ReportService:
    REPORT_TYPE_CONFIG = {
        'executive': {
            'title': 'Executive Summary Report',
            'description': 'High-level strategic overview for C-suite leadership.',
            'sections': ['summary', 'financials', 'health', 'recommendations']
        },
        'board': {
            'title': 'Board of Directors Report',
            'description': 'Governance, compliance, and fiduciary oversight metrics.',
            'sections': ['governance', 'financial_highlights', 'risk_oversight', 'strategic_initiatives']
        },
        'investor': {
            'title': 'Investor Relations Report',
            'description': 'ROI, growth trajectory, valuation indicators, and market position.',
            'sections': ['performance', 'growth_metrics', 'valuation', 'market_position']
        },
        'risk': {
            'title': 'Enterprise Risk Report',
            'description': 'Risk register, heat map data, mitigation status, and compliance.',
            'sections': ['risk_overview', 'risk_register', 'mitigation', 'compliance']
        },
        'growth': {
            'title': 'Growth & Expansion Report',
            'description': 'Customer acquisition, market share, and scaling opportunities.',
            'sections': ['growth_kpis', 'customer_analytics', 'expansion_opps', 'investment_needs']
        },
        'forecast': {
            'title': 'Predictive Forecast Report',
            'description': 'Trend analysis, scenario planning, and confidence intervals.',
            'sections': ['forecast_summary', 'trends', 'scenarios', 'confidence_analysis']
        },
        'financial': {
            'title': 'Detailed Financial Report',
            'description': 'P&L, balance sheet indicators, cash flow, and financial ratios.',
            'sections': ['income_statement', 'cash_flow', 'balance_sheet', 'ratios']
        },
        'operational': {
            'title': 'Operational Efficiency Report',
            'description': 'Throughput, quality scores, supply chain, and process metrics.',
            'sections': ['efficiency', 'supply_chain', 'quality', 'capacity']
        },
        'department': {
            'title': 'Departmental Performance Report',
            'description': 'Headcount, budget vs. actual, and team performance by department.',
            'sections': ['headcount', 'budget_analysis', 'performance', 'salary_breakdown']
        }
    }

    @staticmethod
    def generate_report(company_id, user_id, report_type, name, format_type='pdf', parameters=None, currency_symbol='$'):
        report = Report(
            company_id=company_id,
            user_id=user_id,
            name=name,
            report_type=report_type,
            format=format_type,
            parameters=str(parameters or {})
        )
        db.session.add(report)
        db.session.commit()
        
        try:
            filepath = ReportService.generate_file_for_report(report, company_id, report_type, format_type, parameters, currency_symbol=currency_symbol)
            report.file_path = filepath
            report.last_generated = datetime.utcnow()
            report.summary = f"{ReportService.REPORT_TYPE_CONFIG.get(report_type, {}).get('title', 'Report')} generated successfully on {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}"
            if os.path.exists(filepath):
                report.file_size = os.path.getsize(filepath)
        except Exception as e:
            report.file_path = None
            report.summary = f"Error: {str(e)}"
        
        db.session.commit()
        return report
    
    @staticmethod
    def generate_file_for_report(report, company_id, report_type, format_type=None, parameters=None, currency_symbol='$'):
        """Generate the actual file for an existing report record. Returns absolute file path."""
        if format_type is None:
            format_type = report.format
        if parameters is None:
            parameters = eval(report.parameters) if report.parameters else {}
            
        os.makedirs(REPORTS_DIR, exist_ok=True)
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        
        data = ReportService._get_report_data(company_id, report_type, currency_symbol)
        
        if format_type == 'pdf':
            filepath = ReportService._generate_pdf(report, company_id, report_type, data)
        elif format_type in ('excel', 'xlsx'):
            filepath = ReportService._generate_excel(report, company_id, report_type, data)
        elif format_type == 'csv':
            filepath = ReportService._generate_csv(report, company_id, report_type, data)
        else:
            filepath = ReportService._generate_pdf(report, company_id, report_type, data)
        
        return filepath

    # ==================== DATA COLLECTION ====================
    
    @staticmethod
    def _get_report_data(company_id, report_type, currency_symbol='$'):
        """Fetch all relevant data for a specific report type."""
        from database.models.company import Company
        from database.models.financial import FinancialRecord
        from database.models.employee import Employee
        from database.models.customer import Customer
        from database.models.inventory import Inventory
        from database.models.risk import Risk
        from database.models.forecast import Forecast, ForecastResult
        from database.models.simulation import Simulation
        from database.models.department import Department
        
        company = Company.query.get(company_id)
        now = datetime.utcnow()
        start_30 = (now - timedelta(days=30)).date()
        start_90 = (now - timedelta(days=90)).date()
        start_365 = (now - timedelta(days=365)).date()
        
        # Base financials
        rev_30 = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.transaction_type == 'revenue',
            FinancialRecord.transaction_date >= start_30
        ).scalar() or 0
        
        exp_30 = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.transaction_type == 'expense',
            FinancialRecord.transaction_date >= start_30
        ).scalar() or 0
        
        rev_90 = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.transaction_type == 'revenue',
            FinancialRecord.transaction_date >= start_90
        ).scalar() or 0
        
        exp_90 = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.transaction_type == 'expense',
            FinancialRecord.transaction_date >= start_90
        ).scalar() or 0
        
        rev_365 = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.transaction_type == 'revenue',
            FinancialRecord.transaction_date >= start_365
        ).scalar() or 0
        
        exp_365 = db.session.query(func.sum(FinancialRecord.amount)).filter(
            FinancialRecord.company_id == company_id,
            FinancialRecord.transaction_type == 'expense',
            FinancialRecord.transaction_date >= start_365
        ).scalar() or 0
        
        annual_revenue = company.annual_revenue or rev_365 or 1000000
        annual_cost = exp_365 or (annual_revenue * 0.75)
        profit_30 = rev_30 - exp_30
        profit_365 = rev_365 - exp_365
        margin = (profit_365 / annual_revenue * 100) if annual_revenue else 0
        
        # Employees
        total_employees = db.session.query(func.count(Employee.id)).filter(
            Employee.company_id == company_id, Employee.status == 'active'
        ).scalar() or 0
        
        avg_salary = db.session.query(func.avg(Employee.salary)).filter(
            Employee.company_id == company_id, Employee.status == 'active', Employee.salary.isnot(None)
        ).scalar() or 50000
        
        departments = db.session.query(Department).filter(Department.company_id == company_id).all()
        dept_data = []
        for d in departments:
            count = db.session.query(func.count(Employee.id)).filter(
                Employee.department_id == d.id, Employee.status == 'active'
            ).scalar() or 0
            dept_salary = db.session.query(func.sum(Employee.salary)).filter(
                Employee.department_id == d.id, Employee.status == 'active'
            ).scalar() or 0
            dept_data.append({
                'name': d.name,
                'headcount': count,
                'payroll': float(dept_salary),
                'budget': float(d.budget or dept_salary * 1.2)
            })
        
        if not dept_data:
            dept_data = [
                {'name': 'Sales', 'headcount': max(1, total_employees // 4), 'payroll': 0, 'budget': 0},
                {'name': 'Marketing', 'headcount': max(1, total_employees // 5), 'payroll': 0, 'budget': 0},
                {'name': 'Operations', 'headcount': max(1, total_employees // 3), 'payroll': 0, 'budget': 0},
                {'name': 'Engineering', 'headcount': max(1, total_employees // 4), 'payroll': 0, 'budget': 0}
            ]
        
        # Customers
        total_customers = db.session.query(func.count(Customer.id)).filter(
            Customer.company_id == company_id, Customer.status == 'active'
        ).scalar() or 0
        
        churned = db.session.query(func.count(Customer.id)).filter(
            Customer.company_id == company_id, Customer.is_churned == True
        ).scalar() or 0
        
        churn_rate = (churned / max(total_customers + churned, 1) * 100)
        new_customers_30 = db.session.query(func.count(Customer.id)).filter(
            Customer.company_id == company_id,
            Customer.created_at >= (now - timedelta(days=30))
        ).scalar() or 0
        
        # Inventory
        inventory_items = Inventory.query.filter_by(company_id=company_id, is_active=True).all()
        total_inventory_value = sum((i.quantity_on_hand or 0) * (i.unit_cost or 0) for i in inventory_items)
        low_stock_count = sum(1 for i in inventory_items if getattr(i, 'is_low_stock', lambda: False)())
        total_skus = len(inventory_items)
        
        # Risks
        risks = Risk.query.filter_by(company_id=company_id).all()
        risk_data = [r.to_dict() for r in risks] if risks else []
        avg_risk_score = sum(r.risk_score for r in risks) / len(risks) if risks else 0
        critical_risks = len([r for r in risks if r.risk_level == 'critical'])
        high_risks = len([r for r in risks if r.risk_level == 'high'])
        
        # Forecasts
        recent_forecasts = Forecast.query.filter_by(company_id=company_id).order_by(
            Forecast.created_at.desc()
        ).limit(5).all()
        forecast_data = []
        for f in recent_forecasts:
            results = ForecastResult.query.filter_by(forecast_id=f.id).order_by(ForecastResult.date).all()
            if results:
                first = results[0].value
                last = results[-1].value
                forecast_data.append({
                    'name': f.name,
                    'type': f.forecast_type,
                    'method': f.method,
                    'horizon': f.horizon_days,
                    'trend': ((last - first) / max(abs(first), 1) * 100),
                    'final_value': last
                })
        
        # Simulations
        recent_sims = Simulation.query.filter_by(company_id=company_id).order_by(
            Simulation.created_at.desc()
        ).limit(5).all()
        sim_data = [{
            'name': s.name,
            'type': s.sim_type,
            'revenue_before': s.revenue_before,
            'revenue_after': s.revenue_after,
            'profit_before': s.profit_before,
            'profit_after': s.profit_after,
            'risk_score': s.risk_score
        } for s in recent_sims]
        
        # Health score calculation
        health_score = 50
        if margin > 20: health_score += 15
        elif margin > 10: health_score += 5
        elif margin > 0: health_score -= 5
        else: health_score -= 20
        
        if churn_rate < 5: health_score += 15
        elif churn_rate < 10: health_score += 5
        elif churn_rate < 20: health_score -= 5
        else: health_score -= 15
        
        if critical_risks == 0: health_score += 10
        if high_risks == 0: health_score += 5
        if total_employees > 10: health_score += 5
        health_score = min(max(health_score, 0), 100)
        
        # Ratios
        current_ratio = 1.5
        roe = (profit_365 / max(annual_revenue * 0.3, 1) * 100) if annual_revenue else 0
        burn_rate = exp_30
        runway_months = (annual_revenue - annual_cost) / max(burn_rate, 1) if burn_rate > 0 else 12
        
        return {
            'company': company,
            'company_name': company.company_name if company else 'Business',
            'industry': company.industry if company else 'General',
            'annual_revenue': float(annual_revenue),
            'annual_cost': float(annual_cost),
            'profit_30': float(profit_30),
            'profit_365': float(profit_365),
            'margin': float(margin),
            'rev_30': float(rev_30),
            'exp_30': float(exp_30),
            'rev_90': float(rev_90),
            'exp_90': float(exp_90),
            'rev_365': float(rev_365),
            'exp_365': float(exp_365),
            'total_employees': total_employees,
            'avg_salary': float(avg_salary),
            'departments': dept_data,
            'total_customers': total_customers,
            'churn_rate': float(churn_rate),
            'new_customers_30': new_customers_30,
            'total_inventory_value': float(total_inventory_value),
            'low_stock_count': low_stock_count,
            'total_skus': total_skus,
            'risks': risk_data,
            'avg_risk_score': float(avg_risk_score),
            'critical_risks': critical_risks,
            'high_risks': high_risks,
            'forecasts': forecast_data,
            'simulations': sim_data,
            'health_score': health_score,
            'current_ratio': current_ratio,
            'roe': float(roe),
            'burn_rate': float(burn_rate),
            'runway_months': float(runway_months),
            'generated_at': datetime.utcnow().strftime('%B %d, %Y at %H:%M UTC'),
            'currency_symbol': currency_symbol
        }

    # ==================== PDF GENERATOR ====================
    
    @staticmethod
    def _generate_pdf(report, company_id, report_type, data):
        config = ReportService.REPORT_TYPE_CONFIG.get(report_type, ReportService.REPORT_TYPE_CONFIG['executive'])
        filename = os.path.join(REPORTS_DIR, f"{report_type}_{report.id}.pdf")
        
        doc = SimpleDocTemplate(filename, pagesize=A4,
                               rightMargin=54, leftMargin=54,
                               topMargin=54, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        
        # Custom styles with Unicode-safe font
        title_style = ParagraphStyle(
            'CustomTitle', parent=styles['Heading1'],
            fontSize=22, textColor=colors.HexColor('#1e293b'),
            spaceAfter=6, alignment=TA_CENTER, fontName=_PDF_FONT_BOLD
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle', parent=styles['Normal'],
            fontSize=11, textColor=colors.HexColor('#64748b'),
            alignment=TA_CENTER, spaceAfter=20, fontName=_PDF_FONT
        )
        heading2 = ParagraphStyle(
            'CustomH2', parent=styles['Heading2'],
            fontSize=14, textColor=colors.HexColor('#1e293b'),
            spaceAfter=10, spaceBefore=16, fontName=_PDF_FONT_BOLD,
            borderColor=colors.HexColor('#e2e8f0'), borderWidth=1, borderPadding=5,
            leftIndent=-5, backColor=colors.HexColor('#f8fafc')
        )
        normal = ParagraphStyle(
            'CustomNormal', parent=styles['Normal'],
            fontSize=10, textColor=colors.HexColor('#334155'), leading=14, fontName=_PDF_FONT
        )
        metric_style = ParagraphStyle(
            'Metric', parent=styles['Normal'],
            fontSize=11, textColor=colors.HexColor('#0f172a'), fontName=_PDF_FONT_BOLD
        )
        
        elements = []
        
        # HEADER
        elements.append(Paragraph(f"{data['company_name']}", title_style))
        elements.append(Paragraph(f"{config['title']}", subtitle_style))
        elements.append(Paragraph(f"Generated: {data['generated_at']} &nbsp;|&nbsp; Industry: {data['industry']}", subtitle_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # TYPE-SPECIFIC CONTENT
        if report_type == 'executive':
            ReportService._pdf_executive(elements, data, heading2, normal, metric_style)
        elif report_type == 'board':
            ReportService._pdf_board(elements, data, heading2, normal, metric_style)
        elif report_type == 'investor':
            ReportService._pdf_investor(elements, data, heading2, normal, metric_style)
        elif report_type == 'risk':
            ReportService._pdf_risk(elements, data, heading2, normal, metric_style)
        elif report_type == 'growth':
            ReportService._pdf_growth(elements, data, heading2, normal, metric_style)
        elif report_type == 'forecast':
            ReportService._pdf_forecast(elements, data, heading2, normal, metric_style)
        elif report_type == 'financial':
            ReportService._pdf_financial(elements, data, heading2, normal, metric_style)
        elif report_type == 'operational':
            ReportService._pdf_operational(elements, data, heading2, normal, metric_style)
        elif report_type == 'department':
            ReportService._pdf_department(elements, data, heading2, normal, metric_style)
        else:
            ReportService._pdf_executive(elements, data, heading2, normal, metric_style)
        
        doc.build(elements)
        return filename

    @staticmethod
    def _pdf_table(elements, headers, rows, col_widths=None):
        """Helper to create a styled table."""
        table_data = [headers] + rows
        if not col_widths:
            col_widths = [3*inch] * len(headers)
            if len(headers) == 2:
                col_widths = [3.5*inch, 3.5*inch]
            elif len(headers) == 3:
                col_widths = [2.5*inch, 2.5*inch, 2.5*inch]
            elif len(headers) >= 4:
                total = 7 * inch
                col_widths = [total / len(headers)] * len(headers)
        
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), _PDF_FONT_BOLD),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ffffff')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('FONTNAME', (0, 1), (-1, -1), _PDF_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#334155')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#ffffff'), colors.HexColor('#f8fafc')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.15*inch))

    @staticmethod
    def _pdf_executive(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Executive Summary", h2))
        elements.append(Paragraph(
            f"This report provides a strategic overview of <b>{data['company_name']}</b>. "
            f"The organization operates in the <b>{data['industry']}</b> sector with an annual revenue run-rate of "
            f"<b>{cs}{data['annual_revenue']:,.0f}</b> and a net margin of <b>{data['margin']:.1f}%</b>. "
            f"Overall organizational health is rated <b>{data['health_score']:.0f}/100</b>.",
            normal
        ))
        elements.append(Spacer(1, 0.1*inch))
        
        ReportService._pdf_table(elements, 
            ['Metric', 'Value'],
            [
                ['Annual Revenue Run-Rate', f"{cs}{data['annual_revenue']:,.0f}"],
                ['Annual Cost Base', f"{cs}{data['annual_cost']:,.0f}"],
                ['Net Profit (LTM)', f"{cs}{data['profit_365']:,.0f}"],
                ['Net Margin', f"{data['margin']:.1f}%"],
                ['Health Score', f"{data['health_score']:.0f}/100"],
                ['Active Employees', f"{data['total_employees']}"],
                ['Active Customers', f"{data['total_customers']}"],
                ['Customer Churn Rate', f"{data['churn_rate']:.1f}%"],
                ['Inventory SKUs', f"{data['total_skus']}"],
                ['Low Stock Alerts', f"{data['low_stock_count']}"],
            ]
        )
        
        elements.append(Paragraph("30-Day Financial Pulse", h2))
        elements.append(Paragraph(
            f"Recent trailing-30-day performance shows revenue of <b>{cs}{data['rev_30']:,.0f}</b> "
            f"against expenses of <b>{cs}{data['exp_30']:,.0f}</b>, yielding a short-term profit of "
            f"<b>{cs}{data['profit_30']:,.0f}</b>.",
            normal
        ))
        elements.append(Spacer(1, 0.1*inch))
        
        ReportService._pdf_table(elements,
            ['Period', 'Revenue', 'Expenses', 'Profit'],
            [
                ['Last 30 Days', f"{cs}{data['rev_30']:,.0f}", f"{cs}{data['exp_30']:,.0f}", f"{cs}{data['profit_30']:,.0f}"],
                ['Last 90 Days', f"{cs}{data['rev_90']:,.0f}", f"{cs}{data['exp_90']:,.0f}", f"{cs}{data['rev_90'] - data['exp_90']:,.0f}"],
                ['Last 12 Months', f"{cs}{data['rev_365']:,.0f}", f"{cs}{data['exp_365']:,.0f}", f"{cs}{data['profit_365']:,.0f}"],
            ],
            [1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch]
        )
        
        elements.append(Paragraph("Strategic Recommendations", h2))
        recs = ReportService._get_recommendations(data, 'executive')
        for rec in recs:
            elements.append(Paragraph(f"• {rec}", normal))
            elements.append(Spacer(1, 0.05*inch))

    @staticmethod
    def _pdf_board(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Governance & Fiduciary Overview", h2))
        elements.append(Paragraph(
            f"Presented to the Board of Directors for <b>{data['company_name']}</b>. "
            f"This document covers financial stewardship, risk oversight, and strategic alignment.",
            normal
        ))
        
        elements.append(Paragraph("Financial Highlights", h2))
        ReportService._pdf_table(elements,
            ['Metric', 'Value', 'Board Note'],
            [
                ['Annual Revenue', f"{cs}{data['annual_revenue']:,.0f}", 'Verify against audited statements'],
                ['Annual Costs', f"{cs}{data['annual_cost']:,.0f}", 'Review cost containment policies'],
                ['Net Margin', f"{data['margin']:.1f}%", 'Benchmark vs. industry peers'],
                ['Cash Burn (30d)', f"{cs}{data['burn_rate']:,.0f}", 'Assess runway adequacy'],
                ['Est. Runway', f"{data['runway_months']:.1f} months", 'Approve funding if <6 months'],
            ],
            [2*inch, 2*inch, 3*inch]
        )
        
        elements.append(Paragraph("Risk Oversight", h2))
        elements.append(Paragraph(
            f"The enterprise currently tracks <b>{len(data['risks'])}</b> active risks with an average score of "
            f"<b>{data['avg_risk_score']:.1f}</b>. Critical risks: <b>{data['critical_risks']}</b>. "
            f"High risks: <b>{data['high_risks']}</b>.",
            normal
        ))
        if data['risks']:
            risk_rows = []
            for r in data['risks'][:8]:
                risk_rows.append([
                    r.get('name', 'Unknown'),
                    r.get('category', 'N/A').replace('_', ' ').title(),
                    f"{r.get('risk_score', 0):.0f}",
                    r.get('risk_level', 'N/A').title()
                ])
            ReportService._pdf_table(elements,
                ['Risk Name', 'Category', 'Score', 'Level'],
                risk_rows,
                [2.5*inch, 2*inch, 1.5*inch, 1.5*inch]
            )
        
        elements.append(Paragraph("Strategic Initiatives", h2))
        sims = data.get('simulations', [])
        if sims:
            for s in sims[:3]:
                rev_impact = (s['revenue_after'] or 0) - (s['revenue_before'] or 0)
                elements.append(Paragraph(
                    f"• <b>{s['name']}</b> ({s['type']}): Revenue impact {rev_impact:+,.0f}, "
                    f"Risk score {s['risk_score']:.0f}",
                    normal
                ))
        else:
            elements.append(Paragraph("• No recent simulations on record. Recommend quarterly scenario planning.", normal))

    @staticmethod
    def _pdf_investor(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Investor Relations Summary", h2))
        elements.append(Paragraph(
            f"<b>{data['company_name']}</b> — Investment performance and growth trajectory for the current fiscal period.",
            normal
        ))
        
        elements.append(Paragraph("Performance Metrics", h2))
        ReportService._pdf_table(elements,
            ['KPI', 'Value', 'Investor Relevance'],
            [
                ['LTM Revenue', f"{cs}{data['rev_365']:,.0f}", 'Top-line growth indicator'],
                ['LTM Profit', f"{cs}{data['profit_365']:,.0f}", 'Bottom-line profitability'],
                ['Net Margin', f"{data['margin']:.1f}%", 'Unit economics efficiency'],
                ['Customer Base', f"{data['total_customers']}", 'Market penetration proxy'],
                ['Monthly Churn', f"{data['churn_rate']:.1f}%", 'Retention & LTV health'],
                ['New Customers (30d)', f"{data['new_customers_30']}", 'Growth velocity signal'],
                ['Return on Equity (Est.)', f"{data['roe']:.1f}%", 'Capital efficiency'],
            ],
            [2*inch, 2*inch, 3*inch]
        )
        
        elements.append(Paragraph("Growth Trajectory", h2))
        if data['forecasts']:
            for f in data['forecasts'][:3]:
                direction = "upward" if f['trend'] > 0 else "downward"
                elements.append(Paragraph(
                    f"• <b>{f['name']}</b> ({f['type']}) projects a <b>{direction}</b> trend of "
                    f"<b>{abs(f['trend']):.1f}%</b> over {f['horizon']} days using {f['method']}.",
                    normal
                ))
        else:
            elements.append(Paragraph("• No active forecasts. Recommend initiating predictive modeling for investor guidance.", normal))
        
        elements.append(Paragraph("Valuation Indicators", h2))
        rev_multiple = 3.0 if data['margin'] > 15 else 2.0 if data['margin'] > 5 else 1.5
        estimated_valuation = data['annual_revenue'] * rev_multiple
        elements.append(Paragraph(
            f"Based on a revenue multiple of <b>{rev_multiple:.1f}x</b> (adjusted for margin profile), "
            f"estimated valuation range is <b>{cs}{estimated_valuation * 0.8:,.0f}</b> — "
            f"<b>{cs}{estimated_valuation * 1.2:,.0f}</b>.",
            normal
        ))
        
        elements.append(Paragraph("Investor Recommendations", h2))
        for rec in ReportService._get_recommendations(data, 'investor'):
            elements.append(Paragraph(f"• {rec}", normal))
            elements.append(Spacer(1, 0.05*inch))

    @staticmethod
    def _pdf_risk(elements, data, h2, normal, metric):
        elements.append(Paragraph("Enterprise Risk Assessment", h2))
        elements.append(Paragraph(
            f"Comprehensive risk register for <b>{data['company_name']}</b>. "
            f"Aggregate risk exposure: <b>{data['avg_risk_score']:.1f}/100</b>.",
            normal
        ))
        
        elements.append(Paragraph("Risk Distribution", h2))
        ReportService._pdf_table(elements,
            ['Severity', 'Count', 'Threshold', 'Action Required'],
            [
                ['Critical', str(data['critical_risks']), 'Score ≥ 70', 'Immediate board escalation'],
                ['High', str(data['high_risks']), 'Score 40-69', 'Executive mitigation within 30 days'],
                ['Medium', str(len([r for r in data['risks'] if r.get('risk_level') == 'medium'])), 'Score 20-39', 'Manager monitoring quarterly'],
                ['Low', str(len([r for r in data['risks'] if r.get('risk_level') == 'low'])), 'Score < 20', 'Standard controls'],
            ],
            [1.8*inch, 1.2*inch, 1.8*inch, 2.7*inch]
        )
        
        elements.append(Paragraph("Detailed Risk Register", h2))
        if data['risks']:
            risk_rows = []
            for r in data['risks'][:12]:
                risk_rows.append([
                    r.get('name', 'Unknown')[:30],
                    r.get('category', 'N/A').replace('_', ' ').title()[:15],
                    f"{r.get('probability', 0)*100:.0f}%",
                    f"{r.get('impact', 0)*100:.0f}%",
                    f"{r.get('risk_score', 0):.0f}",
                    r.get('risk_level', 'N/A').title()
                ])
            ReportService._pdf_table(elements,
                ['Risk', 'Category', 'Probability', 'Impact', 'Score', 'Level'],
                risk_rows,
                [1.8*inch, 1.3*inch, 1.1*inch, 1*inch, 0.9*inch, 1*inch]
            )
        else:
            elements.append(Paragraph("No risks registered. Recommend immediate risk assessment.", normal))
        
        elements.append(Paragraph("Mitigation Recommendations", h2))
        for rec in ReportService._get_recommendations(data, 'risk'):
            elements.append(Paragraph(f"• {rec}", normal))
            elements.append(Spacer(1, 0.05*inch))

    @staticmethod
    def _pdf_growth(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Growth & Expansion Analysis", h2))
        elements.append(Paragraph(
            f"Growth performance and scaling readiness for <b>{data['company_name']}</b>.",
            normal
        ))
        
        elements.append(Paragraph("Growth KPIs", h2))
        cac = 150 if data['new_customers_30'] > 0 else 0
        ltv = (data['annual_revenue'] / max(data['total_customers'], 1)) * 2.5
        ReportService._pdf_table(elements,
            ['Growth Metric', 'Value', 'Benchmark'],
            [
                ['New Customers (30d)', str(data['new_customers_30']), 'Target: >5% of base monthly'],
                ['Total Customer Base', str(data['total_customers']), 'Retention-focused growth'],
                ['Churn Rate', f"{data['churn_rate']:.1f}%", 'Best-in-class: <5% annually'],
                ['Est. CAC', f"{cs}{cac:,.0f}" if cac else "N/A", 'Target: <1/3 of LTV'],
                ['Est. LTV', f"{cs}{ltv:,.0f}", 'Higher is better'],
                ['LTV/CAC Ratio', f"{ltv/max(cac,1):.1f}x" if cac else "N/A", 'Target: >3.0x'],
                ['Revenue/Employee', f"{cs}{data['annual_revenue']/max(data['total_employees'],1):,.0f}", 'Efficiency indicator'],
            ],
            [2.2*inch, 2.2*inch, 2.8*inch]
        )
        
        elements.append(Paragraph("Expansion Opportunities", h2))
        elements.append(Paragraph(
            f"With <b>{data['total_employees']}</b> employees and <b>{data['total_customers']}</b> customers, "
            f"the organization has a revenue-per-employee ratio of <b>{cs}{data['annual_revenue']/max(data['total_employees'],1):,.0f}</b>. "
            f"Inventory coverage stands at <b>{data['total_skus']}</b> SKUs with "
            f"<b>{data['low_stock_count']}</b> low-stock alerts.",
            normal
        ))
        
        elements.append(Paragraph("Investment Requirements", h2))
        if data['runway_months'] < 6:
            elements.append(Paragraph(
                f"• <b>Funding Alert:</b> Estimated runway is only <b>{data['runway_months']:.1f} months</b>. "
                f"Secure additional capital or reduce burn rate.", normal
            ))
        if data['low_stock_count'] > 0:
            elements.append(Paragraph(
                f"• <b>Inventory Investment:</b> {data['low_stock_count']} items require replenishment "
                f"to avoid revenue leakage.", normal
            ))
        if data['churn_rate'] > 10:
            elements.append(Paragraph(
                f"• <b>Retention Investment:</b> Churn rate of {data['churn_rate']:.1f}% suggests "
                f"customer success program funding is needed.", normal
            ))
        if data['runway_months'] >= 6 and data['low_stock_count'] == 0 and data['churn_rate'] <= 10:
            elements.append(Paragraph("• Growth profile is stable. Consider market expansion or product line extension.", normal))

    @staticmethod
    def _pdf_forecast(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Predictive Forecast & Scenario Analysis", h2))
        elements.append(Paragraph(
            f"Forward-looking analysis for <b>{data['company_name']}</b> based on recent trends and model projections.",
            normal
        ))
        
        elements.append(Paragraph("Active Forecasts", h2))
        if data['forecasts']:
            for f in data['forecasts']:
                direction = "↑ Growth" if f['trend'] > 5 else "↓ Decline" if f['trend'] < -5 else "→ Stable"
                color = "green" if f['trend'] > 5 else "red" if f['trend'] < -5 else "gray"
                elements.append(Paragraph(
                    f"• <b>{f['name']}</b> | Method: {f['method']} | Horizon: {f['horizon']} days<br/>"
                    f"&nbsp;&nbsp;Projected trend: <b>{direction}</b> ({f['trend']:+.1f}%) | "
                    f"Final projected value: <b>{cs}{f['final_value']:,.0f}</b>",
                    normal
                ))
                elements.append(Spacer(1, 0.08*inch))
        else:
            elements.append(Paragraph("No forecasts available. Generate forecasts in the Forecasting Lab to populate this section.", normal))
        
        elements.append(Paragraph("Scenario Planning", h2))
        sims = data.get('simulations', [])
        if sims:
            for s in sims[:4]:
                rev_impact = (s['revenue_after'] or 0) - (s['revenue_before'] or 0)
                profit_impact = (s['profit_after'] or 0) - (s['profit_before'] or 0)
                elements.append(Paragraph(
                    f"• <b>{s['name']}</b> ({s['type']})<br/>"
                    f"&nbsp;&nbsp;Revenue impact: {rev_impact:+,.0f} | Profit impact: {profit_impact:+,.0f} | Risk: {s['risk_score']:.0f}",
                    normal
                ))
                elements.append(Spacer(1, 0.05*inch))
        else:
            elements.append(Paragraph("No simulation scenarios run. Use the Simulation Center to model 'what-if' situations.", normal))
        
        elements.append(Paragraph("Confidence & Methodology Notes", h2))
        elements.append(Paragraph(
            "Forecasts are generated using statistical models (Prophet, ARIMA, Moving Average) and should be "
            "interpreted with appropriate confidence intervals. Always validate model assumptions against "
            "qualitative market intelligence before making capital allocation decisions.",
            normal
        ))

    @staticmethod
    def _pdf_financial(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Detailed Financial Statement", h2))
        elements.append(Paragraph(
            f"Comprehensive financial analysis for <b>{data['company_name']}</b>. "
            f"All figures in company currency unless otherwise noted.",
            normal
        ))
        
        elements.append(Paragraph("Income Statement (Trailing 12 Months)", h2))
        ReportService._pdf_table(elements,
            ['Line Item', 'Amount', '% of Revenue'],
            [
                ['Revenue', f"{cs}{data['rev_365']:,.0f}", '100.0%'],
                ['Cost of Revenue (Est.)', f"{cs}{data['exp_365'] * 0.6:,.0f}", f"{(data['exp_365']*0.6/max(data['rev_365'],1)*100):.1f}%"],
                ['Gross Profit (Est.)', f"{cs}{data['profit_365'] + (data['exp_365']*0.4):,.0f}", f"{((data['profit_365'] + data['exp_365']*0.4)/max(data['rev_365'],1)*100):.1f}%"],
                ['Operating Expenses', f"{cs}{data['exp_365']:,.0f}", f"{(data['exp_365']/max(data['rev_365'],1)*100):.1f}%"],
                ['Net Profit', f"{cs}{data['profit_365']:,.0f}", f"{(data['profit_365']/max(data['rev_365'],1)*100):.1f}%"],
            ],
            [2.5*inch, 2*inch, 2.5*inch]
        )
        
        elements.append(Paragraph("Cash Flow Indicators", h2))
        ReportService._pdf_table(elements,
            ['Indicator', 'Value', 'Assessment'],
            [
                ['Monthly Burn Rate', f"{cs}{data['burn_rate']:,.0f}", 'High' if data['burn_rate'] > data['rev_30'] else 'Moderate' if data['burn_rate'] > data['rev_30']*0.5 else 'Low'],
                ['Cash Runway (Est.)', f"{data['runway_months']:.1f} months", 'Critical' if data['runway_months'] < 3 else 'Caution' if data['runway_months'] < 6 else 'Healthy'],
                ['Revenue (30d)', f"{cs}{data['rev_30']:,.0f}", 'Current liquidity inflow'],
                ['Expenses (30d)', f"{cs}{data['exp_30']:,.0f}", 'Current liquidity outflow'],
            ],
            [2.2*inch, 2.2*inch, 2.8*inch]
        )
        
        elements.append(Paragraph("Financial Ratios", h2))
        ReportService._pdf_table(elements,
            ['Ratio', 'Value', 'Interpretation'],
            [
                ['Net Margin', f"{data['margin']:.1f}%", 'Strong' if data['margin'] > 20 else 'Average' if data['margin'] > 10 else 'Weak'],
                ['Return on Equity (Est.)', f"{data['roe']:.1f}%", 'Strong' if data['roe'] > 25 else 'Average' if data['roe'] > 10 else 'Weak'],
                ['Revenue per Employee', f"{cs}{data['annual_revenue']/max(data['total_employees'],1):,.0f}", 'Efficiency benchmark'],
                ['Cost Ratio', f"{(data['annual_cost']/max(data['annual_revenue'],1)*100):.1f}%", 'Lower is better'],
            ],
            [2.2*inch, 2.2*inch, 2.8*inch]
        )
        
        elements.append(Paragraph("Financial Recommendations", h2))
        for rec in ReportService._get_recommendations(data, 'financial'):
            elements.append(Paragraph(f"• {rec}", normal))
            elements.append(Spacer(1, 0.05*inch))

    @staticmethod
    def _pdf_operational(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Operational Efficiency Report", h2))
        elements.append(Paragraph(
            f"Process, supply chain, and capacity metrics for <b>{data['company_name']}</b>.",
            normal
        ))
        
        elements.append(Paragraph("Workforce & Capacity", h2))
        ReportService._pdf_table(elements,
            ['Metric', 'Value', 'Status'],
            [
                ['Total Employees', str(data['total_employees']), 'Active headcount'],
                ['Average Salary', f"{cs}{data['avg_salary']:,.0f}", 'Compensation benchmark'],
                ['Annual Payroll (Est.)', f"{cs}{data['avg_salary'] * data['total_employees']:,.0f}", f"{(data['avg_salary']*data['total_employees']/max(data['annual_revenue'],1)*100):.1f}% of revenue"],
                ['Revenue per Employee', f"{cs}{data['annual_revenue']/max(data['total_employees'],1):,.0f}", 'Productivity indicator'],
            ],
            [2.2*inch, 2.2*inch, 2.8*inch]
        )
        
        elements.append(Paragraph("Inventory & Supply Chain", h2))
        stock_health = "Healthy" if data['low_stock_count'] == 0 else "Attention Needed" if data['low_stock_count'] < 5 else "Critical"
        ReportService._pdf_table(elements,
            ['Metric', 'Value', 'Status'],
            [
                ['Total SKUs', str(data['total_skus']), 'Product breadth'],
                ['Inventory Value', f"{cs}{data['total_inventory_value']:,.0f}", 'Working capital tied'],
                ['Low Stock Items', str(data['low_stock_count']), stock_health],
                ['Inventory/Revenue Ratio', f"{(data['total_inventory_value']/max(data['annual_revenue'],1)*100):.1f}%", 'Capital efficiency'],
            ],
            [2.2*inch, 2.2*inch, 2.8*inch]
        )
        
        elements.append(Paragraph("Operational Recommendations", h2))
        for rec in ReportService._get_recommendations(data, 'operational'):
            elements.append(Paragraph(f"• {rec}", normal))
            elements.append(Spacer(1, 0.05*inch))

    @staticmethod
    def _pdf_department(elements, data, h2, normal, metric):
        cs = _pdf_currency_symbol(data)
        elements.append(Paragraph("Departmental Performance Report", h2))
        elements.append(Paragraph(
            f"Team-level breakdown for <b>{data['company_name']}</b> across all active departments.",
            normal
        ))
        
        elements.append(Paragraph("Department Summary", h2))
        if data['departments']:
            dept_rows = []
            for d in data['departments']:
                variance = d['budget'] - d['payroll']
                variance_pct = (variance / max(d['budget'], 1) * 100)
                dept_rows.append([
                    d['name'],
                    str(d['headcount']),
                    f"{cs}{d['payroll']:,.0f}",
                    f"{cs}{d['budget']:,.0f}",
                    f"{cs}{variance:,.0f}",
                    f"{variance_pct:+.1f}%"
                ])
            ReportService._pdf_table(elements,
                ['Department', 'Headcount', 'Actual Payroll', 'Budget', f'Variance ({cs})', 'Variance (%)'],
                dept_rows,
                [1.5*inch, 1*inch, 1.3*inch, 1.3*inch, 1.2*inch, 1.2*inch]
            )
        else:
            elements.append(Paragraph("No department data available. Configure departments in Settings.", normal))
        
        elements.append(Paragraph("Salary Distribution", h2))
        elements.append(Paragraph(
            f"Company-wide average salary is <b>{cs}{data['avg_salary']:,.0f}</b>. "
            f"Total payroll represents approximately "
            f"<b>{(data['avg_salary']*data['total_employees']/max(data['annual_revenue'],1)*100):.1f}%</b> of annual revenue.",
            normal
        ))
        
        elements.append(Paragraph("Department Recommendations", h2))
        for rec in ReportService._get_recommendations(data, 'department'):
            elements.append(Paragraph(f"• {rec}", normal))
            elements.append(Spacer(1, 0.05*inch))

    @staticmethod
    def _get_recommendations(data, report_type):
        """Generate contextual recommendations based on report type and data."""
        cs = _pdf_currency_symbol(data)
        recs = []
        
        if report_type in ('executive', 'financial'):
            if data['margin'] < 10:
                recs.append("URGENT: Net margin is below 10%. Immediate cost review and pricing optimization required.")
            elif data['margin'] < 20:
                recs.append("Net margin is moderate. Explore vendor renegotiation and automation to improve.")
            else:
                recs.append("Strong margin profile. Consider reinvesting in R&D or market expansion.")
            
            if data['runway_months'] < 6:
                recs.append(f"Cash runway is {data['runway_months']:.1f} months. Secure credit line or reduce discretionary spend.")
            if data['churn_rate'] > 10:
                recs.append(f"Customer churn at {data['churn_rate']:.1f}% is eroding growth. Launch retention program.")
            if data['low_stock_count'] > 3:
                recs.append(f"{data['low_stock_count']} inventory items are low. Adjust reorder points to prevent lost sales.")
        
        if report_type in ('executive', 'operational'):
            if data['total_employees'] < 10:
                recs.append("Team size is lean. Document key processes to reduce single-point-of-failure risk.")
            if data['annual_revenue'] / max(data['total_employees'], 1) < 100000:
                recs.append(f"Revenue per employee is below {cs}100K. Evaluate automation or sales training investments.")
        
        if report_type == 'investor':
            if data['churn_rate'] > 5:
                recs.append("Investor Note: Churn above 5% may compress LTV/CAC multiples. Prioritize retention.")
            if data['margin'] > 25:
                recs.append("Investor Positive: Strong margins support premium valuation multiples.")
            if data['new_customers_30'] < data['total_customers'] * 0.02:
                recs.append("Investor Concern: Customer acquisition velocity is low. Review go-to-market spend.")
        
        if report_type == 'risk':
            if data['critical_risks'] > 0:
                recs.append(f"CRITICAL: {data['critical_risks']} risks require immediate mitigation. Assign owners and deadlines.")
            if data['high_risks'] > 3:
                recs.append(f"High risk volume ({data['high_risks']}) suggests systemic exposure. Conduct enterprise risk audit.")
            if not data['risks']:
                recs.append("No risks registered. This is a blind spot. Complete full risk register immediately.")
        
        if report_type == 'growth':
            if data['new_customers_30'] == 0:
                recs.append("No new customers in 30 days. Activate acquisition channels or review sales funnel.")
            if data['churn_rate'] > data['new_customers_30'] / max(data['total_customers'], 1) * 100:
                recs.append("Churn is outpacing acquisition. Net customer base is contracting. Reverse immediately.")
            if data['total_skus'] < 5 and data['annual_revenue'] > 500000:
                recs.append("Revenue concentration risk: Low SKU count with high revenue. Diversify product line.")
        
        if report_type == 'department':
            for d in data['departments']:
                if d['budget'] > 0:
                    variance = (d['payroll'] - d['budget']) / d['budget'] * 100
                    if variance > 10:
                        recs.append(f"{d['name']}: Over budget by {variance:.1f}%. Freeze non-essential hiring.")
                    elif variance < -15:
                        recs.append(f"{d['name']}: Under budget by {abs(variance):.1f}%. Verify headcount plans are on track.")
            if not data['departments']:
                recs.append("Department structure not defined. Configure departments and budgets for accurate tracking.")
        
        if not recs:
            recs.append("Maintain current trajectory. Schedule next review in 30 days.")
        
        return recs

    # ==================== EXCEL GENERATOR ====================
    
    @staticmethod
    def _generate_excel(report, company_id, report_type, data):
        if not _OPENPYXL_AVAILABLE:
            return ReportService._generate_csv(report, company_id, report_type, data)
        
        config = ReportService.REPORT_TYPE_CONFIG.get(report_type, ReportService.REPORT_TYPE_CONFIG['executive'])
        filename = os.path.join(REPORTS_DIR, f"{report_type}_{report.id}.xlsx")
        
        wb = Workbook()
        
        # Remove default sheet and create type-specific sheets
        wb.remove(wb.active)
        
        # Sheet 1: Overview
        ws = wb.create_sheet("Overview")
        ReportService._excel_overview(ws, data, config)
        
        # Sheet 2: Type-specific data
        if report_type == 'executive':
            ws2 = wb.create_sheet("Financials")
            ReportService._excel_financials(ws2, data)
            ws3 = wb.create_sheet("Health Metrics")
            ReportService._excel_health(ws3, data)
        elif report_type == 'board':
            ws2 = wb.create_sheet("Governance")
            ReportService._excel_governance(ws2, data)
            ws3 = wb.create_sheet("Risk Oversight")
            ReportService._excel_risk_oversight(ws3, data)
        elif report_type == 'investor':
            ws2 = wb.create_sheet("Performance")
            ReportService._excel_investor_performance(ws2, data)
            ws3 = wb.create_sheet("Forecasts")
            ReportService._excel_forecasts(ws3, data)
        elif report_type == 'risk':
            ws2 = wb.create_sheet("Risk Register")
            ReportService._excel_risk_register(ws2, data)
            ws3 = wb.create_sheet("Mitigation")
            ReportService._excel_mitigation(ws3, data)
        elif report_type == 'growth':
            ws2 = wb.create_sheet("Growth KPIs")
            ReportService._excel_growth_kpis(ws2, data)
            ws3 = wb.create_sheet("Customer Analytics")
            ReportService._excel_customers(ws3, data)
        elif report_type == 'forecast':
            ws2 = wb.create_sheet("Forecasts")
            ReportService._excel_forecasts(ws2, data)
            ws3 = wb.create_sheet("Scenarios")
            ReportService._excel_scenarios(ws3, data)
        elif report_type == 'financial':
            ws2 = wb.create_sheet("Income Statement")
            ReportService._excel_income_statement(ws2, data)
            ws3 = wb.create_sheet("Ratios")
            ReportService._excel_ratios(ws3, data)
        elif report_type == 'operational':
            ws2 = wb.create_sheet("Operations")
            ReportService._excel_operations(ws2, data)
            ws3 = wb.create_sheet("Inventory")
            ReportService._excel_inventory(ws3, data)
        elif report_type == 'department':
            ws2 = wb.create_sheet("Departments")
            ReportService._excel_departments(ws2, data)
            ws3 = wb.create_sheet("Payroll")
            ReportService._excel_payroll(ws3, data)
        else:
            ws2 = wb.create_sheet("Details")
            ReportService._excel_financials(ws2, data)
            ws3 = wb.create_sheet("Health")
            ReportService._excel_health(ws3, data)
        
        # Sheet 4: Recommendations (all types)
        ws4 = wb.create_sheet("Recommendations")
        ReportService._excel_recommendations(ws4, data, report_type)
        
        wb.save(filename)
        return filename

    @staticmethod
    def _excel_style_header(cell):
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            bottom=Side(style='thin', color='334155')
        )

    @staticmethod
    def _excel_style_data(cell, is_currency=False, is_percent=False, currency_symbol='$'):
        cell.font = Font(size=10, color='334155')
        cell.alignment = Alignment(horizontal='center' if not is_currency else 'right', vertical='center')
        if is_currency:
            cell.number_format = f'{currency_symbol}#,##0'
        elif is_percent:
            cell.number_format = '0.0%'

    @staticmethod
    def _excel_overview(ws, data, config):
        cs = data.get('currency_symbol', '$')
        ws.append([config['title']])
        ws.append([f"Company: {data['company_name']}"])
        ws.append([f"Industry: {data['industry']}"])
        ws.append([f"Generated: {data['generated_at']}"])
        ws.append([])
        ws.append(['Metric', 'Value'])
        ws.append(['Annual Revenue', data['annual_revenue']])
        ws.append(['Annual Costs', data['annual_cost']])
        ws.append(['Net Profit (LTM)', data['profit_365']])
        ws.append(['Net Margin', data['margin'] / 100])
        ws.append(['Health Score', data['health_score']])
        ws.append(['Employees', data['total_employees']])
        ws.append(['Customers', data['total_customers']])
        ws.append(['Churn Rate', data['churn_rate'] / 100])
        ws.append(['Inventory Value', data['total_inventory_value']])
        
        # Style
        for cell in ws[1]:
            cell.font = Font(bold=True, size=16, color='1e293b')
        for cell in ws[6]:
            ReportService._excel_style_header(cell)
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
            ReportService._excel_style_data(row[0])
            if row[1].value and isinstance(row[1].value, (int, float)):
                if 'Margin' in str(row[0].value) or 'Rate' in str(row[0].value):
                    row[1].number_format = '0.0%'
                    row[1].value = row[1].value / 100 if row[1].value > 1 else row[1].value
                else:
                    row[1].number_format = f'{cs}#,##0'
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    @staticmethod
    def _excel_financials(ws, data):
        cs = data.get('currency_symbol', '$')
        ws.append(['Period', 'Revenue', 'Expenses', 'Profit'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Last 30 Days', data['rev_30'], data['exp_30'], data['profit_30']])
        ws.append(['Last 90 Days', data['rev_90'], data['exp_90'], data['rev_90'] - data['exp_90']])
        ws.append(['Last 12 Months', data['rev_365'], data['exp_365'], data['profit_365']])
        for row in ws.iter_rows(min_row=2):
            for cell in row[1:]:
                cell.number_format = f'{cs}#,##0'
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 18

    @staticmethod
    def _excel_health(ws, data):
        ws.append(['Health Indicator', 'Score', 'Status'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        score = data['health_score']
        status = 'Excellent' if score >= 80 else 'Good' if score >= 60 else 'Fair' if score >= 40 else 'Poor'
        ws.append(['Overall Health', score, status])
        ws.append(['Margin Health', data['margin'], 'Strong' if data['margin'] > 20 else 'Moderate' if data['margin'] > 10 else 'Weak'])
        ws.append(['Liquidity Health', data['runway_months'], 'Strong' if data['runway_months'] > 12 else 'Moderate' if data['runway_months'] > 6 else 'Critical'])
        ws.append(['Retention Health', 100 - data['churn_rate'], 'Strong' if data['churn_rate'] < 5 else 'Moderate' if data['churn_rate'] < 15 else 'Weak'])

    @staticmethod
    def _excel_governance(ws, data):
        ws.append(['Governance Item', 'Value', 'Board Action'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Annual Revenue', data['annual_revenue'], 'Verify against audit'])
        ws.append(['Annual Costs', data['annual_cost'], 'Review cost policies'])
        ws.append(['Net Margin', data['margin'] / 100, 'Benchmark vs peers'])
        ws.append(['Cash Burn', data['burn_rate'], 'Assess runway'])
        ws.append(['Runway (months)', data['runway_months'], 'Approve funding if <6'])

    @staticmethod
    def _excel_risk_oversight(ws, data):
        ws.append(['Risk Level', 'Count', 'Action Required'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Critical', data['critical_risks'], 'Immediate escalation'])
        ws.append(['High', data['high_risks'], 'Mitigate within 30 days'])
        medium = len([r for r in data['risks'] if r.get('risk_level') == 'medium'])
        low = len([r for r in data['risks'] if r.get('risk_level') == 'low'])
        ws.append(['Medium', medium, 'Quarterly review'])
        ws.append(['Low', low, 'Standard controls'])

    @staticmethod
    def _excel_investor_performance(ws, data):
        ws.append(['KPI', 'Value', 'Benchmark'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['LTM Revenue', data['rev_365'], 'YoY growth target'])
        ws.append(['LTM Profit', data['profit_365'], 'Bottom-line target'])
        ws.append(['Net Margin', data['margin'] / 100, 'Industry avg 15-20%'])
        ws.append(['Customer Base', data['total_customers'], 'Retention focus'])
        ws.append(['Churn Rate', data['churn_rate'] / 100, 'Target <5%'])
        ws.append(['New Customers (30d)', data['new_customers_30'], 'Growth velocity'])

    @staticmethod
    def _excel_forecasts(ws, data):
        ws.append(['Forecast Name', 'Type', 'Method', 'Horizon', 'Trend %', 'Final Value'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        for f in data['forecasts']:
            ws.append([f['name'], f['type'], f['method'], f['horizon'], f['trend'] / 100, f['final_value']])

    @staticmethod
    def _excel_risk_register(ws, data):
        ws.append(['Risk Name', 'Category', 'Probability', 'Impact', 'Score', 'Level'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        for r in data['risks'][:20]:
            ws.append([
                r.get('name', 'Unknown'),
                r.get('category', 'N/A').replace('_', ' ').title(),
                r.get('probability', 0),
                r.get('impact', 0),
                r.get('risk_score', 0),
                r.get('risk_level', 'N/A').title()
            ])

    @staticmethod
    def _excel_mitigation(ws, data):
        ws.append(['Priority', 'Risk Count', 'Recommended Action'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Critical', data['critical_risks'], 'Board-level intervention required'])
        ws.append(['High', data['high_risks'], 'Executive ownership within 30 days'])
        ws.append(['Process', 'All', 'Implement quarterly risk review cycle'])

    @staticmethod
    def _excel_growth_kpis(ws, data):
        cac = 150 if data['new_customers_30'] > 0 else 0
        ltv = (data['annual_revenue'] / max(data['total_customers'], 1)) * 2.5
        ws.append(['Growth Metric', 'Value', 'Target'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['New Customers (30d)', data['new_customers_30'], '>5% of base monthly'])
        ws.append(['Total Customers', data['total_customers'], 'Retention-focused'])
        ws.append(['Churn Rate', data['churn_rate'] / 100, '<5% annually'])
        ws.append(['Est. CAC', cac, '<1/3 of LTV'])
        ws.append(['Est. LTV', ltv, 'Maximize'])
        ws.append(['LTV/CAC', ltv / max(cac, 1), '>3.0x'])
        ws.append(['Revenue/Employee', data['annual_revenue'] / max(data['total_employees'], 1), '>100K'])

    @staticmethod
    def _excel_customers(ws, data):
        ws.append(['Customer Metric', 'Value'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Total Active', data['total_customers']])
        ws.append(['Churned', int(data['total_customers'] * data['churn_rate'] / max(100 - data['churn_rate'], 1))])
        ws.append(['New (30d)', data['new_customers_30']])
        ws.append(['Revenue per Customer', data['annual_revenue'] / max(data['total_customers'], 1)])

    @staticmethod
    def _excel_income_statement(ws, data):
        ws.append(['Line Item', 'Amount', '% of Revenue'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Revenue', data['rev_365'], 1.0])
        ws.append(['Cost of Revenue (Est.)', data['exp_365'] * 0.6, (data['exp_365'] * 0.6) / max(data['rev_365'], 1)])
        ws.append(['Gross Profit (Est.)', data['profit_365'] + (data['exp_365'] * 0.4), (data['profit_365'] + data['exp_365'] * 0.4) / max(data['rev_365'], 1)])
        ws.append(['Operating Expenses', data['exp_365'], data['exp_365'] / max(data['rev_365'], 1)])
        ws.append(['Net Profit', data['profit_365'], data['profit_365'] / max(data['rev_365'], 1)])

    @staticmethod
    def _excel_ratios(ws, data):
        ws.append(['Ratio', 'Value', 'Assessment'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Net Margin', data['margin'] / 100, 'Strong' if data['margin'] > 20 else 'Average' if data['margin'] > 10 else 'Weak'])
        ws.append(['Return on Equity (Est.)', data['roe'] / 100, 'Strong' if data['roe'] > 25 else 'Average' if data['roe'] > 10 else 'Weak'])
        ws.append(['Revenue per Employee', data['annual_revenue'] / max(data['total_employees'], 1), 'Efficiency'])
        ws.append(['Cost Ratio', data['annual_cost'] / max(data['annual_revenue'], 1), 'Lower is better'])

    @staticmethod
    def _excel_operations(ws, data):
        ws.append(['Operational Metric', 'Value', 'Note'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Total Employees', data['total_employees'], 'Active headcount'])
        ws.append(['Avg Salary', data['avg_salary'], 'Compensation benchmark'])
        ws.append(['Annual Payroll (Est.)', data['avg_salary'] * data['total_employees'], 'Fixed cost'])
        ws.append(['Revenue per Employee', data['annual_revenue'] / max(data['total_employees'], 1), 'Productivity'])

    @staticmethod
    def _excel_inventory(ws, data):
        ws.append(['Inventory Metric', 'Value', 'Status'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        status = 'Healthy' if data['low_stock_count'] == 0 else 'Attention' if data['low_stock_count'] < 5 else 'Critical'
        ws.append(['Total SKUs', data['total_skus'], 'Product breadth'])
        ws.append(['Inventory Value', data['total_inventory_value'], 'Working capital'])
        ws.append(['Low Stock Items', data['low_stock_count'], status])
        ws.append(['Inventory/Revenue', data['total_inventory_value'] / max(data['annual_revenue'], 1), 'Capital efficiency'])

    @staticmethod
    def _excel_departments(ws, data):
        ws.append(['Department', 'Headcount', 'Actual Payroll', 'Budget', 'Variance', 'Variance (%)'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        for d in data['departments']:
            variance = d['budget'] - d['payroll']
            variance_pct = (variance / max(d['budget'], 1) * 100)
            ws.append([d['name'], d['headcount'], d['payroll'], d['budget'], variance, variance_pct / 100])

    @staticmethod
    def _excel_payroll(ws, data):
        ws.append(['Payroll Metric', 'Value'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        ws.append(['Total Employees', data['total_employees']])
        ws.append(['Average Salary', data['avg_salary']])
        ws.append(['Total Payroll', data['avg_salary'] * data['total_employees']])
        ws.append(['Payroll/Revenue %', (data['avg_salary'] * data['total_employees']) / max(data['annual_revenue'], 1)])

    @staticmethod
    def _excel_recommendations(ws, data, report_type):
        ws.append(['Priority', 'Recommendation'])
        for cell in ws[1]:
            ReportService._excel_style_header(cell)
        recs = ReportService._get_recommendations(data, report_type)
        for i, rec in enumerate(recs, 1):
            ws.append([f"{i}", rec])

    # ==================== CSV GENERATOR ====================
    
    @staticmethod
    def _generate_csv(report, company_id, report_type, data):
        config = ReportService.REPORT_TYPE_CONFIG.get(report_type, ReportService.REPORT_TYPE_CONFIG['executive'])
        filename = os.path.join(EXPORTS_DIR, f"{report_type}_{report.id}.csv")
        
        rows = []
        rows.append(['REPORT_TYPE', config['title']])
        rows.append(['COMPANY', data['company_name']])
        rows.append(['INDUSTRY', data['industry']])
        rows.append(['GENERATED', data['generated_at']])
        rows.append([])
        
        if report_type == 'executive':
            rows.append(['SECTION', 'Executive Summary'])
            rows.append(['METRIC', 'VALUE'])
            rows.append(['Annual Revenue', data['annual_revenue']])
            rows.append(['Annual Costs', data['annual_cost']])
            rows.append(['Net Profit (LTM)', data['profit_365']])
            rows.append(['Net Margin %', data['margin']])
            rows.append(['Health Score', data['health_score']])
            rows.append(['Employees', data['total_employees']])
            rows.append(['Customers', data['total_customers']])
            rows.append(['Churn Rate %', data['churn_rate']])
            rows.append([])
            rows.append(['SECTION', '30-Day Financial Pulse'])
            rows.append(['PERIOD', 'REVENUE', 'EXPENSES', 'PROFIT'])
            rows.append(['Last 30 Days', data['rev_30'], data['exp_30'], data['profit_30']])
            rows.append(['Last 90 Days', data['rev_90'], data['exp_90'], data['rev_90'] - data['exp_90']])
            rows.append(['Last 12 Months', data['rev_365'], data['exp_365'], data['profit_365']])
            rows.append([])
            rows.append(['SECTION', 'Recommendations'])
            for rec in ReportService._get_recommendations(data, 'executive'):
                rows.append(['RECOMMENDATION', rec])
        
        elif report_type == 'board':
            rows.append(['SECTION', 'Governance & Financial Highlights'])
            rows.append(['METRIC', 'VALUE', 'BOARD_NOTE'])
            rows.append(['Annual Revenue', data['annual_revenue'], 'Verify against audit'])
            rows.append(['Annual Costs', data['annual_cost'], 'Review cost policies'])
            rows.append(['Net Margin %', data['margin'], 'Benchmark vs peers'])
            rows.append(['Cash Burn (30d)', data['burn_rate'], 'Assess runway'])
            rows.append(['Est Runway (months)', data['runway_months'], 'Approve funding if <6'])
            rows.append([])
            rows.append(['SECTION', 'Risk Oversight'])
            rows.append(['SEVERITY', 'COUNT'])
            rows.append(['Critical', data['critical_risks']])
            rows.append(['High', data['high_risks']])
            rows.append(['Medium', len([r for r in data['risks'] if r.get('risk_level') == 'medium'])])
            rows.append(['Low', len([r for r in data['risks'] if r.get('risk_level') == 'low'])])
        
        elif report_type == 'investor':
            rows.append(['SECTION', 'Investor Performance Metrics'])
            rows.append(['KPI', 'VALUE', 'RELEVANCE'])
            rows.append(['LTM Revenue', data['rev_365'], 'Top-line growth'])
            rows.append(['LTM Profit', data['profit_365'], 'Bottom-line'])
            rows.append(['Net Margin %', data['margin'], 'Unit economics'])
            rows.append(['Customer Base', data['total_customers'], 'Market penetration'])
            rows.append(['Churn Rate %', data['churn_rate'], 'Retention health'])
            rows.append(['New Customers (30d)', data['new_customers_30'], 'Growth velocity'])
            rows.append(['Return on Equity %', data['roe'], 'Capital efficiency'])
            rows.append([])
            rows.append(['SECTION', 'Forecasts'])
            rows.append(['NAME', 'TYPE', 'METHOD', 'HORIZON', 'TREND_%', 'FINAL_VALUE'])
            for f in data['forecasts']:
                rows.append([f['name'], f['type'], f['method'], f['horizon'], f['trend'], f['final_value']])
        
        elif report_type == 'risk':
            rows.append(['SECTION', 'Enterprise Risk Register'])
            rows.append(['RISK_NAME', 'CATEGORY', 'PROBABILITY', 'IMPACT', 'SCORE', 'LEVEL'])
            for r in data['risks'][:50]:
                rows.append([
                    r.get('name', 'Unknown'),
                    r.get('category', 'N/A'),
                    r.get('probability', 0),
                    r.get('impact', 0),
                    r.get('risk_score', 0),
                    r.get('risk_level', 'N/A')
                ])
            rows.append([])
            rows.append(['SECTION', 'Risk Distribution'])
            rows.append(['LEVEL', 'COUNT'])
            rows.append(['Critical', data['critical_risks']])
            rows.append(['High', data['high_risks']])
            rows.append(['Medium', len([r for r in data['risks'] if r.get('risk_level') == 'medium'])])
            rows.append(['Low', len([r for r in data['risks'] if r.get('risk_level') == 'low'])])
        
        elif report_type == 'growth':
            cac = 150 if data['new_customers_30'] > 0 else 0
            ltv = (data['annual_revenue'] / max(data['total_customers'], 1)) * 2.5
            rows.append(['SECTION', 'Growth KPIs'])
            rows.append(['METRIC', 'VALUE', 'BENCHMARK'])
            rows.append(['New Customers (30d)', data['new_customers_30'], '>5% of base monthly'])
            rows.append(['Total Customers', data['total_customers'], 'Retention focus'])
            rows.append(['Churn Rate %', data['churn_rate'], '<5% annually'])
            rows.append(['Est CAC', cac, '<1/3 of LTV'])
            rows.append(['Est LTV', ltv, 'Maximize'])
            rows.append(['LTV/CAC Ratio', ltv / max(cac, 1), '>3.0x'])
            rows.append(['Revenue per Employee', data['annual_revenue'] / max(data['total_employees'], 1), '>100K'])
        
        elif report_type == 'forecast':
            rows.append(['SECTION', 'Active Forecasts'])
            rows.append(['NAME', 'TYPE', 'METHOD', 'HORIZON', 'TREND_%', 'FINAL_VALUE'])
            for f in data['forecasts']:
                rows.append([f['name'], f['type'], f['method'], f['horizon'], f['trend'], f['final_value']])
            rows.append([])
            rows.append(['SECTION', 'Recent Simulations'])
            rows.append(['NAME', 'TYPE', 'REV_IMPACT', 'PROFIT_IMPACT', 'RISK_SCORE'])
            for s in data['simulations'][:10]:
                rev_impact = (s['revenue_after'] or 0) - (s['revenue_before'] or 0)
                profit_impact = (s['profit_after'] or 0) - (s['profit_before'] or 0)
                rows.append([s['name'], s['type'], rev_impact, profit_impact, s['risk_score']])
        
        elif report_type == 'financial':
            rows.append(['SECTION', 'Income Statement (Trailing 12 Months)'])
            rows.append(['LINE_ITEM', 'AMOUNT', 'PCT_OF_REVENUE'])
            rows.append(['Revenue', data['rev_365'], 1.0])
            rows.append(['Cost of Revenue (Est.)', data['exp_365'] * 0.6, (data['exp_365'] * 0.6) / max(data['rev_365'], 1)])
            rows.append(['Gross Profit (Est.)', data['profit_365'] + (data['exp_365'] * 0.4), (data['profit_365'] + data['exp_365'] * 0.4) / max(data['rev_365'], 1)])
            rows.append(['Operating Expenses', data['exp_365'], data['exp_365'] / max(data['rev_365'], 1)])
            rows.append(['Net Profit', data['profit_365'], data['profit_365'] / max(data['rev_365'], 1)])
            rows.append([])
            rows.append(['SECTION', 'Financial Ratios'])
            rows.append(['RATIO', 'VALUE', 'ASSESSMENT'])
            rows.append(['Net Margin %', data['margin'], 'Strong' if data['margin'] > 20 else 'Average' if data['margin'] > 10 else 'Weak'])
            rows.append(['Return on Equity %', data['roe'], 'Strong' if data['roe'] > 25 else 'Average' if data['roe'] > 10 else 'Weak'])
            rows.append(['Revenue per Employee', data['annual_revenue'] / max(data['total_employees'], 1), 'Efficiency'])
            rows.append(['Cost Ratio %', (data['annual_cost'] / max(data['annual_revenue'], 1)) * 100, 'Lower is better'])
        
        elif report_type == 'operational':
            rows.append(['SECTION', 'Operational Metrics'])
            rows.append(['METRIC', 'VALUE', 'NOTE'])
            rows.append(['Total Employees', data['total_employees'], 'Active headcount'])
            rows.append(['Average Salary', data['avg_salary'], 'Compensation benchmark'])
            rows.append(['Annual Payroll (Est.)', data['avg_salary'] * data['total_employees'], 'Fixed cost'])
            rows.append(['Revenue per Employee', data['annual_revenue'] / max(data['total_employees'], 1), 'Productivity'])
            rows.append([])
            rows.append(['SECTION', 'Inventory'])
            rows.append(['METRIC', 'VALUE', 'STATUS'])
            status = 'Healthy' if data['low_stock_count'] == 0 else 'Attention' if data['low_stock_count'] < 5 else 'Critical'
            rows.append(['Total SKUs', data['total_skus'], 'Product breadth'])
            rows.append(['Inventory Value', data['total_inventory_value'], 'Working capital'])
            rows.append(['Low Stock Items', data['low_stock_count'], status])
            rows.append(['Inventory/Revenue %', (data['total_inventory_value'] / max(data['annual_revenue'], 1)) * 100, 'Capital efficiency'])
        
        elif report_type == 'department':
            rows.append(['SECTION', 'Departmental Breakdown'])
            rows.append(['DEPARTMENT', 'HEADCOUNT', 'ACTUAL_PAYROLL', 'BUDGET', 'VARIANCE_$', 'VARIANCE_%'])
            for d in data['departments']:
                variance = d['budget'] - d['payroll']
                variance_pct = (variance / max(d['budget'], 1) * 100)
                rows.append([d['name'], d['headcount'], d['payroll'], d['budget'], variance, variance_pct])
            rows.append([])
            rows.append(['SECTION', 'Company Payroll Summary'])
            rows.append(['METRIC', 'VALUE'])
            rows.append(['Total Employees', data['total_employees']])
            rows.append(['Average Salary', data['avg_salary']])
            rows.append(['Total Payroll', data['avg_salary'] * data['total_employees']])
            rows.append(['Payroll/Revenue %', (data['avg_salary'] * data['total_employees']) / max(data['annual_revenue'], 1) * 100])
        
        else:
            rows.append(['SECTION', 'General Report'])
            rows.append(['METRIC', 'VALUE'])
            rows.append(['Revenue', data['annual_revenue']])
            rows.append(['Costs', data['annual_cost']])
            rows.append(['Profit', data['profit_365']])
            rows.append(['Margin %', data['margin']])
        
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            for row in rows:
                writer.writerow(row)
        
        return filename