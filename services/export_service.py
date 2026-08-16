import os
import csv
from datetime import datetime
from database.db import db
from database.models.export_history import ExportHistory
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ==================== UNICODE FONT REGISTRATION ====================
_PDF_FONT = 'Helvetica'
_PDF_FONT_BOLD = 'Helvetica-Bold'

def _register_unicode_fonts():
    global _PDF_FONT, _PDF_FONT_BOLD
    candidates = [
        ('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
         'DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
        ('LiberationSans', '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
         'LiberationSans-Bold', '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf'),
        ('NotoSans', '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
         'NotoSans-Bold', '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf'),
        ('ArialUnicode', 'C:/Windows/Fonts/arial.ttf',
         'ArialUnicode-Bold', 'C:/Windows/Fonts/arialbd.ttf'),
        ('ArialUnicode', '/Library/Fonts/Arial.ttf',
         'ArialUnicode-Bold', '/Library/Fonts/Arial Bold.ttf'),
        ('ArialUnicode', '/System/Library/Fonts/Supplemental/Arial.ttf',
         'ArialUnicode-Bold', '/System/Library/Fonts/Supplemental/Arial Bold.ttf'),
    ]
    for reg_name, reg_path, bold_name, bold_path in candidates:
        try:
            pdfmetrics.registerFont(TTFont(reg_name, reg_path))
            pdfmetrics.registerFont(TTFont(bold_name, bold_path))
            _PDF_FONT = reg_name
            _PDF_FONT_BOLD = bold_name
            return
        except Exception:
            continue

_register_unicode_fonts()

# Absolute base directories for generated files
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
EXPORTS_DIR = os.path.join(BASE_DIR, 'documents', 'exports')
REPORTS_DIR = os.path.join(BASE_DIR, 'documents', 'reports')

class ExportService:
    @staticmethod
    def export_data(export_type, data, filename=None, format_type='csv', currency_symbol='$'):
        """Export generic data to CSV, PDF, or Excel format."""
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        
        if not filename:
            filename = f"{export_type}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}"
        
        if format_type == 'csv':
            filepath = os.path.join(EXPORTS_DIR, f"{filename}.csv")
            ExportService._export_csv(data, filepath)
        elif format_type == 'pdf':
            filepath = os.path.join(EXPORTS_DIR, f"{filename}.pdf")
            ExportService._export_pdf(data, filepath, export_type)
        elif format_type == 'excel':
            filepath = os.path.join(EXPORTS_DIR, f"{filename}.xlsx")
            ExportService._export_excel(data, filepath)
        else:
            filepath = os.path.join(EXPORTS_DIR, f"{filename}.csv")
            ExportService._export_csv(data, filepath)
        
        # Log export
        export_log = ExportHistory(
            export_type=export_type,
            file_path=filepath,
            file_format=format_type,
            record_count=len(data) if isinstance(data, list) else 0,
            created_at=datetime.utcnow()
        )
        db.session.add(export_log)
        db.session.commit()
        
        return filepath
    
    @staticmethod
    def _export_csv(data, filepath):
        if not data:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['No data available'])
            return
        
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if isinstance(data, list) and len(data) > 0:
                if isinstance(data[0], dict):
                    writer.writerow(data[0].keys())
                    for row in data:
                        writer.writerow(row.values())
                else:
                    writer.writerow(data)
            else:
                writer.writerow(['Data'])
                writer.writerow([str(data)])
    
    @staticmethod
    def _export_pdf(data, filepath, title):
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1a1f3a'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName=_PDF_FONT_BOLD
        )
        
        elements = []
        elements.append(Paragraph(f"{title.replace('_', ' ').title()} Export", title_style))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%B %d, %Y')}", styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            headers = list(data[0].keys())
            table_data = [headers]
            for row in data[:50]:  # Limit to 50 rows for PDF
                table_data.append([str(row.get(h, '')) for h in headers])
            
            col_width = 6 * inch / len(headers)
            t = Table(table_data, colWidths=[col_width] * len(headers))
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1f3a')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), _PDF_FONT_BOLD),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('FONTNAME', (0, 1), (-1, -1), _PDF_FONT),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph(str(data), styles['Normal']))
        
        doc.build(elements)
    
    @staticmethod
    def _export_excel(data, filepath):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Fallback to CSV
            csv_path = filepath.replace('.xlsx', '.csv')
            ExportService._export_csv(data, csv_path)
            return csv_path
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"
        
        if isinstance(data, list) and len(data) > 0 and isinstance(data[0], dict):
            headers = list(data[0].keys())
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1a1f3a', end_color='1a1f3a', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            for row in data:
                ws.append([row.get(h, '') for h in headers])
        else:
            ws.append(['Data'])
            ws.append([str(data)])
        
        wb.save(filepath)
        return filepath